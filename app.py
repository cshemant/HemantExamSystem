import os, csv, io, random, socket, secrets, sys, json, math, sqlite3, tempfile, shutil, base64, time, hashlib, hmac, re, difflib
from pathlib import Path
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from functools import wraps
from urllib.parse import urlparse

from flask import Flask, render_template, request, redirect, url_for, session as web_session, flash, jsonify, abort, send_file, after_this_request
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from sqlalchemy import create_engine, String, Integer, Boolean, Float, ForeignKey, UniqueConstraint, Text, select, func, or_, delete, inspect, text, event
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, scoped_session, sessionmaker
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import qrcode
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from enterprise_core import QUESTION_TYPE_LABELS, canonical_question_type, normalize_answer, normalized_key, is_answer_correct, validate_question_definition
from security_core import generate_totp_secret, verify_totp, totp_uri
from edge_package import seal_envelope, open_sealed_envelope
from audit_core import audit_event_hash, verify_audit_rows
from practical_core import parse_roster_bytes, parse_experiment_bytes, parse_experiment_text, normalize_experiment_sequence, normalize_experiment_code

RESOURCE_DIR=Path(getattr(sys, '_MEIPASS', Path(__file__).resolve().parent))
DATA_DIR=Path(os.getenv('EXAM_DATA_DIR', str(RESOURCE_DIR))).expanduser().resolve()
DATA_DIR.mkdir(parents=True,exist_ok=True)
load_dotenv(RESOURCE_DIR/'.env')

APP_VERSION='2.19.0'
OFFLINE_RELEASE_FILENAME='LearnWithHemant_Offline_Exam_V2.02_Windows.zip'
DEFAULT_OFFLINE_DOWNLOAD_URL=(
    'https://github.com/cshemant/HemantExamSystem/releases/download/v2.02/'
    + OFFLINE_RELEASE_FILENAME
)
OFFLINE_DOWNLOAD_URL=os.getenv('OFFLINE_DOWNLOAD_URL',DEFAULT_OFFLINE_DOWNLOAD_URL).strip() or DEFAULT_OFFLINE_DOWNLOAD_URL
OFFLINE_REQUIRE_SETUP=os.getenv('OFFLINE_REQUIRE_SETUP','0').strip().lower() in {'1','true','yes','on'}
SESSION_TIMEOUT_MINUTES=max(10,int(os.getenv('SESSION_TIMEOUT_MINUTES','45')))
BACKUP_KDF_ITERATIONS=max(100000,int(os.getenv('BACKUP_KDF_ITERATIONS','390000')))
LOGIN_MAX_FAILURES=max(3,int(os.getenv('LOGIN_MAX_FAILURES','5')))
LOGIN_WINDOW_MINUTES=max(1,int(os.getenv('LOGIN_WINDOW_MINUTES','10')))
LOGIN_LOCK_MINUTES=max(1,int(os.getenv('LOGIN_LOCK_MINUTES','10')))
HEARTBEAT_STALE_SECONDS=max(20,int(os.getenv('HEARTBEAT_STALE_SECONDS','45')))
MAX_ANSWER_LENGTH=max(100,int(os.getenv('MAX_ANSWER_LENGTH','4000')))
PRACTICAL_SYNC_BATCH_SIZE=max(5,min(50,int(os.getenv('PRACTICAL_SYNC_BATCH_SIZE','20'))))
INTEGRATION_API_KEY=os.getenv('INTEGRATION_API_KEY','').strip()
EXAM_PACKAGE_SIGNING_KEY=os.getenv('EXAM_PACKAGE_SIGNING_KEY','').strip()

# All user-facing dates/times use an explicit timezone so cloud hosts such as
# Render (which commonly run in UTC) and offline Windows builds show the same time.
APP_TIMEZONE=os.getenv('APP_TIMEZONE','Asia/Kolkata').strip() or 'Asia/Kolkata'
try:
    DISPLAY_TZ=ZoneInfo(APP_TIMEZONE)
except Exception:
    # Safe fallback for Indian deployments if the configured zone is unavailable.
    DISPLAY_TZ=timezone(timedelta(hours=5,minutes=30))

APP_MODE=os.getenv('APP_MODE','offline').strip().lower()
if APP_MODE not in {'offline','online'}: raise RuntimeError('APP_MODE must be offline or online')

def normalize_database_url(raw):
    if not raw:
        return f"sqlite:///{(DATA_DIR/'exam.db').as_posix()}"
    raw=raw.strip()
    if raw.startswith('postgres://'):
        raw='postgresql+psycopg://'+raw[len('postgres://'):]
    elif raw.startswith('postgresql://'):
        raw='postgresql+psycopg://'+raw[len('postgresql://'):]
    return raw

DATABASE_URL=normalize_database_url(os.getenv('DATABASE_URL'))
if APP_MODE=='online' and not DATABASE_URL.startswith('postgresql'):
    raise RuntimeError('Online mode requires a PostgreSQL DATABASE_URL.')

secret=os.getenv('SECRET_KEY','').strip()
# SUPER_ADMIN_USERNAME is the preferred setting. ADMIN_USERNAME remains supported
# for backwards compatibility with existing deployments.
super_admin_username=(os.getenv('SUPER_ADMIN_USERNAME') or os.getenv('ADMIN_USERNAME') or 'admin').strip() or 'admin'
legacy_admin_username=os.getenv('LEGACY_ADMIN_USERNAME','admin').strip() or 'admin'
# Existing deployments already use ADMIN_PASSWORD, so keep it as the fallback.
# SUPER_ADMIN_PASSWORD may be set later if a separate secret is preferred.
admin_password=(os.getenv('SUPER_ADMIN_PASSWORD') or os.getenv('ADMIN_PASSWORD') or '').strip()
# Compatibility alias used by older helper code.
admin_username=super_admin_username
if APP_MODE=='online':
    if len(secret)<24: raise RuntimeError('Online mode requires a strong SECRET_KEY (24+ characters).')
    if len(admin_password)<10: raise RuntimeError('Online mode requires SUPER_ADMIN_PASSWORD or ADMIN_PASSWORD with at least 10 characters.')
if not secret:
    secret='offline-development-secret-change-me'
if not admin_password and not OFFLINE_REQUIRE_SETUP:
    admin_password='Admin@123'

app=Flask(__name__,template_folder=str(RESOURCE_DIR/'templates'),static_folder=str(RESOURCE_DIR/'static'))
app.secret_key=secret
cookie_secure=os.getenv('COOKIE_SECURE','1' if APP_MODE=='online' else '0').strip().lower() in {'1','true','yes','on'}
app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE='Lax',SESSION_COOKIE_SECURE=cookie_secure,MAX_CONTENT_LENGTH=10*1024*1024)
if APP_MODE=='online': app.wsgi_app=ProxyFix(app.wsgi_app,x_for=1,x_proto=1,x_host=1,x_port=1)

engine=create_engine(DATABASE_URL,pool_pre_ping=True,connect_args={'check_same_thread':False,'timeout':10} if DATABASE_URL.startswith('sqlite') else {})
if DATABASE_URL.startswith('sqlite'):
    @event.listens_for(engine,'connect')
    def _sqlite_connection_pragmas(dbapi_connection,_connection_record):
        cursor=dbapi_connection.cursor()
        try:
            cursor.execute('PRAGMA foreign_keys=ON');cursor.execute('PRAGMA busy_timeout=5000')
        finally:
            cursor.close()
DB=scoped_session(sessionmaker(bind=engine,autoflush=False,expire_on_commit=False))

class Base(DeclarativeBase): pass

class Admin(Base):
    __tablename__='admins'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    username:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    password_hash:Mapped[str]=mapped_column(String,nullable=False)
    mfa_secret:Mapped[str]=mapped_column(String,nullable=False,default='')
    mfa_enabled:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)

class Faculty(Base):
    __tablename__='faculty_users'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    username:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    name:Mapped[str]=mapped_column(String,nullable=False)
    password_hash:Mapped[str]=mapped_column(String,nullable=False)
    is_active:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    created_at:Mapped[str]=mapped_column(String,nullable=False)
    mfa_secret:Mapped[str]=mapped_column(String,nullable=False,default='')
    mfa_enabled:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)

class Student(Base):
    __tablename__='students'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    # roll_no is the student's login ID.  Practical-register sync uses the
    # last five digits of the university registration number here.
    roll_no:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    registration_no:Mapped[str]=mapped_column(String,nullable=False,default='')
    name:Mapped[str]=mapped_column(String,nullable=False)
    password_hash:Mapped[str]=mapped_column(String,nullable=False)
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class Exam(Base):
    __tablename__='exams'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    title:Mapped[str]=mapped_column(String,nullable=False)
    duration_minutes:Mapped[int]=mapped_column(Integer,nullable=False)
    is_active:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class Question(Base):
    __tablename__='questions'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    question:Mapped[str]=mapped_column(String,nullable=False)
    option_a:Mapped[str]=mapped_column(String,nullable=False)
    option_b:Mapped[str]=mapped_column(String,nullable=False)
    option_c:Mapped[str]=mapped_column(String,nullable=False)
    option_d:Mapped[str]=mapped_column(String,nullable=False)
    correct_answer:Mapped[str]=mapped_column(String(1),nullable=False)
    question_type:Mapped[str]=mapped_column(String,nullable=False,default='single_choice')
    answer_key:Mapped[str]=mapped_column(Text,nullable=False,default='')
    answer_tolerance:Mapped[str]=mapped_column(String,nullable=False,default='')
    answer_case_sensitive:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    marks:Mapped[int]=mapped_column(Integer,nullable=False,default=1)
    # Snapshot of practical-exam mapping at the time the bank question is copied.
    # Blank means this is a normal/official exam question.
    practical_experiment_no:Mapped[str]=mapped_column(String,nullable=False,default='')

class Attempt(Base):
    __tablename__='attempts'
    __table_args__=(UniqueConstraint('student_id','exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    started_at:Mapped[str]=mapped_column(String,nullable=False)
    end_at:Mapped[str]=mapped_column(String,nullable=False)
    submitted_at:Mapped[str|None]=mapped_column(String,nullable=True)
    status:Mapped[str]=mapped_column(String,nullable=False,default='in_progress')
    score:Mapped[int|None]=mapped_column(Integer,nullable=True)
    total_marks:Mapped[int|None]=mapped_column(Integer,nullable=True)
    grading_status:Mapped[str]=mapped_column(String,nullable=False,default='complete')
    question_order:Mapped[str]=mapped_column(String,nullable=False)

class Answer(Base):
    __tablename__='answers'
    __table_args__=(UniqueConstraint('attempt_id','question_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    attempt_id:Mapped[int]=mapped_column(ForeignKey('attempts.id'),nullable=False)
    question_id:Mapped[int]=mapped_column(ForeignKey('questions.id'),nullable=False)
    selected_answer:Mapped[str|None]=mapped_column(String(1),nullable=True)
    answer_value:Mapped[str]=mapped_column(Text,nullable=False,default='')
    manual_score:Mapped[int|None]=mapped_column(Integer,nullable=True)
    grader_comment:Mapped[str]=mapped_column(Text,nullable=False,default='')
    graded_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    graded_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    saved_at:Mapped[str]=mapped_column(String,nullable=False)

class BankQuestion(Base):
    __tablename__='bank_questions'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    subject:Mapped[str]=mapped_column(String,nullable=False,default='General')
    course_semester:Mapped[str]=mapped_column(String,nullable=False,default='')
    unit:Mapped[str]=mapped_column(String,nullable=False,default='')
    topic:Mapped[str]=mapped_column(String,nullable=False,default='')
    question_type:Mapped[str]=mapped_column(String,nullable=False,default='MCQ')
    question:Mapped[str]=mapped_column(String,nullable=False)
    option_a:Mapped[str]=mapped_column(String,nullable=False)
    option_b:Mapped[str]=mapped_column(String,nullable=False)
    option_c:Mapped[str]=mapped_column(String,nullable=False)
    option_d:Mapped[str]=mapped_column(String,nullable=False)
    correct_answer:Mapped[str]=mapped_column(String(1),nullable=False)
    answer_key:Mapped[str]=mapped_column(Text,nullable=False,default='')
    answer_tolerance:Mapped[str]=mapped_column(String,nullable=False,default='')
    answer_case_sensitive:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    marks:Mapped[int]=mapped_column(Integer,nullable=False,default=1)
    difficulty:Mapped[str]=mapped_column(String,nullable=False,default='Medium')
    bloom_level:Mapped[str]=mapped_column(String,nullable=False,default='Understand')
    co_mapping:Mapped[str]=mapped_column(String,nullable=False,default='')
    po_mapping:Mapped[str]=mapped_column(String,nullable=False,default='')
    pso_mapping:Mapped[str]=mapped_column(String,nullable=False,default='')
    tags:Mapped[str]=mapped_column(String,nullable=False,default='')
    practice_visibility:Mapped[str]=mapped_column(String,nullable=False,default='official_only')
    practical_experiment_no:Mapped[str]=mapped_column(String,nullable=False,default='')
    explanation:Mapped[str]=mapped_column(Text,nullable=False,default='')
    status:Mapped[str]=mapped_column(String,nullable=False,default='draft')
    version:Mapped[int]=mapped_column(Integer,nullable=False,default=1)
    created_by:Mapped[str]=mapped_column(String,nullable=False,default='admin')
    created_at:Mapped[str]=mapped_column(String,nullable=False)
    updated_at:Mapped[str]=mapped_column(String,nullable=False)

class SubjectCatalog(Base):
    __tablename__='subject_catalog'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    name:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    category:Mapped[str]=mapped_column(String,nullable=False,default='Custom / Other')
    course_semester:Mapped[str]=mapped_column(String,nullable=False,default='')
    is_active:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    created_by:Mapped[str]=mapped_column(String,nullable=False,default='system')
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class QuestionRevision(Base):
    __tablename__='question_revisions'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    bank_question_id:Mapped[int]=mapped_column(ForeignKey('bank_questions.id'),nullable=False)
    version:Mapped[int]=mapped_column(Integer,nullable=False)
    snapshot_json:Mapped[str]=mapped_column(String,nullable=False)
    changed_by:Mapped[str]=mapped_column(String,nullable=False)
    changed_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamBankMap(Base):
    __tablename__='exam_bank_map'
    __table_args__=(UniqueConstraint('exam_id','bank_question_id'),UniqueConstraint('exam_question_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    exam_question_id:Mapped[int]=mapped_column(ForeignKey('questions.id'),nullable=False)
    bank_question_id:Mapped[int]=mapped_column(ForeignKey('bank_questions.id'),nullable=False)

class ExamConfig(Base):
    __tablename__='exam_configs'
    __table_args__=(UniqueConstraint('exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    subject:Mapped[str]=mapped_column(String,nullable=False,default='')
    course_semester:Mapped[str]=mapped_column(String,nullable=False,default='')
    question_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    pool_size:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    easy_pct:Mapped[int]=mapped_column(Integer,nullable=False,default=30)
    medium_pct:Mapped[int]=mapped_column(Integer,nullable=False,default=50)
    hard_pct:Mapped[int]=mapped_column(Integer,nullable=False,default=20)
    unit_weights:Mapped[str]=mapped_column(String,nullable=False,default='')
    randomize_questions:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    shuffle_options:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    require_fullscreen:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    tab_switch_limit:Mapped[int]=mapped_column(Integer,nullable=False,default=3)
    # Exam-level classification.  This lets an already-created exam be converted
    # to a Practical Exam without recreating its question pool.
    exam_type:Mapped[str]=mapped_column(String,nullable=False,default='regular')
    practical_experiment_no:Mapped[str]=mapped_column(String,nullable=False,default='')
    # Separate practical-code editing window. Values are stored as local
    # institutional datetimes (YYYY-MM-DDTHH:MM:SS) in APP_TIMEZONE.
    practical_code_start_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    practical_code_end_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    last_generation_summary:Mapped[str]=mapped_column(String,nullable=False,default='')
    updated_at:Mapped[str]=mapped_column(String,nullable=False)

class AttemptQuestion(Base):
    __tablename__='attempt_questions'
    __table_args__=(UniqueConstraint('attempt_id','question_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    attempt_id:Mapped[int]=mapped_column(ForeignKey('attempts.id'),nullable=False)
    question_id:Mapped[int]=mapped_column(ForeignKey('questions.id'),nullable=False)
    position:Mapped[int]=mapped_column(Integer,nullable=False)
    option_order:Mapped[str]=mapped_column(String,nullable=False,default='ABCD')

class IntegrityEvent(Base):
    __tablename__='integrity_events'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    attempt_id:Mapped[int]=mapped_column(ForeignKey('attempts.id'),nullable=False)
    event_type:Mapped[str]=mapped_column(String,nullable=False)
    details:Mapped[str]=mapped_column(String,nullable=False,default='')
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class AuditLog(Base):
    __tablename__='audit_logs'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    actor:Mapped[str]=mapped_column(String,nullable=False)
    action:Mapped[str]=mapped_column(String,nullable=False)
    entity_type:Mapped[str]=mapped_column(String,nullable=False,default='')
    entity_id:Mapped[str]=mapped_column(String,nullable=False,default='')
    details:Mapped[str]=mapped_column(String,nullable=False,default='')
    created_at:Mapped[str]=mapped_column(String,nullable=False)
    prev_hash:Mapped[str]=mapped_column(String(64),nullable=False,default='')
    event_hash:Mapped[str]=mapped_column(String(64),nullable=False,default='')

class InstitutionProfile(Base):
    __tablename__='institution_profile'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    institution_name:Mapped[str]=mapped_column(String,nullable=False,default='Learn with Hemant')
    short_name:Mapped[str]=mapped_column(String,nullable=False,default='Learn with Hemant')
    system_name:Mapped[str]=mapped_column(String,nullable=False,default='Exam System')
    department:Mapped[str]=mapped_column(String,nullable=False,default='')
    academic_year:Mapped[str]=mapped_column(String,nullable=False,default='')
    admin_email:Mapped[str]=mapped_column(String,nullable=False,default='')
    exam_controller:Mapped[str]=mapped_column(String,nullable=False,default='')
    contact_phone:Mapped[str]=mapped_column(String,nullable=False,default='')
    logo_data:Mapped[str]=mapped_column(Text,nullable=False,default='')
    updated_at:Mapped[str]=mapped_column(String,nullable=False)

class FacultyRole(Base):
    __tablename__='faculty_roles'
    __table_args__=(UniqueConstraint('faculty_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    faculty_id:Mapped[int]=mapped_column(ForeignKey('faculty_users.id'),nullable=False)
    role:Mapped[str]=mapped_column(String,nullable=False,default='faculty')
    department:Mapped[str]=mapped_column(String,nullable=False,default='')
    updated_at:Mapped[str]=mapped_column(String,nullable=False)

class AcademicGroup(Base):
    __tablename__='academic_groups'
    __table_args__=(UniqueConstraint('department','program','semester','section','academic_year'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    department:Mapped[str]=mapped_column(String,nullable=False,default='')
    program:Mapped[str]=mapped_column(String,nullable=False)
    semester:Mapped[str]=mapped_column(String,nullable=False)
    section:Mapped[str]=mapped_column(String,nullable=False)
    academic_year:Mapped[str]=mapped_column(String,nullable=False,default='')
    is_active:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class StudentGroup(Base):
    __tablename__='student_groups'
    __table_args__=(UniqueConstraint('student_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    group_id:Mapped[int]=mapped_column(ForeignKey('academic_groups.id'),nullable=False)
    assigned_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamSession(Base):
    __tablename__='exam_sessions'
    __table_args__=(UniqueConstraint('exam_id','group_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    group_id:Mapped[int]=mapped_column(ForeignKey('academic_groups.id'),nullable=False)
    scheduled_start:Mapped[str]=mapped_column(String,nullable=False,default='')
    scheduled_end:Mapped[str]=mapped_column(String,nullable=False,default='')
    venue:Mapped[str]=mapped_column(String,nullable=False,default='')
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamApproval(Base):
    __tablename__='exam_approvals'
    __table_args__=(UniqueConstraint('exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    status:Mapped[str]=mapped_column(String,nullable=False,default='draft')
    requested_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    requested_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    reviewed_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    reviewed_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    comments:Mapped[str]=mapped_column(String,nullable=False,default='')


class ExamPracticeRelease(Base):
    __tablename__='exam_practice_releases'
    __table_args__=(UniqueConstraint('exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    is_released:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    release_after:Mapped[str]=mapped_column(String,nullable=False,default='')
    show_solutions:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    allow_mock:Mapped[bool]=mapped_column(Boolean,nullable=False,default=True)
    updated_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    updated_at:Mapped[str]=mapped_column(String,nullable=False,default='')

class PracticeAttempt(Base):
    __tablename__='practice_attempts'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    mode:Mapped[str]=mapped_column(String,nullable=False,default='practice')
    subject:Mapped[str]=mapped_column(String,nullable=False,default='')
    unit_filter:Mapped[str]=mapped_column(String,nullable=False,default='')
    difficulty_filter:Mapped[str]=mapped_column(String,nullable=False,default='')
    exam_id:Mapped[int|None]=mapped_column(ForeignKey('exams.id'),nullable=True)
    duration_minutes:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    started_at:Mapped[str]=mapped_column(String,nullable=False)
    ends_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    submitted_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    status:Mapped[str]=mapped_column(String,nullable=False,default='in_progress')
    score:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    total_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    question_refs_json:Mapped[str]=mapped_column(Text,nullable=False,default='[]')
    answers_json:Mapped[str]=mapped_column(Text,nullable=False,default='[]')
    incorrect_bank_ids_json:Mapped[str]=mapped_column(Text,nullable=False,default='[]')

class PracticeBookmark(Base):
    __tablename__='practice_bookmarks'
    __table_args__=(UniqueConstraint('student_id','bank_question_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    bank_question_id:Mapped[int]=mapped_column(ForeignKey('bank_questions.id'),nullable=False)
    created_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamSecurityPolicy(Base):
    __tablename__='exam_security_policies'
    __table_args__=(UniqueConstraint('exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    require_candidate_checkin:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    require_exam_pin:Mapped[bool]=mapped_column(Boolean,nullable=False,default=False)
    heartbeat_seconds:Mapped[int]=mapped_column(Integer,nullable=False,default=15)
    updated_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamStudentAccess(Base):
    __tablename__='exam_student_access'
    __table_args__=(UniqueConstraint('exam_id','student_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    pin_hash:Mapped[str]=mapped_column(String,nullable=False)
    pin_ciphertext:Mapped[str]=mapped_column(Text,nullable=False)
    issued_at:Mapped[str]=mapped_column(String,nullable=False)
    issued_by:Mapped[str]=mapped_column(String,nullable=False,default='')

class ExamDeviceLock(Base):
    __tablename__='exam_device_locks'
    __table_args__=(UniqueConstraint('exam_id','student_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    device_hash:Mapped[str]=mapped_column(String(64),nullable=False)
    locked_at:Mapped[str]=mapped_column(String,nullable=False)
    last_seen_at:Mapped[str]=mapped_column(String,nullable=False)

class ExamCandidateCheckin(Base):
    __tablename__='exam_candidate_checkins'
    __table_args__=(UniqueConstraint('exam_id','student_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    status:Mapped[str]=mapped_column(String,nullable=False,default='verified')
    verified_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    verified_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    notes:Mapped[str]=mapped_column(String,nullable=False,default='')

class AttemptHeartbeat(Base):
    __tablename__='attempt_heartbeats'
    __table_args__=(UniqueConstraint('attempt_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    attempt_id:Mapped[int]=mapped_column(ForeignKey('attempts.id'),nullable=False)
    last_seen_at:Mapped[str]=mapped_column(String,nullable=False)
    answer_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    client_state:Mapped[str]=mapped_column(String,nullable=False,default='active')
    client_fingerprint:Mapped[str]=mapped_column(String,nullable=False,default='')

class AuthThrottle(Base):
    __tablename__='auth_throttles'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    auth_key:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    failure_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    first_failed_at:Mapped[str]=mapped_column(String,nullable=False,default='')
    locked_until:Mapped[str]=mapped_column(String,nullable=False,default='')


class EdgePackageReceipt(Base):
    __tablename__='edge_package_receipts'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    package_id:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    source_mode:Mapped[str]=mapped_column(String,nullable=False,default='')
    source_exam_id:Mapped[str]=mapped_column(String,nullable=False,default='')
    imported_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    imported_at:Mapped[str]=mapped_column(String,nullable=False)


class EdgeResultReceipt(Base):
    __tablename__='edge_result_receipts'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    package_id:Mapped[str]=mapped_column(String,unique=True,nullable=False)
    source_mode:Mapped[str]=mapped_column(String,nullable=False,default='')
    source_exam_id:Mapped[str]=mapped_column(String,nullable=False,default='')
    origin_exam_id:Mapped[str]=mapped_column(String,nullable=False,default='')
    target_exam_id:Mapped[int|None]=mapped_column(ForeignKey('exams.id'),nullable=True)
    exam_title:Mapped[str]=mapped_column(String,nullable=False,default='')
    attempts_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    submitted_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    imported_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    imported_at:Mapped[str]=mapped_column(String,nullable=False)


class EdgeResultAttempt(Base):
    __tablename__='edge_result_attempts'
    __table_args__=(UniqueConstraint('receipt_id','roll_no'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    receipt_id:Mapped[int]=mapped_column(ForeignKey('edge_result_receipts.id'),nullable=False)
    roll_no:Mapped[str]=mapped_column(String,nullable=False)
    name:Mapped[str]=mapped_column(String,nullable=False,default='')
    status:Mapped[str]=mapped_column(String,nullable=False,default='')
    grading_status:Mapped[str]=mapped_column(String,nullable=False,default='')
    score:Mapped[int|None]=mapped_column(Integer,nullable=True)
    total_marks:Mapped[int|None]=mapped_column(Integer,nullable=True)
    integrity_count:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    payload_json:Mapped[str]=mapped_column(Text,nullable=False,default='{}')


class PracticalRegister(Base):
    __tablename__='practical_registers'
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    owner_type:Mapped[str]=mapped_column(String,nullable=False)  # admin | faculty
    owner_id:Mapped[int]=mapped_column(Integer,nullable=False)
    owner_name:Mapped[str]=mapped_column(String,nullable=False,default='')
    title:Mapped[str]=mapped_column(String,nullable=False)
    subject:Mapped[str]=mapped_column(String,nullable=False,default='')
    lab_code:Mapped[str]=mapped_column(String,nullable=False,default='')
    section:Mapped[str]=mapped_column(String,nullable=False,default='')
    academic_year:Mapped[str]=mapped_column(String,nullable=False,default='')
    default_max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=30)
    attendance_max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=5)
    record_max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=5)
    performance_max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=10)
    viva_max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=10)
    created_at:Mapped[str]=mapped_column(String,nullable=False)
    updated_at:Mapped[str]=mapped_column(String,nullable=False)


class PracticalStudent(Base):
    __tablename__='practical_students'
    __table_args__=(UniqueConstraint('register_id','roll_no'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    register_id:Mapped[int]=mapped_column(ForeignKey('practical_registers.id'),nullable=False)
    sequence:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    roll_no:Mapped[str]=mapped_column(String,nullable=False)
    name:Mapped[str]=mapped_column(String,nullable=False)
    created_at:Mapped[str]=mapped_column(String,nullable=False)


class PracticalExperiment(Base):
    __tablename__='practical_experiments'
    __table_args__=(UniqueConstraint('register_id','experiment_no'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    register_id:Mapped[int]=mapped_column(ForeignKey('practical_registers.id'),nullable=False)
    experiment_no:Mapped[str]=mapped_column(String,nullable=False)
    title:Mapped[str]=mapped_column(Text,nullable=False)
    # Faculty reference program used for deterministic Practical Code evaluation.
    # It is never shown to students.
    reference_code:Mapped[str]=mapped_column(Text,nullable=False,default='')
    # Optional faculty-defined exact-match penalty rules for Practical Code.
    # One rule per line; rules are never exposed to students.
    penalty_rules:Mapped[str]=mapped_column(Text,nullable=False,default='')
    max_marks:Mapped[int]=mapped_column(Integer,nullable=False,default=10)
    sort_order:Mapped[int]=mapped_column(Integer,nullable=False,default=0)
    created_at:Mapped[str]=mapped_column(String,nullable=False)


class PracticalMark(Base):
    __tablename__='practical_marks'
    __table_args__=(UniqueConstraint('practical_student_id','practical_experiment_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    register_id:Mapped[int]=mapped_column(ForeignKey('practical_registers.id'),nullable=False)
    practical_student_id:Mapped[int]=mapped_column(ForeignKey('practical_students.id'),nullable=False)
    practical_experiment_id:Mapped[int]=mapped_column(ForeignKey('practical_experiments.id'),nullable=False)
    attendance:Mapped[str]=mapped_column(String,nullable=False,default='')  # P | A | blank
    attendance_marks:Mapped[float|None]=mapped_column(Float,nullable=True)
    record_marks:Mapped[float|None]=mapped_column(Float,nullable=True)
    performance_marks:Mapped[float|None]=mapped_column(Float,nullable=True)
    viva_marks:Mapped[float|None]=mapped_column(Float,nullable=True)
    marks:Mapped[float|None]=mapped_column(Float,nullable=True)  # calculated total / legacy total
    remarks:Mapped[str]=mapped_column(Text,nullable=False,default='')
    updated_by:Mapped[str]=mapped_column(String,nullable=False,default='')
    updated_at:Mapped[str]=mapped_column(String,nullable=False,default='')


class PracticalCodeSubmission(Base):
    __tablename__='practical_code_submissions'
    __table_args__=(UniqueConstraint('student_id','exam_id'),)
    id:Mapped[int]=mapped_column(Integer,primary_key=True,autoincrement=True)
    student_id:Mapped[int]=mapped_column(ForeignKey('students.id'),nullable=False)
    exam_id:Mapped[int]=mapped_column(ForeignKey('exams.id'),nullable=False)
    register_id:Mapped[int]=mapped_column(ForeignKey('practical_registers.id'),nullable=False)
    practical_student_id:Mapped[int]=mapped_column(ForeignKey('practical_students.id'),nullable=False)
    practical_experiment_id:Mapped[int]=mapped_column(ForeignKey('practical_experiments.id'),nullable=False)
    experiment_no:Mapped[str]=mapped_column(String,nullable=False,default='')
    source_code:Mapped[str]=mapped_column(Text,nullable=False,default='')
    similarity_pct:Mapped[float]=mapped_column(Float,nullable=False,default=0.0)
    performance_marks:Mapped[float]=mapped_column(Float,nullable=False,default=0.0)
    submitted_at:Mapped[str]=mapped_column(String,nullable=False)


def _configure_database_reliability():
    if not DATABASE_URL.startswith('sqlite'):
        return
    try:
        with engine.begin() as conn:
            conn.exec_driver_sql('PRAGMA journal_mode=WAL')
            conn.exec_driver_sql('PRAGMA synchronous=NORMAL')
            conn.exec_driver_sql('PRAGMA busy_timeout=5000')
            conn.exec_driver_sql('PRAGMA foreign_keys=ON')
    except Exception:
        pass


def _ensure_column(table_name,column_name,ddl):
    columns={c['name'] for c in inspect(engine).get_columns(table_name)}
    if column_name in columns:
        return False
    with engine.begin() as conn:
        conn.execute(text(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl}'))
    return True


def run_schema_upgrades():
    """Idempotent in-app migration for V2.10 enterprise columns.

    Existing V26 SQLite/PostgreSQL deployments can start without losing data.
    New installations receive the columns directly from SQLAlchemy create_all().
    """
    upgrades=(
        ('questions','question_type',"VARCHAR NOT NULL DEFAULT 'single_choice'"),
        ('questions','answer_key',"TEXT NOT NULL DEFAULT ''"),
        ('questions','answer_tolerance',"VARCHAR NOT NULL DEFAULT ''"),
        ('questions','answer_case_sensitive',"BOOLEAN NOT NULL DEFAULT FALSE"),
        ('questions','practical_experiment_no',"VARCHAR NOT NULL DEFAULT ''"),
        ('answers','answer_value',"TEXT NOT NULL DEFAULT ''"),
        ('answers','manual_score',"INTEGER"),
        ('answers','grader_comment',"TEXT NOT NULL DEFAULT ''"),
        ('answers','graded_by',"VARCHAR NOT NULL DEFAULT ''"),
        ('answers','graded_at',"VARCHAR NOT NULL DEFAULT ''"),
        ('attempts','grading_status',"VARCHAR NOT NULL DEFAULT 'complete'"),
        ('bank_questions','answer_key',"TEXT NOT NULL DEFAULT ''"),
        ('bank_questions','answer_tolerance',"VARCHAR NOT NULL DEFAULT ''"),
        ('bank_questions','answer_case_sensitive',"BOOLEAN NOT NULL DEFAULT FALSE"),
        ('bank_questions','po_mapping',"VARCHAR NOT NULL DEFAULT ''"),
        ('bank_questions','pso_mapping',"VARCHAR NOT NULL DEFAULT ''"),
        ('bank_questions','practice_visibility',"VARCHAR NOT NULL DEFAULT 'official_only'"),
        ('bank_questions','practical_experiment_no',"VARCHAR NOT NULL DEFAULT ''"),
        ('bank_questions','explanation',"TEXT NOT NULL DEFAULT ''"),
        ('admins','mfa_secret',"VARCHAR NOT NULL DEFAULT ''"),
        ('admins','mfa_enabled',"BOOLEAN NOT NULL DEFAULT FALSE"),
        ('faculty_users','mfa_secret',"VARCHAR NOT NULL DEFAULT ''"),
        ('faculty_users','mfa_enabled',"BOOLEAN NOT NULL DEFAULT FALSE"),
        ('exam_security_policies','require_exam_pin',"BOOLEAN NOT NULL DEFAULT FALSE"),
        ('exam_configs','exam_type',"VARCHAR NOT NULL DEFAULT 'regular'"),
        ('exam_configs','practical_experiment_no',"VARCHAR NOT NULL DEFAULT ''"),
        ('exam_configs','practical_code_start_at',"VARCHAR NOT NULL DEFAULT ''"),
        ('exam_configs','practical_code_end_at',"VARCHAR NOT NULL DEFAULT ''"),
        ('audit_logs','prev_hash',"VARCHAR(64) NOT NULL DEFAULT ''"),
        ('audit_logs','event_hash',"VARCHAR(64) NOT NULL DEFAULT ''"),
        ('practical_registers','attendance_max_marks','INTEGER NOT NULL DEFAULT 5'),
        ('practical_registers','record_max_marks','INTEGER NOT NULL DEFAULT 5'),
        ('practical_registers','performance_max_marks','INTEGER NOT NULL DEFAULT 10'),
        ('practical_registers','viva_max_marks','INTEGER NOT NULL DEFAULT 10'),
        ('practical_experiments','reference_code',"TEXT NOT NULL DEFAULT ''"),
        ('practical_experiments','penalty_rules',"TEXT NOT NULL DEFAULT ''"),
        ('practical_marks','attendance_marks','FLOAT'),
        ('practical_marks','record_marks','FLOAT'),
        ('practical_marks','performance_marks','FLOAT'),
        ('practical_marks','viva_marks','FLOAT'),
        ('students','registration_no',"VARCHAR NOT NULL DEFAULT ''"),
    )
    for table_name,column_name,ddl in upgrades:
        _ensure_column(table_name,column_name,ddl)
    # Backfill new answer keys from the legacy one-letter columns.
    with engine.begin() as conn:
        conn.execute(text("UPDATE questions SET answer_key=correct_answer WHERE (answer_key IS NULL OR answer_key='') AND correct_answer IS NOT NULL"))
        conn.execute(text("UPDATE bank_questions SET answer_key=correct_answer WHERE (answer_key IS NULL OR answer_key='') AND correct_answer IS NOT NULL"))
        conn.execute(text("UPDATE answers SET answer_value=selected_answer WHERE (answer_value IS NULL OR answer_value='') AND selected_answer IS NOT NULL"))
        conn.execute(text("UPDATE practical_registers SET default_max_marks=attendance_max_marks+record_max_marks+performance_max_marks+viva_max_marks"))
        conn.execute(text("UPDATE practical_experiments SET max_marks=(SELECT default_max_marks FROM practical_registers WHERE practical_registers.id=practical_experiments.register_id) WHERE EXISTS (SELECT 1 FROM practical_registers WHERE practical_registers.id=practical_experiments.register_id)"))


def get_exam_security_policy(s,exam_id,create=False):
    row=s.scalar(select(ExamSecurityPolicy).where(ExamSecurityPolicy.exam_id==exam_id))
    if not row and create:
        row=ExamSecurityPolicy(exam_id=exam_id,require_candidate_checkin=False,require_exam_pin=False,heartbeat_seconds=15,updated_at=now_iso())
        s.add(row);s.flush()
    return row


def candidate_is_checked_in(s,exam_id,student_id):
    row=s.scalar(select(ExamCandidateCheckin).where(ExamCandidateCheckin.exam_id==exam_id,ExamCandidateCheckin.student_id==student_id))
    return bool(row and row.status=='verified')


def _exam_pin_fernet():
    key=base64.urlsafe_b64encode(hashlib.sha256((str(app.secret_key)+'|exam-pin-v1').encode('utf-8')).digest())
    return Fernet(key)


def encrypt_exam_pin(pin):
    return _exam_pin_fernet().encrypt(str(pin).encode('utf-8')).decode('ascii')


def decrypt_exam_pin(value):
    try:return _exam_pin_fernet().decrypt((value or '').encode('ascii')).decode('utf-8')
    except Exception:return ''


def exam_pin_record(s,exam_id,student_id):
    return s.scalar(select(ExamStudentAccess).where(ExamStudentAccess.exam_id==exam_id,ExamStudentAccess.student_id==student_id))


def exam_pin_matches(s,exam_id,student_id,pin):
    row=exam_pin_record(s,exam_id,student_id)
    return bool(row and pin and check_password_hash(row.pin_hash,str(pin).strip()))


ROTATING_EXAM_PIN_SECONDS=60
ROTATING_EXAM_PIN_GRACE_SECONDS=3


def rotating_exam_pin(exam_id,student_id,at_time=None):
    now=int(time.time() if at_time is None else at_time)
    slot=now//ROTATING_EXAM_PIN_SECONDS
    payload=f'exam-rotating-pin-v1|{int(exam_id)}|{int(student_id)}|{slot}'.encode('utf-8')
    digest=hmac.new(str(app.secret_key).encode('utf-8'),payload,hashlib.sha256).digest()
    return f'{int.from_bytes(digest[:8],"big")%1000000:06d}'


def rotating_exam_pin_seconds_remaining(at_time=None):
    now=int(time.time() if at_time is None else at_time)
    return ROTATING_EXAM_PIN_SECONDS-(now%ROTATING_EXAM_PIN_SECONDS)


def rotating_exam_pin_matches(exam_id,student_id,pin,at_time=None):
    value=str(pin or '').strip()
    if not re.fullmatch(r'\d{6}',value):return False
    now=int(time.time() if at_time is None else at_time)
    if secrets.compare_digest(value,rotating_exam_pin(exam_id,student_id,now)):return True
    # Tiny rollover grace avoids rejecting a code typed just as the 60-second window changes.
    if now%ROTATING_EXAM_PIN_SECONDS<=ROTATING_EXAM_PIN_GRACE_SECONDS:
        return secrets.compare_digest(value,rotating_exam_pin(exam_id,student_id,now-ROTATING_EXAM_PIN_SECONDS))
    return False


def eligible_students_for_exam(s,exam_id):
    sessions=s.scalars(select(ExamSession).where(ExamSession.exam_id==exam_id)).all()
    if not sessions:
        return s.scalars(select(Student).order_by(Student.roll_no)).all()
    group_ids={row.group_id for row in sessions}
    if not group_ids:return []
    return s.scalars(select(Student).join(StudentGroup,StudentGroup.student_id==Student.id).where(StudentGroup.group_id.in_(group_ids)).order_by(Student.roll_no)).all()


def generate_exam_pins(s,exam_id,regenerate=False,issued_by=''):
    students=eligible_students_for_exam(s,exam_id);created=0;kept=0;used=set()
    existing=s.scalars(select(ExamStudentAccess)).all()
    for row in existing:
        pin=decrypt_exam_pin(row.pin_ciphertext)
        if pin:used.add(pin)
    for student in students:
        row=exam_pin_record(s,exam_id,student.id)
        if row and not regenerate:
            kept+=1;continue
        if row and regenerate:
            old=decrypt_exam_pin(row.pin_ciphertext)
            if old:used.discard(old)
        pin=''
        for _ in range(50):
            candidate=f'{secrets.randbelow(900000)+100000:06d}'
            if candidate not in used:
                pin=candidate;break
        if not pin:raise RuntimeError('Could not generate a unique Exam PIN.')
        used.add(pin)
        if not row:
            row=ExamStudentAccess(exam_id=exam_id,student_id=student.id,pin_hash='',pin_ciphertext='',issued_at=now_iso(),issued_by=issued_by or '')
            s.add(row)
        row.pin_hash=generate_password_hash(pin);row.pin_ciphertext=encrypt_exam_pin(pin);row.issued_at=now_iso();row.issued_by=issued_by or ''
        created+=1
    return created,kept,len(students)


def _exam_pin_verified_map():
    value=web_session.get('_exam_pin_verified') or {}
    return value if isinstance(value,dict) else {}


def mark_exam_pin_verified(exam_id):
    value=_exam_pin_verified_map();value[str(int(exam_id))]=int(time.time());web_session['_exam_pin_verified']=value


def exam_pin_is_verified(exam_id):
    value=_exam_pin_verified_map();ts=int(value.get(str(int(exam_id))) or 0)
    return ts>0 and int(time.time())-ts<=SESSION_TIMEOUT_MINUTES*60


def create_secure_exam_launch_token(exam_id):
    token=secrets.token_urlsafe(32)
    launches=web_session.get('_secure_exam_launches') or {}
    if not isinstance(launches,dict):launches={}
    launches[str(int(exam_id))]={'token':token,'issued_at':int(time.time())}
    web_session['_secure_exam_launches']=launches
    return token


def secure_exam_launch_token_valid(exam_id,token):
    if not token:return False
    launches=web_session.get('_secure_exam_launches') or {}
    if not isinstance(launches,dict):return False
    row=launches.get(str(int(exam_id))) or {}
    if not isinstance(row,dict):return False
    issued_at=int(row.get('issued_at') or 0)
    expected=str(row.get('token') or '')
    # The token is never exposed on the dashboard. It only authorizes the
    # silent same-session exam shell after a successful PIN verification.
    return bool(expected and issued_at and int(time.time())-issued_at<=SESSION_TIMEOUT_MINUTES*60 and secrets.compare_digest(expected,str(token)))


def clear_secure_exam_launch_token(exam_id):
    launches=web_session.get('_secure_exam_launches') or {}
    if isinstance(launches,dict) and str(int(exam_id)) in launches:
        launches.pop(str(int(exam_id)),None);web_session['_secure_exam_launches']=launches


def _ensure_student_device_token():
    token=(web_session.get('_student_device_token') or '').strip()
    if not token:
        token=secrets.token_urlsafe(24);web_session['_student_device_token']=token
    return token


def current_student_device_hash():
    token=_ensure_student_device_token();ua=(request.headers.get('User-Agent') or '')[:500]
    return hmac.new(str(app.secret_key).encode('utf-8'),f'{token}|{ua}'.encode('utf-8'),hashlib.sha256).hexdigest()


def ensure_exam_device_lock(s,exam_id,student_id):
    device_hash=current_student_device_hash();row=s.scalar(select(ExamDeviceLock).where(ExamDeviceLock.exam_id==exam_id,ExamDeviceLock.student_id==student_id))
    if not row:
        row=ExamDeviceLock(exam_id=exam_id,student_id=student_id,device_hash=device_hash,locked_at=now_iso(),last_seen_at=now_iso());s.add(row);s.flush()
        audit_event(s,'exam_device_locked','exam',exam_id,f'student_id={student_id}')
        return True,row
    if not secrets.compare_digest(row.device_hash,device_hash):
        audit_event(s,'exam_device_lock_blocked','exam',exam_id,f'student_id={student_id}, ip={request.remote_addr or ""}')
        return False,row
    row.last_seen_at=now_iso();return True,row


def secure_exam_device_allowed(s,attempt):
    security=get_exam_security_policy(s,attempt.exam_id,create=False)
    if not security or not security.require_exam_pin:return True
    row=s.scalar(select(ExamDeviceLock).where(ExamDeviceLock.exam_id==attempt.exam_id,ExamDeviceLock.student_id==attempt.student_id))
    return bool(row and secrets.compare_digest(row.device_hash,current_student_device_hash()))


def find_active_exam_for_pin(s,student_id,pin):
    if not pin:return None
    rows=s.execute(select(ExamStudentAccess,Exam).join(Exam,Exam.id==ExamStudentAccess.exam_id).join(ExamSecurityPolicy,ExamSecurityPolicy.exam_id==Exam.id).where(ExamStudentAccess.student_id==student_id,Exam.is_active==True,ExamSecurityPolicy.require_exam_pin==True).order_by(Exam.id.desc())).all()
    for access,exam in rows:
        if check_password_hash(access.pin_hash,str(pin).strip()):return exam
    return None


PRACTICE_VISIBILITY_LABELS={
    'official_only':'Official Exam Only',
    'practice_only':'Practice Only',
    'both':'Official + Practice',
    'practical_exam':'Practical Exam',
}

def normalize_practice_visibility(value):
    value=(value or 'official_only').strip().lower()
    aliases={'official':'official_only','exam':'official_only','practice':'practice_only','published':'practice_only','all':'both','practical':'practical_exam','practical exam':'practical_exam'}
    value=aliases.get(value,value)
    return value if value in PRACTICE_VISIBILITY_LABELS else 'official_only'

def normalize_practical_exam_no(value):
    value=(value or '').strip()
    if not value:
        return ''
    return normalize_experiment_code(value,1)

def get_exam_practice_release(s,exam_id,create=False):
    row=s.scalar(select(ExamPracticeRelease).where(ExamPracticeRelease.exam_id==exam_id))
    if not row and create:
        row=ExamPracticeRelease(exam_id=exam_id,is_released=False,release_after='',show_solutions=True,allow_mock=True,updated_by='',updated_at=now_iso())
        s.add(row);s.flush()
    return row

def practice_release_is_available(row):
    if not row or not row.is_released:return False
    if not (row.release_after or '').strip():return True
    try:return now_dt()>=parse_dt(row.release_after)
    except Exception:return False

def question_is_practice_eligible(q):
    return bool(q and q.status=='approved' and normalize_practice_visibility(q.practice_visibility) in {'practice_only','both'} and canonical_question_type(q.question_type)!='essay')

def question_is_official_eligible(q):
    return bool(q and q.status=='approved' and normalize_practice_visibility(q.practice_visibility) in {'official_only','both','practical_exam'})

def safe_json_load(value,default):
    try:
        parsed=json.loads(value or '')
        return parsed if isinstance(parsed,type(default)) else default
    except Exception:return default

def question_definition_from_form(form):
    qtype=canonical_question_type(form.get('question_type','single_choice'))
    options={key:(form.get(f'option_{key.lower()}') or '').strip() for key in 'ABCD'}
    if qtype=='single_choice':
        answer_key=(form.get('correct_answer') or '').strip().upper()
    elif qtype=='multiple_select':
        selected=form.getlist('correct_answers') if hasattr(form,'getlist') else []
        answer_key=','.join(selected) if selected else (form.get('answer_key') or '')
    elif qtype=='true_false':
        answer_key=(form.get('true_false_answer') or '').strip()
    elif qtype=='numerical':
        answer_key=(form.get('numerical_answer_key') or form.get('answer_key') or '').strip()
    elif qtype=='short_text':
        answer_key=(form.get('short_answer_key') or form.get('answer_key') or '').strip()
    elif qtype=='essay':
        answer_key=''
    else:
        answer_key=(form.get('answer_key') or '').strip()
    tolerance=(form.get('answer_tolerance') or '').strip()
    case_sensitive=form.get('answer_case_sensitive')=='on'
    error=validate_question_definition(qtype,form.get('question',''),options,answer_key,tolerance)
    legacy=(answer_key[:1].upper() if qtype=='single_choice' and answer_key[:1].upper() in {'A','B','C','D'} else 'A')
    return {'question_type':qtype,'options':options,'answer_key':answer_key,'answer_tolerance':tolerance,'answer_case_sensitive':case_sensitive,'legacy_correct_answer':legacy,'error':error}


def answer_record_value(answer):
    return (getattr(answer,'answer_value','') or getattr(answer,'selected_answer','') or '').strip()


def _request_identity_key(kind,username):
    ip=(request.remote_addr or 'unknown').strip()
    digest=hmac.new(app.secret_key.encode('utf-8'),f'{kind}|{username}|{ip}'.encode('utf-8'),hashlib.sha256).hexdigest()
    return digest


def auth_is_locked(s,key):
    row=s.scalar(select(AuthThrottle).where(AuthThrottle.auth_key==key))
    if not row or not row.locked_until:
        return False,0
    try:
        until=parse_dt(row.locked_until)
    except Exception:
        return False,0
    remaining=int((until-now_dt()).total_seconds())
    return remaining>0,max(0,remaining)


def record_auth_failure(s,key):
    now=now_dt();row=s.scalar(select(AuthThrottle).where(AuthThrottle.auth_key==key))
    if not row:
        row=AuthThrottle(auth_key=key,failure_count=0,first_failed_at=now.isoformat(timespec='seconds'),locked_until='');s.add(row);s.flush()
    try:first=parse_dt(row.first_failed_at) if row.first_failed_at else now
    except Exception:first=now
    if (now-first).total_seconds()>LOGIN_WINDOW_MINUTES*60:
        row.failure_count=0;row.first_failed_at=now.isoformat(timespec='seconds');row.locked_until=''
    row.failure_count+=1
    if row.failure_count>=LOGIN_MAX_FAILURES:
        row.locked_until=(now+timedelta(minutes=LOGIN_LOCK_MINUTES)).isoformat(timespec='seconds')
    s.commit()


def clear_auth_failures(s,key):
    row=s.scalar(select(AuthThrottle).where(AuthThrottle.auth_key==key))
    if row:s.delete(row);s.commit()


def heartbeat_status(last_seen):
    if not last_seen:return 'unknown'
    try:age=(now_dt()-parse_dt(last_seen)).total_seconds()
    except Exception:return 'unknown'
    if age<=HEARTBEAT_STALE_SECONDS:return 'online'
    if age<=HEARTBEAT_STALE_SECONDS*3:return 'stale'
    return 'offline'


def now_dt():
    """Current application time in the configured institutional timezone."""
    return datetime.now(DISPLAY_TZ)

def now_iso():
    return now_dt().isoformat(timespec='seconds')

def parse_dt(value):
    """Parse stored timestamps and attach the configured timezone to legacy naive values."""
    dt=datetime.fromisoformat(value)
    return dt if dt.tzinfo else dt.replace(tzinfo=DISPLAY_TZ)

def display_dt(value):
    """Convert any stored timestamp (UTC/local/legacy) to the display timezone."""
    return parse_dt(value).astimezone(DISPLAY_TZ)

def actor_label(s=None):
    role=web_session.get('role','system')
    uid=web_session.get('user_id')
    if role=='admin':
        return f"admin:{web_session.get('username',admin_username)}"
    if role=='faculty' and uid:
        s=s or DB(); row=s.get(Faculty,uid)
        return f'faculty:{row.username if row else uid}'
    if role=='student' and uid:
        return f'student:{uid}'
    return role

def audit_event(s,action,entity_type='',entity_id='',details=''):
    actor=actor_label(s);action=str(action or '');entity_type=str(entity_type or '');entity_id=str(entity_id or '');details=str(details or '')[:1500];created_at=now_iso()
    previous=s.scalar(select(AuditLog).where(AuditLog.event_hash!='').order_by(AuditLog.id.desc()).limit(1))
    prev_hash=previous.event_hash if previous else ''
    event_hash=audit_event_hash(prev_hash=prev_hash,actor=actor,action=action,entity_type=entity_type,entity_id=entity_id,details=details,created_at=created_at)
    s.add(AuditLog(actor=actor,action=action,entity_type=entity_type,entity_id=entity_id,details=details,created_at=created_at,prev_hash=prev_hash,event_hash=event_hash))

def audit_chain_status(s):
    sealed=s.scalars(select(AuditLog).where(AuditLog.event_hash!='').order_by(AuditLog.id.asc())).all()
    legacy=s.scalar(select(func.count()).select_from(AuditLog).where(or_(AuditLog.event_hash=='',AuditLog.event_hash.is_(None)))) or 0
    rows=[{'prev_hash':r.prev_hash,'event_hash':r.event_hash,'actor':r.actor,'action':r.action,'entity_type':r.entity_type,'entity_id':r.entity_id,'details':r.details,'created_at':r.created_at} for r in sealed]
    result=verify_audit_rows(rows);result['legacy_count']=legacy
    return result

ROLE_LABELS={'super_admin':'Super Admin','exam_controller':'Exam Controller','hod':'HOD','faculty':'Faculty'}
APPROVER_ROLES={'super_admin','exam_controller','hod'}
PRACTICAL_ROLES={'super_admin','hod','faculty'}
FACULTY_DAILY_SELF_APPROVAL_LIMIT=3
EXAM_CREATION_AUDIT_ACTIONS={
    'exam_created',
    'catalog_subject_exam_created',
    'ready_exam_created',
    'unit_set_exam_created',
}


def get_institution(s=None,create=True):
    s=s or DB()
    row=s.scalar(select(InstitutionProfile).order_by(InstitutionProfile.id.asc()))
    if not row and create:
        row=InstitutionProfile(institution_name='Learn with Hemant',short_name='Learn with Hemant',system_name='Exam System',department='',academic_year='',admin_email='',exam_controller='',contact_phone='',logo_data='',updated_at=now_iso())
        s.add(row);s.flush()
    return row


def current_staff_role(s=None):
    if web_session.get('role')=='admin': return 'super_admin'
    if web_session.get('role')!='faculty': return web_session.get('role','')
    s=s or DB(); uid=web_session.get('user_id')
    row=s.scalar(select(FacultyRole).where(FacultyRole.faculty_id==uid)) if uid else None
    return row.role if row and row.role in ROLE_LABELS else 'faculty'


def current_staff_name(s=None):
    """Return the human name for the currently signed-in staff account.

    The dashboard must identify the person/account, not merely the permission role
    (for example, show ``Hemant Dashboard`` while the header role chip can still
    show ``HOD``). Existing sessions are supported by resolving the account from
    either ``user_id`` or ``username``.
    """
    role=web_session.get('role')
    username=(web_session.get('username') or '').strip()

    def readable_username(value):
        return value.replace('.',' ').replace('_',' ').replace('-',' ').strip().title()

    if role=='faculty':
        s=s or DB(); uid=web_session.get('user_id')
        row=s.get(Faculty,uid) if uid else None
        if row is None and username:
            row=s.scalar(select(Faculty).where(Faculty.username==username))
        if row:
            account_name=(row.name or '').strip()
            account_username=(row.username or username or '').strip()
            # A few older accounts were created with the role itself in the Name
            # field (e.g. "HOD"). In that case the username is a better account
            # identifier than repeating the role as the dashboard title.
            generic_role_names={'hod','faculty','admin','administrator','exam controller','super admin','staff'}
            if account_name and account_name.casefold() not in generic_role_names:
                return account_name
            if account_username:
                return readable_username(account_username)
            if account_name:
                return account_name

    if role=='admin':
        configured=(os.getenv('SUPER_ADMIN_DISPLAY_NAME') or '').strip()
        if configured:
            return configured

    readable=readable_username(username)
    return readable if readable else 'Staff'


def can_approve_exams(s=None): return current_staff_role(s) in APPROVER_ROLES

def can_approve_content(s=None): return current_staff_role(s) in APPROVER_ROLES


def current_practical_owner(s=None):
    """Return the account identity that owns practical registers."""
    role=web_session.get('role')
    uid=int(web_session.get('user_id') or 0)
    if role=='admin':
        return 'admin',uid,current_staff_name(s)
    if role=='faculty':
        return 'faculty',uid,current_staff_name(s)
    return '',0,''


def practical_register_access(s,register_id):
    row=s.get(PracticalRegister,register_id)
    if not row:abort(404)
    if current_staff_role(s)=='super_admin':
        return row
    owner_type,owner_id,_=current_practical_owner(s)
    if row.owner_type!=owner_type or row.owner_id!=owner_id:
        abort(403)
    return row


def practical_register_stmt(s):
    stmt=select(PracticalRegister).order_by(PracticalRegister.updated_at.desc(),PracticalRegister.id.desc())
    if current_staff_role(s)!='super_admin':
        owner_type,owner_id,_=current_practical_owner(s)
        stmt=stmt.where(PracticalRegister.owner_type==owner_type,PracticalRegister.owner_id==owner_id)
    return stmt

def exam_owner_actor(s,exam_id):
    """Return the staff actor that originally created an exam.

    Ownership is derived from the immutable audit trail so this policy works with
    existing databases without adding a new schema column.
    """
    owner=s.scalar(select(AuditLog.actor).where(
        AuditLog.entity_type=='exam',
        AuditLog.entity_id==str(exam_id),
        AuditLog.action.in_(tuple(EXAM_CREATION_AUDIT_ACTIONS)),
    ).order_by(AuditLog.id.asc()).limit(1))
    if owner:
        return owner
    approval=get_exam_approval(s,exam_id,create=False)
    return (approval.requested_by or '') if approval else ''

def _stored_local_date(value):
    value=(value or '').strip()
    if not value:
        return ''
    try:
        return parse_dt(value).astimezone(DISPLAY_TZ).date().isoformat()
    except Exception:
        try:
            return datetime.fromisoformat(value).date().isoformat()
        except Exception:
            return ''

def exam_scheduled_dates(s,exam_id):
    dates=set()
    for value in s.scalars(select(ExamSession.scheduled_start).where(ExamSession.exam_id==exam_id)).all():
        day=_stored_local_date(value)
        if day:
            dates.add(day)
    return dates

def faculty_exam_ids_for_day(s,owner_actor,target_day,include_exam_id=None):
    """Return a faculty member's exams for one day in conducting order.

    Scheduled exams are ordered by their earliest session start on that day.
    Unscheduled exams fall back to activation/creation time. This lets only the
    4th and subsequent exams require external approval instead of blocking the
    faculty member's first three exams merely because a fourth is also planned.
    """
    created=s.execute(select(AuditLog.entity_id).where(
        AuditLog.entity_type=='exam',
        AuditLog.actor==owner_actor,
        AuditLog.action.in_(tuple(EXAM_CREATION_AUDIT_ACTIONS)),
    )).all()
    ids=set()
    for (raw_id,) in created:
        try:
            ids.add(int(raw_id))
        except (TypeError,ValueError):
            continue
    if include_exam_id:
        ids.add(int(include_exam_id))
    today=now_dt().date().isoformat()
    entries=[]
    for exam_id in ids:
        exam=s.get(Exam,exam_id)
        if not exam:
            continue
        starts=[]
        for value in s.scalars(select(ExamSession.scheduled_start).where(ExamSession.exam_id==exam_id)).all():
            if _stored_local_date(value)==target_day:
                try:
                    starts.append(parse_dt(value).astimezone(DISPLAY_TZ))
                except Exception:
                    pass
        all_scheduled_dates=exam_scheduled_dates(s,exam_id)
        occurs=bool(starts)
        sort_dt=min(starts) if starts else None
        if not all_scheduled_dates:
            activated=s.scalars(select(AuditLog.created_at).where(
                AuditLog.entity_type=='exam',
                AuditLog.entity_id==str(exam_id),
                AuditLog.action=='exam_activated',
                AuditLog.actor==owner_actor,
            ).order_by(AuditLog.id.asc())).all()
            activation_on_day=[]
            for value in activated:
                if _stored_local_date(value)==target_day:
                    try:
                        activation_on_day.append(parse_dt(value).astimezone(DISPLAY_TZ))
                    except Exception:
                        pass
            if activation_on_day:
                occurs=True
                sort_dt=min(activation_on_day)
            elif target_day==today and exam.is_active:
                occurs=True
            if occurs and sort_dt is None:
                try:
                    sort_dt=parse_dt(exam.created_at).astimezone(DISPLAY_TZ)
                except Exception:
                    sort_dt=now_dt()
        if exam_id==include_exam_id:
            occurs=True
            if sort_dt is None:
                try:
                    sort_dt=parse_dt(exam.created_at).astimezone(DISPLAY_TZ)
                except Exception:
                    sort_dt=now_dt()
        if occurs:
            entries.append((sort_dt or now_dt(),exam_id))
    entries.sort(key=lambda item:(item[0],item[1]))
    return [exam_id for _,exam_id in entries]

def exam_counts_for_faculty_day(s,owner_actor,target_day,include_exam_id=None):
    return len(faculty_exam_ids_for_day(s,owner_actor,target_day,include_exam_id=include_exam_id))

def faculty_exam_ordinal_for_day(s,owner_actor,target_day,exam_id):
    ids=faculty_exam_ids_for_day(s,owner_actor,target_day,include_exam_id=exam_id)
    try:
        return ids.index(int(exam_id))+1,len(ids)
    except ValueError:
        return 1,max(1,len(ids))

def exam_approval_policy(s,exam):
    """Return the approval rule for the currently logged-in staff member.

    Ordinary Faculty may self-approve their own first three exams on a day.
    The fourth and subsequent exams on the same day require approval from an HOD,
    Exam Controller or Super Admin. Privileged approvers retain their normal rights.
    """
    role=current_staff_role(s)
    result={
        'role':role,
        'limit':FACULTY_DAILY_SELF_APPROVAL_LIMIT,
        'owner_actor':exam_owner_actor(s,exam.id),
        'self_approval_allowed':False,
        'external_approval_required':False,
        'daily_exam_count':0,
        'date_counts':{},
        'date_labels':[],
        'message':'',
    }
    if role!='faculty':
        return result
    actor=actor_label(s)
    if result['owner_actor'] and result['owner_actor']!=actor:
        result['external_approval_required']=True
        result['message']='This exam was created by another staff account, so faculty self-approval is not available.'
        return result
    if not result['owner_actor']:
        result['external_approval_required']=True
        result['message']='Exam ownership could not be verified from the audit trail; HOD / Exam Controller approval is required.'
        return result
    dates=exam_scheduled_dates(s,exam.id)
    if not dates:
        dates={now_dt().date().isoformat()}
    counts={}
    ordinals={}
    for day in sorted(dates):
        ordinal,total=faculty_exam_ordinal_for_day(s,actor,day,exam.id)
        counts[day]=total
        ordinals[day]=ordinal
    max_count=max(counts.values(),default=1)
    result['date_counts']=counts
    result['daily_exam_count']=max_count
    result['date_labels']=[datetime.fromisoformat(day).strftime('%d %b %Y') for day in counts]
    result['external_approval_required']=any(ordinal>FACULTY_DAILY_SELF_APPROVAL_LIMIT for ordinal in ordinals.values())
    result['self_approval_allowed']=not result['external_approval_required']
    if result['external_approval_required']:
        blocked_day=next(day for day in sorted(ordinals) if ordinals[day]>FACULTY_DAILY_SELF_APPROVAL_LIMIT)
        result['message']=(f'External approval required: this is exam #{ordinals[blocked_day]} of {counts[blocked_day]} on '
                           f'{datetime.fromisoformat(blocked_day).strftime("%d %b %Y")}. '
                           f'Faculty may self-approve only their first {FACULTY_DAILY_SELF_APPROVAL_LIMIT} exams per day.')
    else:
        detail=', '.join(f'{datetime.fromisoformat(day).strftime("%d %b %Y")}: exam #{ordinals[day]} of {counts[day]}' for day in sorted(ordinals))
        result['message']=(f'Self-approval allowed ({detail}). Faculty may self-approve their first '
                           f'{FACULTY_DAILY_SELF_APPROVAL_LIMIT} exams per day; activation will approve this exam automatically.')
    return result

def can_manage_staff_passwords(s=None):
    """Super Admin can reset any staff password; HOD can reset Faculty passwords only."""
    return current_staff_role(s) in {'super_admin','hod'}


def get_exam_approval(s,exam_id,create=True):
    row=s.scalar(select(ExamApproval).where(ExamApproval.exam_id==exam_id))
    if not row and create:
        row=ExamApproval(exam_id=exam_id,status='draft',requested_by='',requested_at='',reviewed_by='',reviewed_at='',comments='')
        s.add(row);s.flush()
    return row


def group_label(group):
    if not group:return 'Unassigned'
    bits=[group.program]
    if group.semester:bits.append(f'Sem {group.semester}')
    if group.section:bits.append(f'Section {group.section}')
    label=' · '.join(bits)
    return f'{label} ({group.academic_year})' if group.academic_year else label


def student_group(s,student_id):
    return s.execute(select(AcademicGroup).join(StudentGroup,StudentGroup.group_id==AcademicGroup.id).where(StudentGroup.student_id==student_id)).scalar_one_or_none()


def find_or_create_group(s,department,program,semester,section,academic_year):
    department=(department or '').strip();program=(program or '').strip();semester=(semester or '').strip();section=(section or '').strip();academic_year=(academic_year or '').strip()
    if not program or not semester or not section:return None
    row=s.scalar(select(AcademicGroup).where(AcademicGroup.department==department,AcademicGroup.program==program,AcademicGroup.semester==semester,AcademicGroup.section==section,AcademicGroup.academic_year==academic_year))
    if not row:
        row=AcademicGroup(department=department,program=program,semester=semester,section=section,academic_year=academic_year,is_active=True,created_at=now_iso());s.add(row);s.flush()
    return row


def assign_student_group(s,student_id,group_id):
    row=s.scalar(select(StudentGroup).where(StudentGroup.student_id==student_id))
    if group_id:
        if row:row.group_id=group_id;row.assigned_at=now_iso()
        else:s.add(StudentGroup(student_id=student_id,group_id=group_id,assigned_at=now_iso()))
    elif row:s.delete(row)


def _canonical_registration_no(value):
    """Normalize a university registration number only for duplicate matching."""
    return ''.join(ch for ch in (value or '').strip().upper() if ch.isalnum())


def _student_login_from_registration(value):
    """Return the last five digits used as both student login ID and default password."""
    digits=''.join(ch for ch in (value or '') if ch.isdigit())
    return digits[-5:] if len(digits)>=5 else ''


def _section_code(value):
    """Extract a compact section code such as E/K from a label when possible."""
    raw=(value or '').strip()
    if not raw:return ''
    match=re.search(r'\bsection\s*[-:/]?\s*([A-Za-z0-9]+)\b',raw,re.I)
    if match:return match.group(1).strip().upper()
    # A practical register often stores only the section letter in this field.
    if re.fullmatch(r'[A-Za-z][A-Za-z0-9_-]{0,7}',raw) and not re.search(r'\bsem(?:ester)?\b',raw,re.I):
        return raw.upper()
    return ''


def _practical_register_section_code(register):
    """Prefer the register's Section/Batch field, then its title."""
    return _section_code(getattr(register,'section','')) or _section_code(getattr(register,'title',''))


def parse_local_schedule(value):
    value=(value or '').strip()
    if not value:return ''
    try:return datetime.fromisoformat(value).strftime('%Y-%m-%dT%H:%M:%S')
    except ValueError:raise ValueError('Schedule date/time is invalid.')


def practical_code_window_state(cfg,at_time=None):
    """Return the server-authoritative Practical Code editing-window state.

    Practical-code timestamps are saved as naive local institutional times, the
    same format used by ExamSession.  This keeps Render/UTC hosts and offline
    Windows deployments consistent with the APP_TIMEZONE selected by the college.
    """
    start_value=(getattr(cfg,'practical_code_start_at','') or '').strip() if cfg else ''
    end_value=(getattr(cfg,'practical_code_end_at','') or '').strip() if cfg else ''
    now_local=(at_time or now_dt()).astimezone(DISPLAY_TZ).replace(tzinfo=None)
    start=None;end=None
    try:start=datetime.fromisoformat(start_value) if start_value else None
    except ValueError:start=None
    try:end=datetime.fromisoformat(end_value) if end_value else None
    except ValueError:end=None

    configured=bool(start and end)
    if not configured:
        return {
            'configured':False,'status':'unrestricted','can_edit':True,
            'start_at':start_value,'end_at':end_value,'end_epoch_ms':None,'label':''
        }
    if now_local<start:
        return {
            'configured':True,'status':'upcoming','can_edit':False,
            'start_at':start_value,'end_at':end_value,
            'end_epoch_ms':int(end.replace(tzinfo=DISPLAY_TZ).timestamp()*1000),
            'label':f'Practical code entry opens on {start.strftime("%d %b %Y, %I:%M %p")}.'
        }
    if now_local>=end:
        return {
            'configured':True,'status':'closed','can_edit':False,
            'start_at':start_value,'end_at':end_value,
            'end_epoch_ms':int(end.replace(tzinfo=DISPLAY_TZ).timestamp()*1000),
            'label':f'Editing closed on {end.strftime("%d %b %Y, %I:%M %p")}. Your code and marks are view-only.'
        }
    return {
        'configured':True,'status':'open','can_edit':True,
        'start_at':start_value,'end_at':end_value,
        'end_epoch_ms':int(end.replace(tzinfo=DISPLAY_TZ).timestamp()*1000),'label':''
    }


def practical_exam_assigned_to_student(s,student_id,exam):
    """Check only batch/section assignment, not the normal exam clock window."""
    sessions=s.scalars(select(ExamSession).where(ExamSession.exam_id==exam.id)).all()
    if not sessions:
        return True
    membership=s.scalar(select(StudentGroup).where(StudentGroup.student_id==student_id))
    if not membership:
        return False
    return any(row.group_id==membership.group_id for row in sessions)


def exam_access_for_student(s,student_id,exam):
    security=get_exam_security_policy(s,exam.id,create=False)
    if security and security.require_candidate_checkin and not candidate_is_checked_in(s,exam.id,student_id):
        return False,'Identity check-in required at the exam centre',None
    sessions=s.scalars(select(ExamSession).where(ExamSession.exam_id==exam.id)).all()
    if not sessions:return True,'Available',None
    membership=s.scalar(select(StudentGroup).where(StudentGroup.student_id==student_id))
    if not membership:return False,'Not assigned to your batch/section',None
    matched=next((x for x in sessions if x.group_id==membership.group_id),None)
    if not matched:return False,'Not assigned to your batch/section',None
    now=now_dt().replace(tzinfo=None)
    if matched.scheduled_start:
        start=datetime.fromisoformat(matched.scheduled_start)
        if now<start:return False,f'Scheduled for {start.strftime("%d %b, %I:%M %p")}',matched
    if matched.scheduled_end:
        end=datetime.fromisoformat(matched.scheduled_end)
        if now>end:return False,'Exam window closed',matched
    return True,'Available',matched


def local_lan_ip():
    """Return the IPv4 address used by the machine's current default LAN route.

    The UDP connect does not send application data; it asks the OS which local
    interface would be used for that route. This keeps the Exam Centre URL in
    sync when the server laptop moves between Wi-Fi/LAN/hotspot networks.
    """
    probes=(('8.8.8.8',80),('1.1.1.1',80),('192.0.2.1',80))
    for host,port in probes:
        sock=None
        try:
            sock=socket.socket(socket.AF_INET,socket.SOCK_DGRAM)
            sock.connect((host,port))
            ip=sock.getsockname()[0]
            if ip and not ip.startswith('127.') and not ip.startswith('169.254.'):
                return ip
        except OSError:
            pass
        finally:
            if sock:
                try:sock.close()
                except OSError:pass
    try:
        candidates=[]
        for info in socket.getaddrinfo(socket.gethostname(),None,socket.AF_INET,socket.SOCK_DGRAM):
            ip=info[4][0]
            if ip and not ip.startswith('127.') and not ip.startswith('169.254.') and ip not in candidates:
                candidates.append(ip)
        private=next((ip for ip in candidates if ip.startswith('10.') or ip.startswith('192.168.') or (ip.startswith('172.') and 16<=int(ip.split('.')[1])<=31)),None)
        return private or (candidates[0] if candidates else 'SERVER-IP')
    except Exception:
        return 'SERVER-IP'


def qr_data_uri(text):
    img=qrcode.make(text);buf=io.BytesIO();img.save(buf,format='PNG')
    return 'data:image/png;base64,'+base64.b64encode(buf.getvalue()).decode('ascii')


def current_staff_account(s=None):
    s=s or DB();role=web_session.get('role');uid=web_session.get('user_id')
    if role=='admin':return s.get(Admin,uid) if uid else s.scalar(select(Admin).where(Admin.username==web_session.get('username','')))
    if role=='faculty':return s.get(Faculty,uid) if uid else s.scalar(select(Faculty).where(Faculty.username==web_session.get('username','')))
    return None


def backup_key(password,salt):
    kdf=PBKDF2HMAC(algorithm=hashes.SHA256(),length=32,salt=salt,iterations=BACKUP_KDF_ITERATIONS)
    return base64.urlsafe_b64encode(kdf.derive(password.encode('utf-8')))


def encrypt_backup_bytes(raw,password):
    salt=os.urandom(16);token=Fernet(backup_key(password,salt)).encrypt(raw)
    return b'LWHBK1'+salt+token


def decrypt_backup_bytes(raw,password):
    if not raw.startswith(b'LWHBK1') or len(raw)<23:raise ValueError('This is not a valid encrypted Learn with Hemant backup.')
    salt=raw[6:22]
    try:return Fernet(backup_key(password,salt)).decrypt(raw[22:])
    except InvalidToken as exc:raise ValueError('Backup password is incorrect or the backup is damaged.') from exc

PRELOADED_BANK_FILE=RESOURCE_DIR/'preloaded_question_banks.json'
_PRELOADED_BANK_CACHE=None

def load_preloaded_question_banks():
    global _PRELOADED_BANK_CACHE
    if _PRELOADED_BANK_CACHE is None:
        try:
            payload=json.loads(PRELOADED_BANK_FILE.read_text(encoding='utf-8'))
            packs=payload.get('packs',[]) if isinstance(payload,dict) else []
            _PRELOADED_BANK_CACHE={str(p.get('slug')):p for p in packs if p.get('slug') and p.get('subject')}
        except Exception:
            _PRELOADED_BANK_CACHE={}
    return _PRELOADED_BANK_CACHE

def activate_preloaded_pack(s,pack):
    marker=f"preloaded:{pack['slug']}"
    ensure_subject_catalog_entry(s,pack.get('subject',''),pack.get('category','Engineering'),pack.get('course_semester',''),'preloaded-library')
    existing=set(s.scalars(select(BankQuestion.question).where(BankQuestion.created_by==marker)).all())
    added=0
    for row in pack.get('questions',[]):
        question=(row.get('question') or '').strip()
        if not question or question in existing:
            continue
        ans=(row.get('correct_answer') or 'A').upper()
        if ans not in {'A','B','C','D'}:
            ans='A'
        bq=BankQuestion(
            subject=(row.get('subject') or pack.get('subject') or 'General').strip(),
            course_semester=(row.get('course_semester') or pack.get('course_semester') or '').strip(),
            unit=str(row.get('unit') or '').strip(),
            topic=(row.get('topic') or row.get('unit_title') or '').strip(),
            question_type='single_choice',
            question=question,
            option_a=str(row.get('option_a') or '').strip(),
            option_b=str(row.get('option_b') or '').strip(),
            option_c=str(row.get('option_c') or '').strip(),
            option_d=str(row.get('option_d') or '').strip(),
            correct_answer=ans,answer_key=ans,answer_tolerance='',answer_case_sensitive=False,
            marks=max(1,int(row.get('marks') or 1)),
            difficulty=canonical_difficulty(row.get('difficulty')),
            bloom_level=canonical_bloom(row.get('bloom_level')),
            co_mapping=str(row.get('co_mapping') or '').strip(),
            tags=(str(row.get('tags') or '').strip()+f',preloaded:{pack["slug"]}').strip(','),
            status='approved',
            version=1,
            created_by=marker,
            created_at=now_iso(),
            updated_at=now_iso()
        )
        s.add(bq); added+=1
    return added,len(existing)

def ensure_subject_catalog_entry(s,name,category='Custom / Other',course_semester='',created_by='system',reactivate=True):
    name=(name or '').strip()
    if not name:
        return None
    category=(category or '').strip() or 'Custom / Other'
    course_semester=(course_semester or '').strip()
    row=s.scalar(select(SubjectCatalog).where(func.lower(SubjectCatalog.name)==name.lower()))
    if row:
        # Preserve a faculty-defined category, but upgrade generic placeholders when better metadata arrives.
        if (not row.category or row.category in {'Custom / Other','Imported / Other','General'}) and category not in {'Custom / Other','Imported / Other','General'}:
            row.category=category
        if not row.course_semester and course_semester:
            row.course_semester=course_semester
        if reactivate:
            row.is_active=True
        return row
    row=SubjectCatalog(name=name,category=category,course_semester=course_semester,is_active=True,created_by=created_by or 'system',created_at=now_iso())
    s.add(row)
    return row

def seed_subject_catalog(s):
    # Register all bundled subjects so the faculty form can use a categorized subject selector immediately.
    for pack in load_preloaded_question_banks().values():
        ensure_subject_catalog_entry(s,pack.get('subject',''),pack.get('category','Engineering'),pack.get('course_semester',''),'preloaded-library',False)
    # Preserve subjects already present in older databases even if they were created before this feature existed.
    for subject,course in s.execute(select(BankQuestion.subject,BankQuestion.course_semester).distinct()).all():
        ensure_subject_catalog_entry(s,subject,'Custom / Other',course,'legacy-question-bank',False)
    s.flush()

def subject_catalog_groups(s,include_inactive=False):
    stmt=select(SubjectCatalog)
    if not include_inactive:
        stmt=stmt.where(SubjectCatalog.is_active==True)
    rows=s.scalars(stmt.order_by(SubjectCatalog.category,SubjectCatalog.name)).all()
    grouped=[]
    current=None
    for row in rows:
        if current is None or current['category']!=row.category:
            current={'category':row.category,'subjects':[]}
            grouped.append(current)
        current['subjects'].append(row)
    return grouped

def preloaded_pack_statuses(s):
    output=[]
    for pack in load_preloaded_question_banks().values():
        marker=f"preloaded:{pack['slug']}"
        active=s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.created_by==marker)) or 0
        item=dict(pack)
        item['active_count']=int(active)
        item['is_active']=active>=int(pack.get('question_count') or 0) and int(pack.get('question_count') or 0)>0
        item['is_partial']=active>0 and not item['is_active']
        output.append(item)
    output.sort(key=lambda p:(p.get('category',''),p.get('subject','')))
    return output

def ensure_super_admin_identity(s):
    """Ensure one configured Super Admin login and optionally demote legacy ``admin``.

    When SUPER_ADMIN_USERNAME differs from LEGACY_ADMIN_USERNAME (``admin`` by
    default), the old Admin record is converted into an ordinary Faculty account
    using the *same password hash*. This preserves the legacy username/password
    while moving Super Admin authority to the newly configured username.
    """
    super_admin=s.scalar(select(Admin).where(Admin.username==super_admin_username))
    if not super_admin:
        super_admin=Admin(username=super_admin_username,password_hash=generate_password_hash(admin_password))
        s.add(super_admin)
        s.flush()
    elif admin_password and not check_password_hash(super_admin.password_hash,admin_password):
        super_admin.password_hash=generate_password_hash(admin_password)

    if legacy_admin_username==super_admin_username:
        return

    legacy=s.scalar(select(Admin).where(Admin.username==legacy_admin_username))
    if not legacy:
        return

    faculty=s.scalar(select(Faculty).where(Faculty.username==legacy_admin_username))
    if faculty:
        # The legacy Admin identity wins if the same username already exists in
        # faculty_users, because the user's requirement is to preserve the old
        # Admin credentials while changing only its role.
        faculty.password_hash=legacy.password_hash
        faculty.is_active=True
    else:
        faculty=Faculty(
            username=legacy.username,
            name=legacy.username.replace('.', ' ').replace('_', ' ').strip().title() or 'Faculty',
            password_hash=legacy.password_hash,
            is_active=True,
            created_at=now_iso(),
        )
        s.add(faculty)
        s.flush()

    role_row=s.scalar(select(FacultyRole).where(FacultyRole.faculty_id==faculty.id))
    if role_row:
        role_row.role='faculty'
        role_row.updated_at=now_iso()
    else:
        s.add(FacultyRole(faculty_id=faculty.id,role='faculty',department='',updated_at=now_iso()))

    # Removing the old row is essential: staff login checks Admin before Faculty,
    # so leaving it here would still grant the legacy username Super Admin rights.
    s.delete(legacy)


def init_db():
    Base.metadata.create_all(engine)
    _configure_database_reliability()
    run_schema_upgrades()
    s=DB()
    try:
        seed_subject_catalog(s)
        if APP_MODE=='online' or not OFFLINE_REQUIRE_SETUP:
            ensure_super_admin_identity(s)
            get_institution(s,create=True)
        s.commit()
    except IntegrityError:
        s.rollback()

def needs_offline_setup():
    if APP_MODE!='offline' or not OFFLINE_REQUIRE_SETUP:
        return False
    s=DB()
    return (s.scalar(select(func.count()).select_from(Admin)) or 0)==0

init_db()

@app.teardown_appcontext
def cleanup(_exc=None): DB.remove()

@app.before_request
def csrf_and_session_setup():
    now_ts=int(time.time())
    if web_session.get('role'):
        last=int(web_session.get('_last_activity',now_ts))
        if now_ts-last>SESSION_TIMEOUT_MINUTES*60:
            device_token=web_session.get('_student_device_token');web_session.clear();web_session['_csrf_token']=secrets.token_urlsafe(32)
            if device_token:web_session['_student_device_token']=device_token
            flash('Your session expired after inactivity. Please sign in again.','error')
            if request.endpoint not in {'home','static','health'}:return redirect(url_for('home'))
        else:web_session['_last_activity']=now_ts
    if '_csrf_token' not in web_session: web_session['_csrf_token']=secrets.token_urlsafe(32)
    if request.method in {'POST','PUT','PATCH','DELETE'}:
        supplied=request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not supplied or not secrets.compare_digest(str(supplied),str(web_session.get('_csrf_token',''))):
            abort(400,'Security token validation failed. Refresh the page and try again.')

@app.context_processor
def globals_for_templates():
    practical_code_available=False
    try:
        s=DB();institution=get_institution(s,create=True);staff_role=current_staff_role(s) if web_session.get('role') in {'admin','faculty'} else web_session.get('role','');staff_display_name=current_staff_name(s) if web_session.get('role') in {'admin','faculty'} else ''
        if web_session.get('role')=='student':
            practical_code_available=student_practical_code_available(s,web_session.get('user_id'))
    except Exception:
        institution=None;staff_role=web_session.get('role','');staff_display_name=(web_session.get('username') or '').replace('.',' ').replace('_',' ').strip().title()
    return {'csrf_token':web_session.get('_csrf_token',''),'web_session':web_session,'is_online':APP_MODE=='online','app_version':APP_VERSION,'institution':institution,'staff_role':staff_role,'staff_role_label':ROLE_LABELS.get(staff_role,staff_role.replace('_',' ').title() if staff_role else ''),'staff_display_name':staff_display_name,'edge_package_enabled':len(EXAM_PACKAGE_SIGNING_KEY.encode('utf-8'))>=32,'practical_code_available':practical_code_available}

@app.before_request
def offline_first_run_guard():
    if not needs_offline_setup(): return None
    allowed={'setup_admin','static','health'}
    if request.endpoint not in allowed: return redirect(url_for('setup_admin'))
    return None

@app.after_request
def security_headers(response):
    response.headers.setdefault('X-Content-Type-Options','nosniff')
    # Every page remains non-frameable except the authenticated secure exam
    # document loaded by our own PIN-verification shell.  That single route is
    # same-origin only, so external sites still cannot embed the application.
    secure_exam_frame=bool(request.endpoint=='take_exam' and request.args.get('secure_shell')=='1')
    response.headers['X-Frame-Options']='SAMEORIGIN' if secure_exam_frame else 'DENY'
    response.headers.setdefault('Referrer-Policy','same-origin')
    response.headers.setdefault('Permissions-Policy','camera=(), microphone=(), geolocation=(), payment=(), usb=()')
    if secure_exam_frame:
        response.headers['Content-Security-Policy']="default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; connect-src 'self'; frame-ancestors 'self'; base-uri 'self'; form-action 'self'"
    else:
        response.headers['Content-Security-Policy']="default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; connect-src 'self'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    response.headers.setdefault('Cross-Origin-Opener-Policy','same-origin')
    response.headers.setdefault('Cross-Origin-Resource-Policy','same-origin')
    if APP_MODE=='online' and request.is_secure: response.headers.setdefault('Strict-Transport-Security','max-age=31536000; includeSubDomains')
    response.headers.setdefault('Cache-Control','no-store' if request.path.startswith('/admin') or request.path.startswith('/student') else 'no-cache')
    return response

@app.route('/setup',methods=['GET','POST'])
def setup_admin():
    if not needs_offline_setup(): return redirect(url_for('home'))
    if request.method=='POST':
        username=request.form.get('username','').strip(); password=request.form.get('password',''); confirm=request.form.get('confirm_password','')
        if len(username)<3: flash('Administrator username must contain at least 3 characters.','error')
        elif len(password)<10: flash('Administrator password must contain at least 10 characters.','error')
        elif password!=confirm: flash('Passwords do not match.','error')
        else:
            s=DB()
            try:
                inst=get_institution(s,create=True)
                inst.institution_name=request.form.get('institution_name','').strip() or 'Learn with Hemant'
                inst.short_name=request.form.get('short_name','').strip() or inst.institution_name
                inst.system_name=request.form.get('system_name','').strip() or 'Examination Management System'
                inst.department=request.form.get('department','').strip();inst.academic_year=request.form.get('academic_year','').strip();inst.admin_email=request.form.get('admin_email','').strip();inst.exam_controller=request.form.get('exam_controller','').strip();inst.updated_at=now_iso()
                s.add(Admin(username=username,password_hash=generate_password_hash(password))); s.commit(); flash('Institution and administrator account created. You can now sign in.'); return redirect(url_for('home'))
            except IntegrityError:
                s.rollback(); flash('That administrator username is already in use.','error')
    return render_template('setup.html',login_page=True)

@app.route('/download/offline')
def offline_download():
    parsed=urlparse(OFFLINE_DOWNLOAD_URL)
    if parsed.scheme not in {'https','http'} or not parsed.netloc: abort(503,'Offline download is temporarily unavailable.')
    return redirect(OFFLINE_DOWNLOAD_URL,code=302)

@app.template_filter('dt')
def format_dt(v):
    if not v:return '-'
    try:return display_dt(v).strftime('%d %b %Y, %I:%M %p')
    except Exception:return v

def integration_api_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if len(INTEGRATION_API_KEY)<24:
            return jsonify(error='Integration API is disabled.'),503
        supplied=request.headers.get('Authorization','')
        token=supplied[7:].strip() if supplied.lower().startswith('bearer ') else ''
        if not token or not secrets.compare_digest(token,INTEGRATION_API_KEY):
            return jsonify(error='Unauthorized'),401
        return fn(*a,**kw)
    return inner


def admin_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if web_session.get('role')!='admin': return redirect(url_for('home'))
        return fn(*a,**kw)
    return inner

def staff_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if web_session.get('role') not in {'admin','faculty'}: return redirect(url_for('home'))
        return fn(*a,**kw)
    return inner

def practical_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if web_session.get('role') not in {'admin','faculty'}: return redirect(url_for('home'))
        s=DB()
        if current_staff_role(s) not in PRACTICAL_ROLES:abort(403)
        return fn(*a,**kw)
    return inner

def student_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if web_session.get('role')!='student': return redirect(url_for('home'))
        return fn(*a,**kw)
    return inner

def approver_required(fn):
    @wraps(fn)
    def inner(*a,**kw):
        if web_session.get('role') not in {'admin','faculty'}:return redirect(url_for('home'))
        if not can_approve_exams(DB()):abort(403)
        return fn(*a,**kw)
    return inner

def get_attempt(s,student_id,exam_id): return s.scalar(select(Attempt).where(Attempt.student_id==student_id,Attempt.exam_id==exam_id))

def student_exam_display_title(s, exam):
    """Return the stored exam title, adding a Part suffix for duplicates."""
    raw=(exam.title or '').strip()
    if not raw:
        return 'Exam'
    siblings=s.scalars(select(Exam).where(Exam.title==raw).order_by(Exam.id.asc())).all()
    if len(siblings)<=1:
        return raw
    for idx,row in enumerate(siblings,1):
        if row.id==exam.id:
            return f"{raw} - Part {idx}"
    return raw

def student_grouped_exam_display_title(s, exam, subject):
    """Avoid repeating the subject name inside its own dashboard section."""
    display=student_exam_display_title(s,exam)
    subject=(subject or '').strip()
    if not subject or subject.casefold()=='general':
        return display
    # Example:
    #   Subject header: Mobile Application Development
    #   Stored title:   Mobile Application Development - Unit 1 - Set A
    #   Row title:      Unit 1 - Set A
    pattern=r'^'+re.escape(subject)+r'\s*(?:[-–—:|]\s*)+'
    shortened=re.sub(pattern,'',display,count=1,flags=re.IGNORECASE).strip()
    return shortened or display

def result_performance(score,total_marks):
    total=total_marks or 0
    pct=round(((score or 0)/total)*100) if total else 0
    if pct>=90:
        return pct,'Outstanding','outstanding'
    if pct>=75:
        return pct,'Very Good','very-good'
    if pct>=60:
        return pct,'Good','good'
    if pct>=40:
        return pct,'Average','average'
    return pct,'Poor','poor'

def get_exam_config(s,exam_id,create=False):
    cfg=s.scalar(select(ExamConfig).where(ExamConfig.exam_id==exam_id))
    if not cfg and create:
        cfg=ExamConfig(exam_id=exam_id,question_count=0,pool_size=0,easy_pct=30,medium_pct=50,hard_pct=20,unit_weights='',randomize_questions=True,shuffle_options=True,require_fullscreen=False,tab_switch_limit=3,exam_type='regular',practical_experiment_no='',last_generation_summary='',updated_at=now_iso())
        s.add(cfg); s.flush()
    return cfg

def attempt_question_ids(s,attempt):
    rows=s.scalars(select(AttemptQuestion).where(AttemptQuestion.attempt_id==attempt.id).order_by(AttemptQuestion.position)).all()
    if rows: return [r.question_id for r in rows]
    return [int(x) for x in (attempt.question_order or '').split(',') if x]

def save_answer_record(s,attempt_id,question_id,answer,question=None):
    question=question or s.get(Question,question_id)
    if not question:raise ValueError('Question not found.')
    row=s.scalar(select(Answer).where(Answer.attempt_id==attempt_id,Answer.question_id==question_id))
    raw='' if answer is None else str(answer).strip()
    # Free-response and multi-select controls can legitimately be cleared. Delete the
    # previously autosaved row so a stale answer is never submitted by accident.
    if not raw:
        if row:s.delete(row)
        return
    normalized=normalize_answer(question,raw)
    if not normalized:raise ValueError('Answer format is invalid for this question.')
    legacy=normalized if canonical_question_type(question.question_type)=='single_choice' and normalized in {'A','B','C','D'} else None
    if row:
        row.selected_answer=legacy;row.answer_value=normalized;row.saved_at=now_iso()
    else:
        s.add(Answer(attempt_id=attempt_id,question_id=question_id,selected_answer=legacy,answer_value=normalized,saved_at=now_iso()))

def recalculate_attempt_score(s,attempt,questions=None,saved=None):
    qids=attempt_question_ids(s,attempt)
    questions=questions if questions is not None else (s.scalars(select(Question).where(Question.id.in_(qids))).all() if qids else [])
    saved=saved if saved is not None else s.scalars(select(Answer).where(Answer.attempt_id==attempt.id)).all()
    amap={a.question_id:a for a in saved};score=0;pending=False
    for q in questions:
        answer=amap.get(q.id);value=answer_record_value(answer) if answer else ''
        if canonical_question_type(q.question_type)=='essay':
            if value:
                if answer.manual_score is None:pending=True
                else:score+=max(0,min(int(answer.manual_score),int(q.marks or 0)))
        elif is_answer_correct(q,value):score+=q.marks
    attempt.score=score;attempt.total_marks=sum(q.marks for q in questions);attempt.grading_status='pending' if pending else 'complete'
    return attempt


def finalize_attempt(s,attempt):
    if attempt.status!='submitted':
        recalculate_attempt_score(s,attempt);attempt.status='submitted';attempt.submitted_at=now_iso();s.commit()
    # Practical-viva sync is deliberately best-effort: a practical mapping issue
    # must never break the student's normal exam submission/result flow.
    try:
        result=sync_practical_viva_from_attempt(s,attempt)
        if result.get('updated'):
            s.commit()
    except Exception:
        s.rollback()
        try:app.logger.exception('Practical Exam viva auto-sync failed for attempt %s',attempt.id)
        except Exception:pass
    return attempt

def canonical_difficulty(value):
    v=(value or 'Medium').strip().lower()
    return {'easy':'Easy','medium':'Medium','hard':'Hard'}.get(v,'Medium')

def canonical_bloom(value):
    allowed=['Remember','Understand','Apply','Analyze','Evaluate','Create']
    v=(value or 'Understand').strip().lower()
    for item in allowed:
        if item.lower()==v: return item
    return 'Understand'

def parse_unit_weights(text):
    text=(text or '').strip()
    if not text: return {}
    result={}
    for part in text.replace(';',',').split(','):
        part=part.strip()
        if not part: continue
        if ':' not in part: raise ValueError('Use unit distribution like 1:20, 2:30, 3:50.')
        unit,weight=part.split(':',1)
        unit=unit.strip()
        if unit.lower().startswith('unit '): unit=unit[5:].strip()
        try: weight=int(weight.strip())
        except ValueError as exc: raise ValueError('Unit weights must be whole numbers.') from exc
        if not unit or weight<0: raise ValueError('Unit distribution contains an invalid value.')
        result[unit]=weight
    if result and sum(result.values())<=0: raise ValueError('Unit distribution must contain a positive weight.')
    return result

def allocate_counts(total,weights):
    weights={str(k):max(0,float(v)) for k,v in weights.items() if float(v)>=0}
    if total<=0 or not weights or sum(weights.values())<=0: return {k:0 for k in weights}
    s=sum(weights.values()); raw={k:total*v/s for k,v in weights.items()}; out={k:int(math.floor(v)) for k,v in raw.items()}
    remainder=total-sum(out.values())
    for k,_ in sorted(raw.items(),key=lambda item:item[1]-math.floor(item[1]),reverse=True)[:remainder]: out[k]+=1
    return out

def choose_blueprint_questions(candidates,total,unit_weights,diff_weights):
    pool=list(candidates); random.shuffle(pool); total=min(total,len(pool))
    unit_targets=allocate_counts(total,unit_weights) if unit_weights else {}
    diff_targets=allocate_counts(total,diff_weights)
    selected=[]; unit_counts={k:0 for k in unit_targets}; diff_counts={k:0 for k in diff_targets}
    while len(selected)<total and pool:
        best_score=None; best=[]
        for q in pool:
            unit_key=(q.unit or '').strip()
            diff=canonical_difficulty(q.difficulty)
            unit_need=(unit_targets.get(unit_key,0)-unit_counts.get(unit_key,0)) if unit_targets else 0
            diff_need=diff_targets.get(diff,0)-diff_counts.get(diff,0)
            score=(4 if unit_need>0 else 0)+(3 if diff_need>0 else 0)+random.random()
            if unit_targets and unit_key not in unit_targets: score-=3
            if best_score is None or score>best_score: best_score=score; best=[q]
            elif abs(score-best_score)<1e-9: best.append(q)
        chosen=random.choice(best); pool.remove(chosen); selected.append(chosen)
        if chosen.unit in unit_counts: unit_counts[chosen.unit]+=1
        diff=canonical_difficulty(chosen.difficulty); diff_counts[diff]=diff_counts.get(diff,0)+1
    return selected,unit_targets,diff_targets,unit_counts,diff_counts

def copy_bank_question_to_exam(s,bq,exam_id):
    if s.scalar(select(ExamBankMap).where(ExamBankMap.exam_id==exam_id,ExamBankMap.bank_question_id==bq.id)): return False
    q=Question(exam_id=exam_id,question=bq.question,option_a=bq.option_a,option_b=bq.option_b,option_c=bq.option_c,option_d=bq.option_d,correct_answer=bq.correct_answer,question_type=canonical_question_type(bq.question_type),answer_key=(bq.answer_key or bq.correct_answer),answer_tolerance=bq.answer_tolerance or '',answer_case_sensitive=bool(bq.answer_case_sensitive),marks=bq.marks,practical_experiment_no=(normalize_practical_exam_no(bq.practical_experiment_no) if normalize_practice_visibility(bq.practice_visibility)=='practical_exam' else ''))
    s.add(q); s.flush(); s.add(ExamBankMap(exam_id=exam_id,exam_question_id=q.id,bank_question_id=bq.id)); return True

def sync_manual_exam_question_count(s,exam_id):
    """For manually curated exams, make every question in the pool visible to a new attempt.

    The Blueprint page can still be used afterwards to deliberately choose a smaller
    randomized per-student count. Calling this helper again after a later manual add
    intentionally expands the per-student count to include the newly added question.
    """
    pool_count=s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==exam_id)) or 0
    cfg=get_exam_config(s,exam_id,create=True)
    cfg.pool_size=pool_count
    cfg.question_count=pool_count
    cfg.updated_at=now_iso()
    return pool_count

def normalize_legacy_manual_subject_exam(s,exam_id,cfg=None):
    """Upgrade V50 subject-exam configs to the new manual-pool semantics."""
    cfg=cfg or get_exam_config(s,exam_id,create=False)
    if not cfg or not (cfg.last_generation_summary or '').startswith('Created as a separate exam from '):
        return cfg
    pool_count=s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==exam_id)) or 0
    if cfg.pool_size!=pool_count or cfg.question_count!=pool_count:
        cfg.pool_size=pool_count
        cfg.question_count=pool_count
        cfg.updated_at=now_iso()
        s.commit()
    return cfg

def edge_exam_payload(s,exam):
    cfg=get_exam_config(s,exam.id,create=False);security=get_exam_security_policy(s,exam.id,create=False);questions=s.scalars(select(Question).where(Question.exam_id==exam.id).order_by(Question.id)).all();inst=get_institution(s,create=True)
    return {
        'kind':'exam','schema_version':1,'issued_at':now_iso(),
        'source':{'mode':APP_MODE,'exam_id':exam.id,'institution':inst.short_name if inst else ''},
        'exam':{'title':exam.title,'duration_minutes':exam.duration_minutes,'created_at':exam.created_at},
        'config':({
            'subject':cfg.subject,'course_semester':cfg.course_semester,'question_count':cfg.question_count,'pool_size':cfg.pool_size,
            'easy_pct':cfg.easy_pct,'medium_pct':cfg.medium_pct,'hard_pct':cfg.hard_pct,'unit_weights':cfg.unit_weights,
            'randomize_questions':bool(cfg.randomize_questions),'shuffle_options':bool(cfg.shuffle_options),'require_fullscreen':bool(cfg.require_fullscreen),'tab_switch_limit':cfg.tab_switch_limit
        } if cfg else {}),
        'security':({'require_candidate_checkin':bool(security.require_candidate_checkin),'require_exam_pin':bool(security.require_exam_pin),'heartbeat_seconds':security.heartbeat_seconds} if security else {}),
        'questions':[{'question':q.question,'question_type':canonical_question_type(q.question_type),'option_a':q.option_a,'option_b':q.option_b,'option_c':q.option_c,'option_d':q.option_d,'correct_answer':q.correct_answer,'answer_key':q.answer_key,'answer_tolerance':q.answer_tolerance,'answer_case_sensitive':bool(q.answer_case_sensitive),'marks':q.marks,'practical_experiment_no':q.practical_experiment_no or ''} for q in questions],
    }


def edge_results_payload(s,exam):
    attempts=s.scalars(select(Attempt).where(Attempt.exam_id==exam.id).order_by(Attempt.id)).all();student_ids={a.student_id for a in attempts};students={x.id:x for x in (s.scalars(select(Student).where(Student.id.in_(student_ids))).all() if student_ids else [])};attempt_ids={a.id for a in attempts};answers=s.scalars(select(Answer).where(Answer.attempt_id.in_(attempt_ids))).all() if attempt_ids else [];events=s.scalars(select(IntegrityEvent).where(IntegrityEvent.attempt_id.in_(attempt_ids)).order_by(IntegrityEvent.id)).all() if attempt_ids else [];answers_by={};events_by={}
    for a in answers:answers_by.setdefault(a.attempt_id,[]).append(a)
    for ev in events:events_by.setdefault(ev.attempt_id,[]).append(ev)
    link=s.scalar(select(EdgePackageReceipt).where(EdgePackageReceipt.exam_id==exam.id).order_by(EdgePackageReceipt.id.desc()).limit(1))
    origin={'exam_package_id':link.package_id,'exam_id':link.source_exam_id,'mode':link.source_mode} if link else {}
    return {'kind':'results','schema_version':1,'issued_at':now_iso(),'source':{'mode':APP_MODE,'exam_id':exam.id},'origin':origin,'exam':{'id':exam.id,'title':exam.title},'attempts':[{'attempt_id':a.id,'roll_no':students.get(a.student_id).roll_no if students.get(a.student_id) else str(a.student_id),'name':students.get(a.student_id).name if students.get(a.student_id) else '','status':a.status,'grading_status':a.grading_status,'score':a.score,'total_marks':a.total_marks,'started_at':a.started_at,'submitted_at':a.submitted_at,'answers':[{'question_id':x.question_id,'answer_value':answer_record_value(x),'manual_score':x.manual_score,'grader_comment':x.grader_comment,'graded_by':x.graded_by,'graded_at':x.graded_at} for x in answers_by.get(a.id,[])],'integrity':[{'event_type':e.event_type,'details':e.details,'created_at':e.created_at} for e in events_by.get(a.id,[])]} for a in attempts]}


def _edge_filename(text):
    cleaned=''.join(ch if ch.isalnum() else '_' for ch in str(text or '')).strip('_')
    return cleaned[:80] or 'exam'

def _cell_text(value):
    if value is None:return ''
    if isinstance(value,float) and value.is_integer():return str(int(value))
    return str(value).strip()

def _canonical_header(value):
    key=_cell_text(value).lower().replace('-','_').replace(' ','_')
    while '__' in key:key=key.replace('__','_')
    aliases={
        'roll':'roll_no','rollno':'roll_no','roll_number':'roll_no','student_id':'roll_no','registration_no':'registration_no','registration_number':'registration_no','enrollment_no':'registration_no','enrollment_number':'registration_no',
        'student_name':'name','full_name':'name','candidate_name':'name','pass':'password','login_password':'password','student_password':'password',
        'semester':'course_semester','course':'course_semester','co':'co_mapping','bloom':'bloom_level','answer':'correct_answer','correct':'correct_answer','type':'question_type'
    }
    return aliases.get(key,key)

def _rows_from_upload(upload):
    filename=(upload.filename or '').lower(); raw=upload.read()
    if not raw: raise ValueError('The uploaded file is empty.')
    if filename.endswith('.csv'):
        try:text=raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:raise ValueError('CSV must be UTF-8 encoded.') from exc
        all_rows=list(csv.reader(io.StringIO(text)))
    elif filename.endswith('.xlsx'):
        try:
            wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True); ws=wb.active; all_rows=[list(r) for r in ws.iter_rows(values_only=True)]
        except Exception as exc:raise ValueError('The Excel file could not be read. Please upload a valid .xlsx file.') from exc
    else: raise ValueError('Please upload a CSV or Excel (.xlsx) file.')
    if not all_rows:raise ValueError('The uploaded file has no rows.')
    headers=[_canonical_header(v) for v in all_rows[0]]; rows=[]
    for row_number,row in enumerate(all_rows[1:],start=2):
        values={headers[i]:_cell_text(row[i] if i<len(row) else '') for i in range(len(headers)) if headers[i]}
        if any(values.values()): values['_row_number']=row_number; rows.append(values)
    return headers,rows

def complete_staff_session(s,row,role):
    csrf=web_session.get('_csrf_token');web_session.clear();web_session['_csrf_token']=csrf
    web_session.update(role=role,user_id=row.id,username=row.username,display_name=((getattr(row,'name','') or row.username).strip()),_last_activity=int(time.time()))
    audit_event(s,'staff_login','user',row.id,role);s.commit()
    return redirect(url_for('admin_dashboard'))


def begin_staff_mfa(row,role):
    csrf=web_session.get('_csrf_token');web_session.clear();web_session['_csrf_token']=csrf
    web_session.update(_mfa_pending_role=role,_mfa_pending_user_id=row.id,_mfa_pending_username=row.username,_mfa_pending_expires=int(time.time())+300,_mfa_failures=0)
    return redirect(url_for('mfa_verify'))


@app.route('/',methods=['GET','POST'])
def home():
    if web_session.get('role') in {'admin','faculty'}: return redirect(url_for('admin_dashboard'))
    if web_session.get('role')=='student': return redirect(url_for('student_dashboard'))
    if request.method=='POST':
        s=DB();typ=request.form.get('login_type');kind='staff' if typ=='admin' else 'student'
        username=(request.form.get('username','') if typ=='admin' else request.form.get('roll_no','')).strip()
        password=request.form.get('password','')
        throttle_key=_request_identity_key(kind,username.casefold())
        locked,remaining=auth_is_locked(s,throttle_key)
        if locked:
            flash(f'Too many failed sign-in attempts. Try again in {max(1,math.ceil(remaining/60))} minute(s).','error')
            return render_template('login.html',login_page=True),429
        row=None;role='student'
        if typ=='admin':
            row=s.scalar(select(Admin).where(Admin.username==username));role='admin'
            if not row:row=s.scalar(select(Faculty).where(Faculty.username==username,Faculty.is_active==True));role='faculty'
        else:
            row=s.scalar(select(Student).where(Student.roll_no==username))
        valid=bool(row and check_password_hash(row.password_hash,password))
        if valid:
            clear_auth_failures(s,throttle_key)
            if role=='student':
                csrf=web_session.get('_csrf_token');device_token=web_session.get('_student_device_token') or secrets.token_urlsafe(24);web_session.clear();web_session['_csrf_token']=csrf;web_session['_student_device_token']=device_token
                web_session.update(role='student',user_id=row.id,username=row.roll_no,_last_activity=int(time.time()))
                audit_event(s,'student_login','user',row.id,'student')
                s.commit();return redirect(url_for('student_dashboard'))
            if bool(getattr(row,'mfa_enabled',False)) and (getattr(row,'mfa_secret','') or '').strip():
                return begin_staff_mfa(row,role)
            return complete_staff_session(s,row,role)
        record_auth_failure(s,throttle_key)
        flash('Invalid login credentials.','error')
    return render_template('login.html',login_page=True)

@app.route('/mfa-verify',methods=['GET','POST'])
def mfa_verify():
    role=web_session.get('_mfa_pending_role');uid=web_session.get('_mfa_pending_user_id');expires=int(web_session.get('_mfa_pending_expires') or 0)
    if role not in {'admin','faculty'} or not uid or expires<int(time.time()):
        for key in ['_mfa_pending_role','_mfa_pending_user_id','_mfa_pending_username','_mfa_pending_expires','_mfa_failures']:web_session.pop(key,None)
        flash('Your verification session expired. Sign in again.','error');return redirect(url_for('home'))
    s=DB();row=s.get(Admin,uid) if role=='admin' else s.get(Faculty,uid)
    if not row or not bool(getattr(row,'mfa_enabled',False)) or not (getattr(row,'mfa_secret','') or '').strip():
        flash('Multi-factor authentication is not available for this account.','error');return redirect(url_for('home'))
    if request.method=='POST':
        code=request.form.get('code','')
        if verify_totp(row.mfa_secret,code):
            return complete_staff_session(s,row,role)
        failures=int(web_session.get('_mfa_failures') or 0)+1;web_session['_mfa_failures']=failures
        audit_event(s,'staff_mfa_failed','user',row.id,role);s.commit()
        if failures>=5:
            csrf=web_session.get('_csrf_token');web_session.clear();web_session['_csrf_token']=csrf
            flash('Too many invalid verification codes. Sign in again.','error');return redirect(url_for('home'))
        flash('Invalid authentication code.','error')
    return render_template('mfa_verify.html',login_page=True,username=getattr(row,'username',''))


@app.route('/admin/security/mfa',methods=['GET','POST'])
@staff_required
def staff_mfa_settings():
    s=DB();account=current_staff_account(s)
    if not account:abort(404)
    enabled=bool(getattr(account,'mfa_enabled',False) and (getattr(account,'mfa_secret','') or '').strip())
    if request.method=='POST':
        action=request.form.get('action','')
        password=request.form.get('password','')
        if not check_password_hash(account.password_hash,password):
            flash('Current password is incorrect.','error');return redirect(url_for('staff_mfa_settings'))
        if action=='enable':
            secret_value=(web_session.get('_mfa_enroll_secret') or '').strip();code=request.form.get('code','')
            if not secret_value or not verify_totp(secret_value,code):
                flash('Enter the current 6-digit code from your authenticator app.','error');return redirect(url_for('staff_mfa_settings'))
            account.mfa_secret=secret_value;account.mfa_enabled=True;web_session.pop('_mfa_enroll_secret',None);audit_event(s,'staff_mfa_enabled','user',account.id,current_staff_role(s));s.commit();flash('Multi-factor authentication is now enabled.');return redirect(url_for('staff_mfa_settings'))
        if action=='disable':
            code=request.form.get('code','')
            if not enabled or not verify_totp(account.mfa_secret,code):
                flash('Enter a valid authenticator code to disable MFA.','error');return redirect(url_for('staff_mfa_settings'))
            account.mfa_enabled=False;account.mfa_secret='';web_session.pop('_mfa_enroll_secret',None);audit_event(s,'staff_mfa_disabled','user',account.id,current_staff_role(s));s.commit();flash('Multi-factor authentication has been disabled.');return redirect(url_for('staff_mfa_settings'))
        abort(400)
    secret_value=''
    qr_uri=''
    if not enabled:
        secret_value=(web_session.get('_mfa_enroll_secret') or '').strip()
        if not secret_value:
            secret_value=generate_totp_secret();web_session['_mfa_enroll_secret']=secret_value
        qr_uri=qr_data_uri(totp_uri(secret_value,account.username))
    return render_template('mfa_settings.html',enabled=enabled,secret_value=secret_value,qr_uri=qr_uri,account=account)


@app.route('/logout')
def logout():
    device_token=web_session.get('_student_device_token');web_session.clear()
    if device_token:web_session['_student_device_token']=device_token
    return redirect(url_for('home'))

@app.route('/health')
def health():
    started=time.perf_counter()
    try:
        s=DB();s.execute(select(1));latency_ms=round((time.perf_counter()-started)*1000,1)
        active=s.scalar(select(func.count()).select_from(Exam).where(Exam.is_active==True)) or 0
        in_progress=s.scalar(select(func.count()).select_from(Attempt).where(Attempt.status=='in_progress')) or 0
        return jsonify(status='ok',version=APP_VERSION,mode=APP_MODE,database='postgresql' if DATABASE_URL.startswith('postgresql') else 'sqlite',db_latency_ms=latency_ms,active_exams=active,in_progress_attempts=in_progress)
    except Exception as exc:return jsonify(status='error',version=APP_VERSION,detail=type(exc).__name__),503

@app.route('/api/v1/exams')
@integration_api_required
def integration_exams():
    s=DB();rows=s.scalars(select(Exam).order_by(Exam.id.desc())).all()
    return jsonify(exams=[{'id':e.id,'title':e.title,'duration_minutes':e.duration_minutes,'is_active':bool(e.is_active),'created_at':e.created_at} for e in rows])


@app.route('/api/v1/results')
@integration_api_required
def integration_results():
    exam_id=request.args.get('exam_id',type=int);s=DB();rows=result_rows(s,exam_id)
    return jsonify(results=[{'attempt_id':r.attempt_id,'roll_no':r.roll_no,'name':r.name,'batch_section':r.group_label,'exam_id':r.exam_id,'exam':r.title,'status':r.status,'grading_status':r.grading_status,'score':r.score,'total_marks':r.total_marks,'percentage':r.percentage if r.status=='submitted' else None,'grade':r.grade if r.status=='submitted' else None,'integrity_events':r.violations,'started_at':r.started_at,'submitted_at':r.submitted_at} for r in rows])


@app.route('/admin')
@staff_required
def admin_dashboard():
    s=DB(); stats={
        'students':s.scalar(select(func.count()).select_from(Student)) or 0,
        'exams':s.scalar(select(func.count()).select_from(Exam)) or 0,
        'questions':s.scalar(select(func.count()).select_from(Question)) or 0,
        'bank_questions':s.scalar(select(func.count()).select_from(BankQuestion)) or 0,
        'attempts':s.scalar(select(func.count()).select_from(Attempt)) or 0,
        'approved':s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.status=='approved')) or 0
    }
    if current_staff_role(s) in PRACTICAL_ROLES:
        stats['practical_registers']=len(s.scalars(practical_register_stmt(s)).all())
    else:
        stats['practical_registers']=0
    return render_template('admin_dashboard.html',stats=stats)

def practical_marks_maxima(register):
    return {
        'attendance': max(0, int(getattr(register, 'attendance_max_marks', 5) or 0)),
        'record': max(0, int(getattr(register, 'record_max_marks', 5) or 0)),
        'performance': max(0, int(getattr(register, 'performance_max_marks', 10) or 0)),
        'viva': max(0, int(getattr(register, 'viva_max_marks', 10) or 0)),
    }


def practical_total_max(register):
    return sum(practical_marks_maxima(register).values())


def repair_practical_experiment_numbers(s, register):
    """Normalize legacy practical labels (e.g. B-7/C-4) to 12-B/12-C.

    The repair uses experiment order only; marks stay attached to the same
    PracticalExperiment IDs, so existing evaluation data is not moved or lost.
    """
    rows=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id).order_by(PracticalExperiment.sort_order,PracticalExperiment.id)).all()
    if not rows:
        return 0
    targets=normalize_experiment_sequence([row.experiment_no for row in rows])
    if len({code.casefold() for code in targets}) != len(targets):
        return 0
    changed=[(row,target) for row,target in zip(rows,targets) if row.experiment_no!=target]
    if not changed:
        return 0
    # Two-phase rename avoids transient unique-constraint collisions.
    for row,_ in changed:
        row.experiment_no=f'__tmp_practical_{row.id}'
    s.flush()
    for row,target in changed:
        row.experiment_no=target
    register.updated_at=now_iso()
    audit_event(s,'practical_experiment_numbers_normalized','practical_register',register.id,f'updated={len(changed)}')
    s.commit()
    return len(changed)


def _component_mark(value, label, maximum):
    if value in {None, ''}:
        return None
    try:
        mark=float(value)
    except (TypeError, ValueError):
        raise ValueError(f'{label} marks must be numeric.')
    if mark < 0 or mark > maximum:
        raise ValueError(f'{label} marks must be between 0 and {maximum}.')
    return mark


@app.route('/admin/practicals',methods=['GET','POST'])
@practical_required
def practical_registers():
    s=DB();owner_type,owner_id,owner_name=current_practical_owner(s)
    if request.method=='POST':
        title=(request.form.get('title') or '').strip();subject=(request.form.get('subject') or '').strip();section=(request.form.get('section') or '').strip();academic_year=(request.form.get('academic_year') or '').strip();lab_code=(request.form.get('lab_code') or '').strip()
        attendance_max=5;record_max=5;performance_max=10;viva_max=10;default_marks=attendance_max+record_max+performance_max+viva_max
        if not title:title=subject or 'Practical Register'
        if not subject:
            flash('Subject / lab name is required.','error');return redirect(url_for('practical_registers'))
        row=PracticalRegister(owner_type=owner_type,owner_id=owner_id,owner_name=owner_name,title=title,subject=subject,lab_code=lab_code,section=section,academic_year=academic_year,default_max_marks=default_marks,attendance_max_marks=attendance_max,record_max_marks=record_max,performance_max_marks=performance_max,viva_max_marks=viva_max,created_at=now_iso(),updated_at=now_iso());s.add(row);s.flush()
        roster=request.files.get('roster_file');added=0
        if roster and roster.filename:
            try:
                imported=parse_roster_bytes(roster.filename,roster.read())
                for seq,item in enumerate(imported,start=1):
                    s.add(PracticalStudent(register_id=row.id,sequence=seq,roll_no=item['roll_no'],name=item['name'],created_at=now_iso()));added+=1
            except ValueError as exc:
                s.rollback();flash(str(exc),'error');return redirect(url_for('practical_registers'))
        audit_event(s,'practical_register_created','practical_register',row.id,f'subject={subject}, students={added}');s.commit();flash(f'Practical register created'+(f' with {added} students.' if added else '.'))
        return redirect(url_for('practical_register_detail',register_id=row.id))
    rows=s.scalars(practical_register_stmt(s)).all();counts={}
    for row in rows:
        counts[row.id]={
            'students':s.scalar(select(func.count()).select_from(PracticalStudent).where(PracticalStudent.register_id==row.id)) or 0,
            'experiments':s.scalar(select(func.count()).select_from(PracticalExperiment).where(PracticalExperiment.register_id==row.id)) or 0,
            'marks':s.scalar(select(func.count()).select_from(PracticalMark).where(PracticalMark.register_id==row.id,PracticalMark.marks.is_not(None))) or 0,
        }
    return render_template('practical_registers.html',registers=rows,counts=counts,viewer_role=current_staff_role(s))


@app.route('/admin/practicals/<int:register_id>/delete',methods=['POST'])
@practical_required
def delete_practical_register(register_id):
    s=DB();register=practical_register_access(s,register_id)
    label=register.title or register.section or register.subject or f'Practical Register {register.id}'
    student_count=s.scalar(select(func.count()).select_from(PracticalStudent).where(PracticalStudent.register_id==register.id)) or 0
    experiment_count=s.scalar(select(func.count()).select_from(PracticalExperiment).where(PracticalExperiment.register_id==register.id)) or 0
    mark_count=s.scalar(select(func.count()).select_from(PracticalMark).where(PracticalMark.register_id==register.id)) or 0
    audit_event(
        s,'practical_register_deleted','practical_register',register.id,
        f'title={label}, students={student_count}, experiments={experiment_count}, mark_rows={mark_count}'
    )
    # Delete dependent practical data explicitly so this works consistently on
    # both PostgreSQL production databases and SQLite/LAN deployments.
    s.execute(delete(PracticalCodeSubmission).where(PracticalCodeSubmission.register_id==register.id))
    s.execute(delete(PracticalMark).where(PracticalMark.register_id==register.id))
    s.execute(delete(PracticalStudent).where(PracticalStudent.register_id==register.id))
    s.execute(delete(PracticalExperiment).where(PracticalExperiment.register_id==register.id))
    s.delete(register);s.commit()
    flash(f'{label} and all of its practical data were deleted.')
    return redirect(url_for('practical_registers'))


@app.route('/admin/practicals/<int:register_id>')
@practical_required
def practical_register_detail(register_id):
    s=DB();register=practical_register_access(s,register_id);repair_practical_experiment_numbers(s,register)
    students=s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id).order_by(PracticalStudent.sequence,PracticalStudent.roll_no)).all()
    experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id).order_by(PracticalExperiment.sort_order,PracticalExperiment.id)).all()
    marks=s.scalars(select(PracticalMark).where(PracticalMark.register_id==register.id)).all();by_student={}
    for mark in marks:by_student.setdefault(mark.practical_student_id,[]).append(mark)
    possible=sum(e.max_marks for e in experiments);summary=[]
    for st in students:
        student_marks=by_student.get(st.id,[]);scored=sum(float(m.marks or 0) for m in student_marks if m.marks is not None);evaluated=sum(1 for m in student_marks if m.marks is not None or m.attendance=='A')
        summary.append({'student':st,'scored':round(scored,2),'possible':possible,'evaluated':evaluated,'percent':round(scored*100/possible,1) if possible else 0})
    return render_template('practical_register_detail.html',register=register,students=students,experiments=experiments,summary=summary,component_maxima=practical_marks_maxima(register),total_max=practical_total_max(register))


@app.route('/admin/practicals/<int:register_id>/marks-settings',methods=['POST'])
@practical_required
def practical_marks_settings(register_id):
    s=DB();register=practical_register_access(s,register_id)
    values={}
    labels=(('attendance','Attendance'),('record','Record'),('performance','Performance'),('viva','Viva'))
    try:
        for key,label in labels:
            raw=(request.form.get(f'{key}_max_marks') or '').strip()
            value=int(raw)
            if value < 0 or value > 100:raise ValueError(f'{label} maximum must be between 0 and 100.')
            values[key]=value
    except (TypeError,ValueError) as exc:
        flash(str(exc) if str(exc) else 'Enter valid maximum marks.','error');return redirect(url_for('practical_register_detail',register_id=register.id))
    total=sum(values.values())
    if total <= 0:
        flash('The practical total must be greater than zero.','error');return redirect(url_for('practical_register_detail',register_id=register.id))
    component_columns=(('attendance',PracticalMark.attendance_marks,'Attendance'),('record',PracticalMark.record_marks,'Record'),('performance',PracticalMark.performance_marks,'Performance'),('viva',PracticalMark.viva_marks,'Viva'))
    for key,column,label in component_columns:
        highest=s.scalar(select(func.max(column)).where(PracticalMark.register_id==register.id))
        if highest is not None and float(highest)>values[key]:
            flash(f'{label} maximum cannot be reduced below an existing mark of {highest:g}.','error');return redirect(url_for('practical_register_detail',register_id=register.id))
    register.attendance_max_marks=values['attendance'];register.record_max_marks=values['record'];register.performance_max_marks=values['performance'];register.viva_max_marks=values['viva'];register.default_max_marks=total;register.updated_at=now_iso()
    experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id)).all()
    for experiment in experiments:experiment.max_marks=total
    audit_event(s,'practical_marks_settings_updated','practical_register',register.id,f"attendance={values['attendance']}, record={values['record']}, performance={values['performance']}, viva={values['viva']}, total={total}")
    s.commit();flash(f'Practical marks distribution updated. Total: {total}.')
    return redirect(url_for('practical_register_detail',register_id=register.id))


@app.route('/admin/practicals/<int:register_id>/students/import',methods=['POST'])
@practical_required
def practical_students_import(register_id):
    s=DB();register=practical_register_access(s,register_id);upload=request.files.get('roster_file')
    if not upload or not upload.filename:flash('Choose a CSV or Excel student sheet.','error');return redirect(url_for('practical_register_detail',register_id=register.id))
    try:rows=parse_roster_bytes(upload.filename,upload.read())
    except ValueError as exc:flash(str(exc),'error');return redirect(url_for('practical_register_detail',register_id=register.id))
    existing={x.roll_no.casefold():x for x in s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id)).all()};added=updated=0;seq=max([x.sequence for x in existing.values()] or [0])
    for item in rows:
        key=item['roll_no'].casefold();current=existing.get(key)
        if current:
            if current.name!=item['name']:current.name=item['name'];updated+=1
        else:
            seq+=1;s.add(PracticalStudent(register_id=register.id,sequence=seq,roll_no=item['roll_no'],name=item['name'],created_at=now_iso()));added+=1
    register.updated_at=now_iso();audit_event(s,'practical_students_imported','practical_register',register.id,f'added={added}, updated={updated}');s.commit();flash(f'Student sheet processed: {added} added, {updated} updated.');return redirect(url_for('practical_register_detail',register_id=register.id))


@app.route('/admin/practicals/<int:register_id>/experiments/import',methods=['POST'])
@practical_required
def practical_experiments_import(register_id):
    s=DB();register=practical_register_access(s,register_id);upload=request.files.get('experiment_file');pasted=(request.form.get('experiment_text') or '').strip()
    try:
        if upload and upload.filename:rows=parse_experiment_bytes(upload.filename,upload.read(),register.default_max_marks)
        elif pasted:rows=parse_experiment_text(pasted,register.default_max_marks)
        else:raise ValueError('Upload an experiment file or paste an experiment list.')
    except ValueError as exc:flash(str(exc),'error');return redirect(url_for('practical_register_detail',register_id=register.id))
    repair_practical_experiment_numbers(s,register)
    existing_rows=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id).order_by(PracticalExperiment.sort_order,PracticalExperiment.id)).all()
    by_code={x.experiment_no.casefold():x for x in existing_rows};by_title={x.title.casefold():x for x in existing_rows};added=updated=0;order=max([x.sort_order for x in existing_rows] or [0])
    for item in rows:
        key=item['experiment_no'].casefold();title_key=item['title'].casefold();current=by_code.get(key) or by_title.get(title_key)
        total_max=practical_total_max(register)
        if current:
            # If the same experiment title was previously imported with a bad
            # legacy label, correct only its label; its ID/marks remain intact.
            if current.experiment_no.casefold()!=key and key not in by_code:
                old_key=current.experiment_no.casefold();current.experiment_no=item['experiment_no'];by_code.pop(old_key,None);by_code[key]=current
            current.title=item['title'];current.max_marks=total_max
            # Re-uploading the experiment sheet may add/update a reference program.
            # A blank reference_code never erases an already configured solution.
            if (item.get('reference_code') or '').strip():
                current.reference_code=(item.get('reference_code') or '').strip()
            updated+=1
        else:
            order+=1;current=PracticalExperiment(register_id=register.id,experiment_no=item['experiment_no'],title=item['title'],reference_code=(item.get('reference_code') or '').strip(),max_marks=total_max,sort_order=order,created_at=now_iso());s.add(current);s.flush();by_code[key]=current;by_title[title_key]=current;added+=1
    register.updated_at=now_iso();audit_event(s,'practical_experiments_imported','practical_register',register.id,f'added={added}, updated={updated}');s.commit();flash(f'Experiment list processed: {added} added, {updated} updated.');return redirect(url_for('practical_register_detail',register_id=register.id))


@app.route('/admin/practicals/<int:register_id>/experiment/<int:experiment_id>/reference-code',methods=['GET','POST'])
@practical_required
def practical_reference_code(register_id,experiment_id):
    s=DB();register=practical_register_access(s,register_id);experiment=s.get(PracticalExperiment,experiment_id)
    if not experiment or experiment.register_id!=register.id:abort(404)
    if request.method=='POST':
        upload=request.files.get('code_file');code=(request.form.get('reference_code') or '')
        penalty_rules=(request.form.get('penalty_rules') or '')
        if upload and upload.filename:
            raw=upload.read()
            if len(raw)>512*1024:
                flash('Reference code file is too large. Keep it below 512 KB.','error');return redirect(url_for('practical_reference_code',register_id=register.id,experiment_id=experiment.id))
            try:code=raw.decode('utf-8-sig')
            except UnicodeDecodeError:
                flash('Reference code file must be UTF-8 text.','error');return redirect(url_for('practical_reference_code',register_id=register.id,experiment_id=experiment.id))
        code=code.strip()
        if len(code)>250000:
            flash('Reference code is too large.','error');return redirect(url_for('practical_reference_code',register_id=register.id,experiment_id=experiment.id))
        if len(penalty_rules)>20000:
            flash('Penalty rules are too large.','error');return redirect(url_for('practical_reference_code',register_id=register.id,experiment_id=experiment.id))
        # Keep one non-empty rule per line and remove exact duplicates while
        # preserving faculty order.  Rules remain private on the admin side.
        cleaned_rules=[];seen_rules=set()
        for rule in penalty_rules.replace('\r\n','\n').replace('\r','\n').split('\n'):
            rule=rule.strip()
            if rule and rule not in seen_rules:
                cleaned_rules.append(rule);seen_rules.add(rule)
        experiment.reference_code=code
        experiment.penalty_rules='\n'.join(cleaned_rules)
        register.updated_at=now_iso()
        audit_event(s,'practical_reference_code_updated','practical_experiment',experiment.id,f'register={register.id}, experiment={experiment.experiment_no}, configured={bool(code)}, penalty_rules={len(cleaned_rules)}')
        s.commit();flash('Reference code saved.' if code else 'Reference code cleared.')
        return redirect(url_for('practical_register_detail',register_id=register.id))
    return render_template('practical_reference_code.html',register=register,experiment=experiment)


@app.route('/admin/practicals/<int:register_id>/mark-entry')
@practical_required
def practical_mark_entry(register_id):
    s=DB();register=practical_register_access(s,register_id);repair_practical_experiment_numbers(s,register);experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id).order_by(PracticalExperiment.sort_order,PracticalExperiment.id)).all();students=s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id).order_by(PracticalStudent.sequence,PracticalStudent.roll_no)).all()
    if not experiments:
        flash('Upload the experiment list before entering marks.','error');return redirect(url_for('practical_register_detail',register_id=register.id))
    experiment_id=request.args.get('experiment_id',type=int) or experiments[0].id;experiment=next((x for x in experiments if x.id==experiment_id),experiments[0]);marks=s.scalars(select(PracticalMark).where(PracticalMark.register_id==register.id,PracticalMark.practical_experiment_id==experiment.id)).all();mark_map={m.practical_student_id:m for m in marks};evaluated=sum(1 for m in marks if m.marks is not None or m.attendance=='A')
    return render_template('practical_mark_entry.html',register=register,students=students,experiments=experiments,experiment=experiment,mark_map=mark_map,evaluated=evaluated,component_maxima=practical_marks_maxima(register),total_max=practical_total_max(register))


def _save_practical_mark(s,register,student_id,experiment_id,attendance,attendance_marks_value='',record_marks_value='',performance_marks_value='',viva_marks_value='',remarks=''):
    student=s.get(PracticalStudent,student_id);experiment=s.get(PracticalExperiment,experiment_id)
    if not student or not experiment or student.register_id!=register.id or experiment.register_id!=register.id:raise ValueError('Student/experiment does not belong to this practical register.')
    attendance=(attendance or '').strip().upper()
    if attendance not in {'','P','A'}:attendance=''
    maxima=practical_marks_maxima(register)
    component_values={
        'attendance': _component_mark(attendance_marks_value,'Attendance',maxima['attendance']),
        'record': _component_mark(record_marks_value,'Record',maxima['record']),
        'performance': _component_mark(performance_marks_value,'Performance',maxima['performance']),
        'viva': _component_mark(viva_marks_value,'Viva',maxima['viva']),
    }
    # Attendance marks follow the attendance status automatically. The configured
    # Attendance Max is awarded for Present; Absent receives zero for the practical.
    # Untouched rows (no attendance and no component marks) remain blank.
    non_attendance_component=any(component_values[key] is not None for key in ('record','performance','viva'))
    if not attendance and non_attendance_component:
        attendance='P'
    if attendance=='P':
        component_values['attendance']=float(maxima['attendance'])
    elif attendance=='A':
        component_values={'attendance':0.0,'record':None,'performance':None,'viva':None}
    has_component=any(value is not None for value in component_values.values())
    row=s.scalar(select(PracticalMark).where(PracticalMark.practical_student_id==student.id,PracticalMark.practical_experiment_id==experiment.id))
    legacy_total=(row.marks if row and row.marks is not None and all(getattr(row,name,None) is None for name in ('attendance_marks','record_marks','performance_marks','viva_marks')) else None)
    total=(sum(value or 0 for value in component_values.values()) if has_component else legacy_total)
    if attendance=='A':total=0.0
    remarks_value=(remarks or '').strip()[:500]
    # A faculty save with a Performance value is treated as the manual review
    # decision for an anti-copy flag, so the temporary Review marker is cleared.
    if component_values['performance'] is not None and '[CODE REVIEW]' in remarks_value:
        remarks_value=remarks_value.split('[CODE REVIEW]',1)[0].rstrip()[:500]
    if not row:
        row=PracticalMark(register_id=register.id,practical_student_id=student.id,practical_experiment_id=experiment.id,attendance=attendance,attendance_marks=component_values['attendance'],record_marks=component_values['record'],performance_marks=component_values['performance'],viva_marks=component_values['viva'],marks=total,remarks=remarks_value,updated_by=actor_label(s),updated_at=now_iso());s.add(row)
    else:
        row.attendance=attendance;row.attendance_marks=component_values['attendance'];row.record_marks=component_values['record'];row.performance_marks=component_values['performance'];row.viva_marks=component_values['viva'];row.marks=total;row.remarks=remarks_value;row.updated_by=actor_label(s);row.updated_at=now_iso()
    experiment.max_marks=practical_total_max(register)
    return row


@app.route('/admin/practicals/<int:register_id>/marks/save',methods=['POST'])
@practical_required
def practical_mark_save(register_id):
    s=DB();register=practical_register_access(s,register_id);payload=request.get_json(silent=True) or request.form
    try:
        student_id=int(payload.get('student_id') or 0);experiment_id=int(payload.get('experiment_id') or 0)
        row=_save_practical_mark(s,register,student_id,experiment_id,payload.get('attendance',''),payload.get('attendance_marks',''),payload.get('record_marks',''),payload.get('performance_marks',''),payload.get('viva_marks',''),payload.get('remarks',''));register.updated_at=now_iso();s.commit()
        return jsonify(ok=True,attendance=row.attendance,attendance_marks=row.attendance_marks,record_marks=row.record_marks,performance_marks=row.performance_marks,viva_marks=row.viva_marks,total=row.marks,updated_at=row.updated_at)
    except ValueError as exc:s.rollback();return jsonify(ok=False,error=str(exc)),400


@app.route('/admin/practicals/<int:register_id>/marks/bulk',methods=['POST'])
@practical_required
def practical_mark_bulk(register_id):
    s=DB();register=practical_register_access(s,register_id)
    try:experiment_id=int(request.form.get('experiment_id') or 0)
    except ValueError:experiment_id=0
    students=s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id)).all();errors=[];saved=0
    for st in students:
        attendance=request.form.get(f'attendance_{st.id}','');attendance_marks=request.form.get(f'attendance_marks_{st.id}','');record_marks=request.form.get(f'record_marks_{st.id}','');performance_marks=request.form.get(f'performance_marks_{st.id}','');viva_marks=request.form.get(f'viva_marks_{st.id}','');remarks=request.form.get(f'remarks_{st.id}','')
        if attendance or attendance_marks or record_marks or performance_marks or viva_marks or remarks:
            try:_save_practical_mark(s,register,st.id,experiment_id,attendance,attendance_marks,record_marks,performance_marks,viva_marks,remarks);saved+=1
            except ValueError as exc:errors.append(f'{st.roll_no}: {exc}')
    if errors:s.rollback();flash(' '.join(errors[:5]),'error')
    else:register.updated_at=now_iso();audit_event(s,'practical_marks_bulk_saved','practical_register',register.id,f'experiment={experiment_id}, rows={saved}');s.commit();flash(f'Saved {saved} practical mark row(s).')
    return redirect(url_for('practical_mark_entry',register_id=register.id,experiment_id=experiment_id))


@app.route('/admin/practicals/<int:register_id>/marks/all-present',methods=['POST'])
@practical_required
def practical_mark_all_present(register_id):
    s=DB();register=practical_register_access(s,register_id)
    try:experiment_id=int(request.form.get('experiment_id') or 0)
    except ValueError:experiment_id=0
    experiment=s.get(PracticalExperiment,experiment_id)
    if not experiment or experiment.register_id!=register.id:abort(404)
    students=s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id)).all();attendance_max=float(practical_marks_maxima(register)['attendance'])
    for st in students:
        existing=s.scalar(select(PracticalMark).where(PracticalMark.practical_student_id==st.id,PracticalMark.practical_experiment_id==experiment.id))
        if not existing:
            s.add(PracticalMark(register_id=register.id,practical_student_id=st.id,practical_experiment_id=experiment.id,attendance='P',attendance_marks=attendance_max,marks=attendance_max,remarks='',updated_by=actor_label(s),updated_at=now_iso()))
        elif existing.attendance!='A':
            existing.attendance='P';existing.attendance_marks=attendance_max
            if any(getattr(existing,name,None) is not None for name in ('record_marks','performance_marks','viva_marks')):
                existing.marks=sum((getattr(existing,name,None) or 0) for name in ('attendance_marks','record_marks','performance_marks','viva_marks'))
            elif existing.marks is None:
                existing.marks=attendance_max
            existing.updated_by=actor_label(s);existing.updated_at=now_iso()
    register.updated_at=now_iso();s.commit();flash('All unmarked students set to Present.');return redirect(url_for('practical_mark_entry',register_id=register.id,experiment_id=experiment.id))


def _roll_identity_key(value):
    return re.sub(r'[^a-z0-9]+','',(value or '').strip().casefold())


def _practical_subject_key(value):
    tokens=re.findall(r'[a-z0-9]+',(value or '').casefold())
    ignored={'lab','laboratory','practical'}
    return ''.join(token for token in tokens if token not in ignored)


def practical_exam_metadata_for_exam(s,exam_id):
    """Resolve one safe practical mapping for an exam.

    Preferred mapping is stored at exam level so an already-created exam can be
    converted to a Practical Exam from the Exams -> Edit dialog.  Existing V64
    question-level mappings remain supported as a backward-compatible fallback.

    Safety rule: if an exam-level Practical Exam mapping exists, any question
    that already carries a practical serial must either match that same serial
    or the mapping is rejected.  Unmapped legacy questions are allowed because
    converting an existing regular exam is an explicit administrator action.
    """
    cfg=get_exam_config(s,exam_id,create=False)
    subject=(cfg.subject or '').strip() if cfg else ''
    exam_type=((getattr(cfg,'exam_type','') or 'regular').strip().lower() if cfg else 'regular')
    configured_experiment=normalize_practical_exam_no(getattr(cfg,'practical_experiment_no','') if cfg else '')
    questions=s.scalars(select(Question).where(Question.exam_id==exam_id).order_by(Question.id)).all()

    if exam_type=='practical_exam':
        if not configured_experiment:
            return {'is_practical':True,'valid':False,'reason':'missing_practical_experiment_no'}
        if not subject:
            return {'is_practical':True,'valid':False,'reason':'missing_exam_subject'}
        # A mapped question may never silently point to a different experiment.
        conflicting=[normalize_practical_exam_no(q.practical_experiment_no) for q in questions
                     if normalize_practical_exam_no(q.practical_experiment_no)
                     and normalize_practical_exam_no(q.practical_experiment_no)!=configured_experiment]
        if conflicting:
            return {'is_practical':True,'valid':False,'reason':'question_experiment_conflicts_with_exam'}
        return {'is_practical':True,'valid':True,'reason':'','experiment_no':configured_experiment,'subject':subject,'source':'exam'}

    # Backward compatibility for practical exams created under V64.1, where the
    # mapping was snapshotted only onto individual Exam Question rows.
    if not questions:
        return {'is_practical':False,'valid':False,'reason':'no_questions'}
    mapped=[normalize_practical_exam_no(q.practical_experiment_no) for q in questions]
    practical=[value for value in mapped if value]
    if not practical:
        return {'is_practical':False,'valid':True,'reason':'normal_exam'}
    if len(practical)!=len(questions):
        return {'is_practical':True,'valid':False,'reason':'mixed_practical_and_regular_questions'}
    experiment_numbers=set(practical)
    if len(experiment_numbers)!=1:
        return {'is_practical':True,'valid':False,'reason':'multiple_experiment_numbers'}
    if not subject:
        return {'is_practical':True,'valid':False,'reason':'missing_exam_subject'}
    return {'is_practical':True,'valid':True,'reason':'','experiment_no':next(iter(experiment_numbers)),'subject':subject,'source':'questions'}



def resolve_practical_target_for_student(s,student,experiment_no,subject=''):
    """Resolve one practical row using full roll identity + exact experiment serial.

    This intentionally mirrors the safe Viva mapping rules.  Section/year and
    subject are only disambiguators; if more than one target remains, nothing is
    graded automatically.
    """
    if not student:
        return {'ok':False,'reason':'student_not_found','matches':0}
    registration_key=_roll_identity_key(student.registration_no)
    login_key=_roll_identity_key(student.roll_no)
    if not registration_key and not login_key:
        return {'ok':False,'reason':'student_roll_missing','matches':0}
    target_experiment=normalize_practical_exam_no(experiment_no)
    if not target_experiment:
        return {'ok':False,'reason':'missing_practical_experiment_no','matches':0}

    stmt=select(PracticalStudent)
    roll_filters=[]
    if student.registration_no:
        roll_filters.append(func.lower(PracticalStudent.roll_no)==student.registration_no.strip().casefold())
    if student.roll_no:
        roll_filters.append(PracticalStudent.roll_no.endswith(student.roll_no))
    if roll_filters:
        stmt=stmt.where(or_(*roll_filters))
    student_rows=s.scalars(stmt).all()
    candidates=[]
    for practical_student in student_rows:
        practical_roll_key=_roll_identity_key(practical_student.roll_no)
        if registration_key:
            if practical_roll_key!=registration_key:
                continue
        elif not (login_key and practical_roll_key.endswith(login_key)):
            continue
        register=s.get(PracticalRegister,practical_student.register_id)
        if not register:
            continue
        experiment=s.scalar(select(PracticalExperiment).where(
            PracticalExperiment.register_id==register.id,
            PracticalExperiment.experiment_no==target_experiment
        ))
        if not experiment:
            experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id)).all()
            experiment=next((row for row in experiments if normalize_practical_exam_no(row.experiment_no)==target_experiment),None)
        if experiment:
            candidates.append((register,practical_student,experiment))

    match_basis='roll+experiment'
    if len(candidates)>1:
        group=student_group(s,student.id)
        group_section=(getattr(group,'section','') or '').strip().upper() if group else ''
        group_year=(getattr(group,'academic_year','') or '').strip().casefold() if group else ''
        if group_section:
            section_matches=[item for item in candidates if _practical_register_section_code(item[0])==group_section]
            if group_year and len(section_matches)>1:
                year_matches=[item for item in section_matches if (item[0].academic_year or '').strip().casefold()==group_year]
                if year_matches:
                    section_matches=year_matches
            if len(section_matches)==1:
                candidates=section_matches;match_basis='roll+experiment+section'
            elif section_matches:
                candidates=section_matches
    if len(candidates)>1:
        target_subject=_practical_subject_key(subject)
        if target_subject:
            subject_matches=[item for item in candidates if _practical_subject_key(item[0].subject)==target_subject]
            if len(subject_matches)==1:
                candidates=subject_matches;match_basis+=' + subject'
            elif subject_matches:
                candidates=subject_matches
    if len(candidates)!=1:
        return {'ok':False,'reason':'practical_target_not_unique','matches':len(candidates)}
    register,practical_student,experiment=candidates[0]
    return {'ok':True,'reason':'','register':register,'practical_student':practical_student,'experiment':experiment,'match_basis':match_basis}


def _strip_code_comments(value):
    """Remove common source comments before similarity scoring."""
    text=value or ''
    text=re.sub(r'/\*.*?\*/',' ',text,flags=re.S)
    text=re.sub(r'(?m)//[^\n\r]*$',' ',text)
    text=re.sub(r'(?m)^\s*#(?!\s*(include|define|if|ifdef|ifndef|endif|pragma)\b).*$',' ',text,flags=re.I)
    return text


def _code_tokens(value):
    text=_strip_code_comments(value).replace('\r\n','\n').replace('\r','\n')
    # Strings/numbers are normalized so harmless literal changes do not dominate
    # the structural comparison.  Identifiers and operators are retained.
    text=re.sub(r'"(?:\\.|[^"\\])*"', ' STR ', text)
    text=re.sub(r"'(?:\\.|[^'\\])*'", ' CHR ', text)
    text=re.sub(r'\b\d+(?:\.\d+)?\b',' NUM ',text)
    return re.findall(r'[A-Za-z_][A-Za-z0-9_]*|==|!=|<=|>=|&&|\|\||\+\+|--|->|::|[{}()\[\].,;:+\-*/%<>=!?:]',text.casefold())


PRACTICAL_CODE_LANGUAGE_KEYWORDS={
    'package','import','class','interface','enum','extends','implements','public','private','protected',
    'static','final','abstract','void','int','long','float','double','boolean','char','byte','short','string',
    'new','this','super','return','if','else','for','while','do','switch','case','default','break','continue',
    'try','catch','finally','throw','throws','synchronized','volatile','transient','native','instanceof',
    'override','const','auto','struct','namespace','using','include','define','def','lambda','in','is','not',
    'and','or','true','false','null','none','with','as','from','pass','yield','async','await'
}


def _code_semantic_tokens(value):
    """Normalize renameable identifiers while preserving behavior/API structure.

    A student's local variable names and Android resource names are implementation
    choices, not correctness requirements.  This normalization therefore treats
    ``toggleButton`` and ``myButton`` (or ``R.id.toggleButton`` and
    ``R.id.my_toggle``) as equivalent, while keeping method/API names such as
    ``findViewById`` / ``setText`` so genuinely different behavior still affects
    the score.
    """
    tokens=_code_tokens(value)
    out=[];identifier_map={};next_identifier=1;i=0
    while i<len(tokens):
        token=tokens[i]
        # Android-style resource references: R.id.some_name, R.layout.some_name,
        # R.string.some_name, etc.  The final resource identifier is arbitrary.
        if (i+4<len(tokens) and token=='r' and tokens[i+1]=='.' and
                re.fullmatch(r'[a-z_][a-z0-9_]*',tokens[i+2] or '') and
                tokens[i+3]=='.' and re.fullmatch(r'[a-z_][a-z0-9_]*',tokens[i+4] or '')):
            out.extend(['r','.',tokens[i+2],'.','RESOURCE_ID'])
            i+=5;continue
        if re.fullmatch(r'[a-z_][a-z0-9_]*',token or ''):
            previous=tokens[i-1] if i else ''
            following=tokens[i+1] if i+1<len(tokens) else ''
            # Language words and callable/member API names carry semantic meaning.
            if token in PRACTICAL_CODE_LANGUAGE_KEYWORDS or previous in {'.','::'} or following=='(':
                out.append(token)
            else:
                if token not in identifier_map:
                    identifier_map[token]=f'ID{next_identifier}'
                    next_identifier+=1
                out.append(identifier_map[token])
        else:
            out.append(token)
        i+=1
    return out


def _code_behavior_tokens(value):
    """Extract callable/member tokens so behavior has explicit grading weight."""
    tokens=_code_tokens(value);out=[]
    for i,token in enumerate(tokens):
        if not re.fullmatch(r'[a-z_][a-z0-9_]*',token or ''):
            continue
        previous=tokens[i-1] if i else ''
        following=tokens[i+1] if i+1<len(tokens) else ''
        if following=='(' and token not in {'if','for','while','switch','catch','synchronized'}:
            out.append('CALL:'+token)
        elif previous in {'.','::'}:
            # Ignore the arbitrary final name in R.id.foo / R.layout.foo.
            if i>=4 and tokens[i-4]=='r' and tokens[i-3]=='.' and tokens[i-1]=='.':
                continue
            out.append('MEMBER:'+token)
    return out


def evaluate_practical_code(reference_code,student_code):
    """Return deterministic 0..1 logic-oriented similarity for practical code.

    The score deliberately ignores harmless local-variable/resource-ID renaming,
    but still rewards matching APIs, operators, control structure and overall
    program coverage.  It never executes student code.
    """
    ref=_code_semantic_tokens(reference_code);sub=_code_semantic_tokens(student_code)
    if len(ref)<6 or len(sub)<3:
        return 0.0
    sequence=difflib.SequenceMatcher(None,ref,sub,autojunk=False).ratio()
    from collections import Counter
    rc,sc=Counter(ref),Counter(sub)
    common=sum((rc & sc).values())
    multiset=(2.0*common/(len(ref)+len(sub))) if ref or sub else 0.0

    ref_behavior=_code_behavior_tokens(reference_code);sub_behavior=_code_behavior_tokens(student_code)
    if ref_behavior or sub_behavior:
        rb,sb=Counter(ref_behavior),Counter(sub_behavior)
        behavior_common=sum((rb & sb).values())
        behavior=(2.0*behavior_common/(len(ref_behavior)+len(sub_behavior))) if (ref_behavior or sub_behavior) else 0.0
    else:
        behavior=sequence

    length_ratio=min(len(ref),len(sub))/max(len(ref),len(sub))
    coverage_penalty=0.20+0.80*length_ratio
    score=(0.35*sequence+0.20*multiset+0.45*behavior)*coverage_penalty
    return max(0.0,min(1.0,score))


def practical_code_exact_penalty(student_code,penalty_rules,per_word=1.0):
    """Calculate faculty-defined exact-match deductions without executing code.

    Rules are private and one-per-line.  A rule containing exactly one normal
    identifier/keyword is matched as a case-sensitive whole word and deducts
    ``per_word`` for every occurrence.  Any other rule is treated as an exact
    code line after harmless whitespace normalization; each exact occurrence
    deducts ``per_word`` for every word/token configured in that rule.
    """
    source=(student_code or '').replace('\r\n','\n').replace('\r','\n')
    rules=(penalty_rules or '').replace('\r\n','\n').replace('\r','\n')
    normalized_source_lines=[' '.join(line.strip().split()) for line in source.split('\n')]
    details=[];total_units=0
    seen=set()
    for raw_rule in rules.split('\n'):
        rule=raw_rule.strip()
        if not rule or rule in seen:
            continue
        seen.add(rule)
        if re.fullmatch(r'[A-Za-z_][A-Za-z0-9_]*',rule):
            pattern=rf'(?<![A-Za-z0-9_]){re.escape(rule)}(?![A-Za-z0-9_])'
            occurrences=len(re.findall(pattern,source))
            units=occurrences
            kind='word'
        else:
            normalized_rule=' '.join(rule.split())
            occurrences=sum(1 for line in normalized_source_lines if line==normalized_rule)
            # "0.5 per word": count source-language words/identifiers/numbers in
            # the configured line, with a minimum of one penalty unit.
            words=re.findall(r'[A-Za-z_][A-Za-z0-9_]*|\d+(?:\.\d+)?',rule)
            units=occurrences*max(1,len(words))
            kind='line'
        if occurrences:
            total_units+=units
            details.append({'rule':rule,'kind':kind,'occurrences':occurrences,'units':units})
    deduction=round(float(total_units)*float(per_word),2)
    return {'deduction':deduction,'units':total_units,'matches':details}


PRACTICAL_CODE_PEER_COPY_THRESHOLD=0.97
PRACTICAL_CODE_PEER_MIN_TOKENS=40
PRACTICAL_CODE_LANGUAGE_WORDS={
    'package','import','class','interface','enum','extends','implements','public','private','protected',
    'static','final','abstract','void','int','long','float','double','boolean','char','byte','short','string',
    'new','this','super','return','if','else','for','while','do','switch','case','default','break','continue',
    'try','catch','finally','throw','throws','synchronized','volatile','transient','native','instanceof',
    'override','const','auto','struct','namespace','using','include','define','def','lambda','in','is','not',
    'and','or','true','false','null','none','with','as','from','pass','yield','async','await'
}


def _code_clone_tokens(value):
    """Normalize cosmetic identifier renames for plagiarism/clone review only."""
    tokens=_code_tokens(value);out=[];previous=''
    for token in tokens:
        if re.fullmatch(r'[a-z_][a-z0-9_]*',token):
            # Keep language words and member/API names after . or ::, while local
            # variable / helper identifiers collapse to ID.  Thus changing only
            # toggleButton -> myButton does not evade the clone detector.
            if token in PRACTICAL_CODE_LANGUAGE_WORDS or previous in {'.','::'}:
                out.append(token)
            else:
                out.append('ID')
        else:
            out.append(token)
        previous=token
    return out


def practical_code_peer_similarity(first_code,second_code):
    """Return 0..1 similarity used only to flag near-clone student submissions.

    Formatting, comments, literals and small cosmetic changes are intentionally
    discounted by _code_tokens(), so changing whitespace or a message string does
    not defeat the copy check.  The threshold is deliberately very high because
    beginner practical programs can legitimately share framework boilerplate.
    """
    first=_code_clone_tokens(first_code);second=_code_clone_tokens(second_code)
    if min(len(first),len(second))<PRACTICAL_CODE_PEER_MIN_TOKENS:
        return 0.0
    sequence=difflib.SequenceMatcher(None,first,second,autojunk=False).ratio()
    # Token 4-grams make the test resistant to whitespace/comment/literal edits
    # while still requiring almost the same program structure before we flag it.
    def grams(tokens,size=4):
        return {tuple(tokens[i:i+size]) for i in range(max(0,len(tokens)-size+1))}
    ga,gb=grams(first),grams(second)
    gram_score=(len(ga & gb)/len(ga | gb)) if ga and gb else 0.0
    length_ratio=min(len(first),len(second))/max(len(first),len(second))
    return max(0.0,min(1.0,(0.70*sequence+0.30*gram_score)*length_ratio))


def find_practical_code_peer_clone(s,student_id,exam_id,experiment_id,source_code):
    """Find the closest prior submission from another student in this exam."""
    submissions=s.scalars(select(PracticalCodeSubmission).where(
        PracticalCodeSubmission.exam_id==exam_id,
        PracticalCodeSubmission.practical_experiment_id==experiment_id,
        PracticalCodeSubmission.student_id!=student_id
    ).order_by(PracticalCodeSubmission.id)).all()
    best=None;best_similarity=0.0
    for item in submissions:
        similarity=practical_code_peer_similarity(source_code,item.source_code or '')
        if similarity>best_similarity:
            best=item;best_similarity=similarity
    if best and best_similarity>=PRACTICAL_CODE_PEER_COPY_THRESHOLD:
        return best,best_similarity
    return None,best_similarity


def append_code_review_remark(existing,text_value):
    existing=(existing or '').strip()
    marker='[CODE REVIEW]'
    # Replace an older auto-generated review note instead of growing the remarks
    # on every re-submission. Faculty-written remarks before the marker are kept.
    base=existing.split(marker,1)[0].rstrip()
    review=f'{marker} {text_value}'.strip()
    return f'{base} {review}'.strip()[:500]


def clear_code_review_remark(existing):
    existing=(existing or '').strip()
    return existing.split('[CODE REVIEW]',1)[0].rstrip()[:500]


def practical_code_exam_rows_for_student(s,student):
    rows=[]
    if not student:return rows

    # Closed practical windows remain visible as a read-only marks/history page.
    # Upcoming practical work is shown only after the exam has been activated.
    exams_list=s.scalars(select(Exam).order_by(Exam.id.desc())).all()
    for exam in exams_list:
        meta=practical_exam_metadata_for_exam(s,exam.id)
        if not meta.get('is_practical') or not meta.get('valid'):
            continue
        if not practical_exam_assigned_to_student(s,student.id,exam):
            continue

        cfg=get_exam_config(s,exam.id,create=False)
        window=practical_code_window_state(cfg)
        submission=s.scalar(select(PracticalCodeSubmission).where(
            PracticalCodeSubmission.student_id==student.id,
            PracticalCodeSubmission.exam_id==exam.id
        ))

        # Do not expose an inactive future practical exam.  Once a practical
        # window has closed (or the student already submitted), retain it so
        # the student can continue to view the frozen code and marks.
        visible=bool(exam.is_active or submission or window.get('status')=='closed')
        if not visible:
            continue

        target=resolve_practical_target_for_student(
            s,student,meta.get('experiment_no',''),meta.get('subject','')
        )
        practical_mark=None;mark_maxima=None;mark_total_max=None
        if target.get('ok'):
            practical_mark=s.scalar(select(PracticalMark).where(
                PracticalMark.practical_student_id==target['practical_student'].id,
                PracticalMark.practical_experiment_id==target['experiment'].id
            ))
            mark_maxima=practical_marks_maxima(target['register'])
            mark_total_max=sum(mark_maxima.values())

        can_submit=bool(exam.is_active and window.get('can_edit'))
        if not exam.is_active and window.get('status')!='closed':
            access_label='This Practical Exam is not active.'
        elif not window.get('can_edit'):
            access_label=window.get('label') or 'Practical code editing is closed.'
        else:
            access_label=''

        rows.append({
            'exam':exam,'meta':meta,'target':target,'submission':submission,
            'practical_mark':practical_mark,'mark_maxima':mark_maxima,'mark_total_max':mark_total_max,
            'can_submit':can_submit,'access_label':access_label,'code_window':window
        })
    return rows


def student_practical_code_available(s,student_id):
    student=s.get(Student,student_id) if student_id else None
    return bool(practical_code_exam_rows_for_student(s,student))

def sync_practical_viva_from_attempt(s,attempt):
    """Copy a submitted Practical Exam score into exactly one Viva column.

    Matching is intentionally strict: exam subject + practical experiment serial +
    the student's full registration/roll number must resolve to one practical row.
    If the mapping is ambiguous, no practical mark is changed.
    """
    if not attempt or attempt.status!='submitted':
        return {'updated':False,'reason':'attempt_not_submitted'}
    if (attempt.grading_status or 'complete')!='complete':
        return {'updated':False,'reason':'grading_pending'}
    meta=practical_exam_metadata_for_exam(s,attempt.exam_id)
    if not meta.get('is_practical'):
        return {'updated':False,'reason':'normal_exam'}
    if not meta.get('valid'):
        return {'updated':False,'reason':meta.get('reason') or 'invalid_practical_mapping'}
    student=s.get(Student,attempt.student_id)
    if not student:
        return {'updated':False,'reason':'student_not_found'}
    registration_key=_roll_identity_key(student.registration_no)
    login_key=_roll_identity_key(student.roll_no)
    if not registration_key and not login_key:
        return {'updated':False,'reason':'student_roll_missing'}

    # Match the student primarily by the full university registration / roll number.
    # The practical register subject is NOT a hard requirement: an administrator may
    # convert an already-created exam to a Practical Exam even if the exam's catalog
    # subject label differs from the lab-register subject.  Safety is instead enforced
    # by the exact roll identity + exact experiment serial, with batch/section used as
    # a disambiguator when the same student exists in more than one practical register.
    stmt=select(PracticalStudent)
    roll_filters=[]
    if student.registration_no:
        roll_filters.append(func.lower(PracticalStudent.roll_no)==student.registration_no.strip().casefold())
    if student.roll_no:
        roll_filters.append(PracticalStudent.roll_no.endswith(student.roll_no))
    if roll_filters:
        stmt=stmt.where(or_(*roll_filters))
    student_rows=s.scalars(stmt).all()
    candidates=[]
    target_experiment=normalize_practical_exam_no(meta['experiment_no'])
    for practical_student in student_rows:
        practical_roll_key=_roll_identity_key(practical_student.roll_no)
        if registration_key:
            if practical_roll_key!=registration_key:
                continue
        elif not (login_key and practical_roll_key.endswith(login_key)):
            continue
        register=s.get(PracticalRegister,practical_student.register_id)
        if not register:
            continue
        experiment=s.scalar(select(PracticalExperiment).where(
            PracticalExperiment.register_id==register.id,
            PracticalExperiment.experiment_no==target_experiment
        ))
        if not experiment:
            # Backward-compatible fallback for legacy serial formatting such as 01 / Exp-1.
            experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id)).all()
            experiment=next((row for row in experiments if normalize_practical_exam_no(row.experiment_no)==target_experiment),None)
        if experiment:
            candidates.append((register,practical_student,experiment))

    match_basis='roll+experiment'
    if len(candidates)>1:
        group=student_group(s,student.id)
        group_section=(getattr(group,'section','') or '').strip().upper() if group else ''
        group_year=(getattr(group,'academic_year','') or '').strip().casefold() if group else ''
        if group_section:
            section_matches=[item for item in candidates if _practical_register_section_code(item[0])==group_section]
            if group_year and len(section_matches)>1:
                year_matches=[item for item in section_matches if (item[0].academic_year or '').strip().casefold()==group_year]
                if year_matches:
                    section_matches=year_matches
            if len(section_matches)==1:
                candidates=section_matches
                match_basis='roll+experiment+section'
            elif section_matches:
                candidates=section_matches

    # Subject is only a final tie-breaker. It must never block an otherwise unique
    # roll + experiment mapping, which is the mapping requested for Practical Exams.
    if len(candidates)>1:
        target_subject=_practical_subject_key(meta.get('subject',''))
        if target_subject:
            subject_matches=[item for item in candidates if _practical_subject_key(item[0].subject)==target_subject]
            if len(subject_matches)==1:
                candidates=subject_matches
                match_basis+=' + subject'
            elif subject_matches:
                candidates=subject_matches

    if len(candidates)!=1:
        return {'updated':False,'reason':'practical_target_not_unique','matches':len(candidates)}

    register,practical_student,experiment=candidates[0]
    score=float(attempt.score or 0)
    total=float(attempt.total_marks or 0)
    viva_max=float(practical_marks_maxima(register)['viva'])
    if total<=0 or viva_max<0:
        return {'updated':False,'reason':'invalid_exam_or_viva_max'}
    # Normal case: a 10-mark viva exam copies its earned score directly. If a
    # practical exam is configured above the Viva maximum, scale proportionally
    # so the value cannot exceed the configured Viva column maximum.
    viva_value=score if total<=viva_max else (score*viva_max/total)
    viva_value=max(0.0,min(viva_max,round(viva_value,2)))
    mark=s.scalar(select(PracticalMark).where(PracticalMark.practical_student_id==practical_student.id,PracticalMark.practical_experiment_id==experiment.id))
    if mark and mark.viva_marks is not None and abs(float(mark.viva_marks)-viva_value)<0.0001:
        return {'updated':False,'reason':'already_synced','register_id':register.id,'practical_student_id':practical_student.id,'experiment_id':experiment.id,'experiment_no':target_experiment,'viva_marks':viva_value,'viva_max':viva_max}
    if not mark:
        mark=PracticalMark(register_id=register.id,practical_student_id=practical_student.id,practical_experiment_id=experiment.id,attendance='',attendance_marks=None,record_marks=None,performance_marks=None,viva_marks=viva_value,marks=viva_value,remarks='',updated_by='Practical Exam',updated_at=now_iso())
        s.add(mark);s.flush()
    else:
        mark.viva_marks=viva_value
        if (mark.attendance or '').upper()=='A':
            mark.marks=0.0
        else:
            components=[mark.attendance_marks,mark.record_marks,mark.performance_marks,mark.viva_marks]
            mark.marks=sum(value or 0 for value in components) if any(value is not None for value in components) else viva_value
        mark.updated_by='Practical Exam';mark.updated_at=now_iso()
    register.updated_at=now_iso()
    audit_event(s,'practical_viva_auto_synced','practical_mark',mark.id or '',f'exam={attempt.exam_id}, attempt={attempt.id}, roll={practical_student.roll_no}, experiment={target_experiment}, viva={viva_value}/{viva_max}, match={match_basis}')
    return {'updated':True,'reason':'','register_id':register.id,'practical_student_id':practical_student.id,'experiment_id':experiment.id,'experiment_no':target_experiment,'viva_marks':viva_value,'viva_max':viva_max,'match_basis':match_basis}


def resync_submitted_practical_attempts(s,exam_id):
    """Repair Viva marks for attempts submitted before an exam was mapped as practical."""
    attempts=s.scalars(select(Attempt).where(Attempt.exam_id==exam_id,Attempt.status=='submitted').order_by(Attempt.id)).all()
    updated=0;skipped=0;reasons={}
    for attempt in attempts:
        try:
            result=sync_practical_viva_from_attempt(s,attempt)
            if result.get('updated'):
                s.commit();updated+=1
            else:
                skipped+=1;reason=result.get('reason') or 'not_updated';reasons[reason]=reasons.get(reason,0)+1
        except Exception:
            s.rollback();skipped+=1;reasons['exception']=reasons.get('exception',0)+1
            try:app.logger.exception('Practical Exam retroactive viva sync failed for attempt %s',attempt.id)
            except Exception:pass
    return {'attempts':len(attempts),'updated':updated,'skipped':skipped,'reasons':reasons}


@app.route('/admin/practicals/<int:register_id>/template/<kind>/<fmt>')
@practical_required
def practical_template(register_id,kind,fmt):
    s=DB();register=practical_register_access(s,register_id)
    if kind=='students':headers=['roll_no','name'];example=['2024/17008','Student Name'];base='practical_student_roster'
    elif kind=='experiments':headers=['experiment_no','title','reference_code'];example=['1','Installation of Android Studio','Paste the faculty reference program here (optional)'];base='practical_experiment_list'
    else:abort(404)
    if fmt=='csv':
        out=io.StringIO(newline='');w=csv.writer(out);w.writerow(headers);w.writerow(example);data=io.BytesIO(out.getvalue().encode('utf-8-sig'));return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'{base}.csv')
    if fmt=='xlsx':
        wb=Workbook();ws=wb.active;ws.title='Students' if kind=='students' else 'Experiments';ws.append(headers);ws.append(example)
        for cell in ws[1]:cell.font=Font(bold=True)
        ws.column_dimensions['A'].width=20;ws.column_dimensions['B'].width=70 if kind=='experiments' else 34
        if kind=='experiments':ws.column_dimensions['C'].width=80
        data=io.BytesIO();wb.save(data);data.seek(0);return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=f'{base}.xlsx')
    abort(404)


@app.route('/admin/practicals/<int:register_id>/export/<fmt>')
@practical_required
def practical_export(register_id,fmt):
    s=DB();register=practical_register_access(s,register_id);repair_practical_experiment_numbers(s,register);students=s.scalars(select(PracticalStudent).where(PracticalStudent.register_id==register.id).order_by(PracticalStudent.sequence,PracticalStudent.roll_no)).all();experiments=s.scalars(select(PracticalExperiment).where(PracticalExperiment.register_id==register.id).order_by(PracticalExperiment.sort_order,PracticalExperiment.id)).all();marks=s.scalars(select(PracticalMark).where(PracticalMark.register_id==register.id)).all();mark_map={(m.practical_student_id,m.practical_experiment_id):m for m in marks};headers=['S.No','Roll No','Student Name']+[e.experiment_no for e in experiments]+['Total','Possible','Percentage'];possible=sum(e.max_marks for e in experiments);matrix=[]
    for idx,st in enumerate(students,start=1):
        cells=[];total=0.0
        for e in experiments:
            m=mark_map.get((st.id,e.id))
            if not m:cells.append('')
            elif m.attendance=='A':cells.append('A')
            elif m.marks is None:cells.append('P')
            else:cells.append(m.marks);total+=float(m.marks)
        matrix.append([idx,st.roll_no,st.name]+cells+[round(total,2),possible,round(total*100/possible,1) if possible else 0])
    safe=''.join(ch if ch.isalnum() else '_' for ch in register.title).strip('_')[:60] or 'practical_marks'
    if fmt=='csv':
        out=io.StringIO(newline='');w=csv.writer(out);w.writerow(headers);w.writerows(matrix);data=io.BytesIO(out.getvalue().encode('utf-8-sig'));return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'{safe}.csv')
    if fmt=='xlsx':
        wb=Workbook();ws=wb.active;ws.title='Practical Marks';ws.append([register.title]);ws.append([f'Subject: {register.subject}',f'Section: {register.section}',f'Academic Year: {register.academic_year}']);ws.append(headers)
        for row in matrix:ws.append(row)
        for cell in ws[3]:cell.font=Font(bold=True)
        ws.freeze_panes='D4';ws.column_dimensions['A'].width=8;ws.column_dimensions['B'].width=18;ws.column_dimensions['C'].width=30
        for col in range(4,4+len(experiments)):ws.column_dimensions[ws.cell(3,col).column_letter].width=11
        data=io.BytesIO();wb.save(data);data.seek(0);return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=f'{safe}.xlsx')
    abort(404)


@app.route('/admin/students',methods=['GET','POST'])
@staff_required
def students():
    s=DB()
    if request.method=='POST':
        roll=request.form.get('roll_no','').strip(); name=request.form.get('name','').strip(); pw=request.form.get('password','')
        if not roll or not name or not pw: flash('All student fields are required.','error')
        else:
            try:
                st=Student(roll_no=roll,registration_no='',name=name,password_hash=generate_password_hash(pw),created_at=now_iso()); s.add(st); s.flush();
                try:group_id=int(request.form.get('group_id','0') or 0)
                except ValueError:group_id=0
                if group_id and s.get(AcademicGroup,group_id):assign_student_group(s,st.id,group_id)
                audit_event(s,'student_created','student',st.id,roll); s.commit(); flash('Student added.')
            except IntegrityError:s.rollback();flash('Roll number already exists.','error')
    rows=s.scalars(select(Student).order_by(Student.roll_no)).all();groups=s.scalars(select(AcademicGroup).where(AcademicGroup.is_active==True).order_by(AcademicGroup.program,AcademicGroup.semester,AcademicGroup.section)).all();membership=dict(s.execute(select(StudentGroup.student_id,StudentGroup.group_id)).all());group_map={g.id:g for g in groups};practical_registers=s.scalars(practical_register_stmt(s)).all();return render_template('students.html',students=rows,groups=groups,membership=membership,group_map=group_map,group_label=group_label,practical_registers=practical_registers,section_code=_section_code,practical_section_code=_practical_register_section_code)

@app.route('/admin/students/sync-practical',methods=['POST'])
@staff_required
def sync_students_from_practical():
    """Create/update exam logins from an existing practical-register roster.

    Browser requests are processed in small, committed chunks (20 rows by
    default).  This keeps production requests short and makes the sync
    resumable/idempotent.  Inside every chunk each student is protected by a
    SAVEPOINT, so one legacy/duplicate row cannot abort the remaining rows.
    """
    s=DB()
    wants_json=(
        request.form.get('ajax')=='1'
        or request.headers.get('X-Requested-With')=='XMLHttpRequest'
        or request.accept_mimetypes.best=='application/json'
    )

    def respond_error(message,status=400):
        if wants_json:
            return jsonify({'ok':False,'error':message}),status
        flash(message,'error')
        return redirect(url_for('students'))

    try:
        register_id=int(request.form.get('register_id','0') or 0)
        group_id=int(request.form.get('group_id','0') or 0)
    except ValueError:
        register_id=group_id=0
    if not register_id or not group_id:
        return respond_error('Choose a practical list and a batch / section.')

    register=practical_register_access(s,register_id)
    group=s.get(AcademicGroup,group_id)
    if not group or not group.is_active:
        return respond_error('Choose a valid active batch / section.')

    register_section=_practical_register_section_code(register)
    group_section=_section_code(group.section)
    if register_section and group_section and register_section!=group_section:
        return respond_error(f'This practical list belongs to Section {register_section}; it cannot be synced to Section {group_section}.')
    register_year=(register.academic_year or '').strip()
    group_year=(group.academic_year or '').strip()
    if register_year and group_year and register_year.casefold()!=group_year.casefold():
        return respond_error(f'This practical list is for {register_year}; choose a batch / section from the same academic year.')

    practical_students=s.scalars(
        select(PracticalStudent)
        .where(PracticalStudent.register_id==register.id)
        .order_by(PracticalStudent.sequence,PracticalStudent.id)
    ).all()
    if not practical_students:
        return respond_error('The selected practical list has no students.')

    total=len(practical_students)
    if wants_json:
        try:offset=max(0,int(request.form.get('offset','0') or 0))
        except ValueError:offset=0
        try:batch_size=int(request.form.get('batch_size',str(PRACTICAL_SYNC_BATCH_SIZE)) or PRACTICAL_SYNC_BATCH_SIZE)
        except ValueError:batch_size=PRACTICAL_SYNC_BATCH_SIZE
        batch_size=max(1,min(PRACTICAL_SYNC_BATCH_SIZE,batch_size))
        if offset>=total:
            return jsonify({'ok':True,'done':True,'offset':offset,'next_offset':total,'total':total,'processed':0,'created':0,'updated':0,'assigned':0,'collisions':0,'invalid':0,'failed':0,'issues':[]})
        selected_students=practical_students[offset:offset+batch_size]
    else:
        # Compatibility fallback when JavaScript is disabled.  Keep the route
        # functional, but the normal UI uses sequential 20-row requests.
        offset=0
        selected_students=practical_students
        batch_size=len(selected_students)

    # Build deterministic lookup maps for this chunk. Registration numbers are
    # not a database unique key in older installations, so ambiguous legacy
    # registrations are skipped instead of attaching the wrong account.
    master_students=s.scalars(select(Student)).all()
    by_login={row.roll_no.strip().casefold():row for row in master_students if row.roll_no}
    by_registration={}
    ambiguous_registrations=set()
    for row in master_students:
        keys=[]
        key=_canonical_registration_no(getattr(row,'registration_no',''))
        if key:keys.append(key)
        legacy_key=_canonical_registration_no(row.roll_no)
        if len(legacy_key)>5 and legacy_key not in keys:keys.append(legacy_key)
        for item_key in keys:
            previous=by_registration.get(item_key)
            if previous is not None and previous.id!=row.id:
                ambiguous_registrations.add(item_key)
            else:
                by_registration[item_key]=row

    created=updated=assigned=collisions=invalid=failed=0
    seen_registration=set()
    seen_login={}
    failure_types={}
    issues=[]

    for local_index,practical_student in enumerate(selected_students,start=offset+1):
        registration=(practical_student.roll_no or '').strip()
        name=(practical_student.name or '').strip()
        reg_key=_canonical_registration_no(registration)
        login_id=_student_login_from_registration(registration)

        def add_issue(kind,detail=''):
            if len(issues)>=12:return
            label=registration or f'row {local_index}'
            issues.append({'row':local_index,'registration':label,'type':kind,'detail':detail[:160]})

        if not reg_key or not login_id or not name:
            invalid+=1;add_issue('invalid','Missing name or a registration number with at least five digits.');continue
        if reg_key in seen_registration:
            # Duplicate inside the same practical chunk. The first row wins.
            invalid+=1;add_issue('duplicate','Duplicate registration number in this practical list chunk.');continue
        seen_registration.add(reg_key)
        previous_reg=seen_login.get(login_id)
        if previous_reg and previous_reg!=reg_key:
            collisions+=1;add_issue('login_collision',f'Last five digits {login_id} are shared by another registration.');continue
        seen_login[login_id]=reg_key
        if reg_key in ambiguous_registrations:
            collisions+=1;add_issue('registration_collision','Multiple existing exam accounts use this registration number.');continue

        student=by_registration.get(reg_key)
        login_owner=by_login.get(login_id.casefold())
        if student is None and login_owner is not None:
            owner_reg=_canonical_registration_no(getattr(login_owner,'registration_no',''))
            owner_name=(login_owner.name or '').strip().casefold()
            if owner_reg and owner_reg!=reg_key:
                collisions+=1;add_issue('login_collision',f'Login {login_id} already belongs to another registration.');continue
            if not owner_reg and owner_name and owner_name!=name.casefold():
                collisions+=1;add_issue('legacy_login_collision',f'Login {login_id} belongs to an older account with a different name.');continue
            student=login_owner
        if student is not None and login_owner is not None and login_owner.id!=student.id:
            collisions+=1;add_issue('login_collision',f'Login {login_id} is already owned by another account.');continue

        row_created=row_updated=row_assigned=False
        old_login=(student.roll_no or '').strip().casefold() if student is not None else ''
        had_registration=bool(_canonical_registration_no(getattr(student,'registration_no',''))) if student is not None else False
        try:
            with s.begin_nested():
                if student is None:
                    student=Student(
                        roll_no=login_id,
                        registration_no=registration,
                        name=name,
                        password_hash=generate_password_hash(login_id),
                        created_at=now_iso(),
                    )
                    s.add(student);s.flush();row_created=True
                else:
                    changed=False;login_changed=False
                    if student.roll_no!=login_id:
                        existing_login=by_login.get(login_id.casefold())
                        if existing_login is not None and existing_login.id!=student.id:
                            raise IntegrityError('login collision',None,None)
                        student.roll_no=login_id;changed=True;login_changed=True
                    if getattr(student,'registration_no','')!=registration:
                        student.registration_no=registration;changed=True
                    if student.name!=name:
                        student.name=name;changed=True
                    # Preserve a student's password on ordinary re-sync. Reset
                    # only while converting/adopting a legacy account.
                    if login_changed or not had_registration:
                        student.password_hash=generate_password_hash(login_id)
                    row_updated=changed

                membership=s.scalar(select(StudentGroup).where(StudentGroup.student_id==student.id))
                if not membership or membership.group_id!=group.id:
                    assign_student_group(s,student.id,group.id);row_assigned=True
                s.flush()
        except IntegrityError as exc:
            collisions+=1;add_issue('database_collision',str(getattr(exc,'orig',None) or 'Duplicate login/assignment constraint.'));continue
        except Exception as exc:
            failed+=1
            failure_types[type(exc).__name__]=failure_types.get(type(exc).__name__,0)+1
            add_issue(type(exc).__name__,str(exc))
            continue

        if row_created:created+=1
        if row_updated:updated+=1
        if row_assigned:assigned+=1
        if old_login and old_login!=login_id.casefold():by_login.pop(old_login,None)
        by_login[login_id.casefold()]=student
        by_registration[reg_key]=student

    processed=len(selected_students)
    next_offset=min(total,offset+processed)
    detail=(
        f'register={register.id}, group={group.id}, offset={offset}, processed={processed}, '
        f'created={created}, updated={updated}, assigned={assigned}, collisions={collisions}, '
        f'invalid={invalid}, failed={failed}'
    )
    if failure_types:detail+=', failure_types='+json.dumps(failure_types,sort_keys=True)
    audit_event(s,'students_practical_sync_batch','student','',detail)
    try:
        s.commit()
    except IntegrityError as exc:
        s.rollback()
        return respond_error('This batch could not be saved because two student records would create the same login ID. The completed earlier batches are safe.',409)
    except Exception as exc:
        s.rollback()
        return respond_error(f'This batch could not be saved ({type(exc).__name__}). The completed earlier batches are safe and you can retry.',500)

    result={
        'ok':True,
        'done':next_offset>=total,
        'offset':offset,
        'next_offset':next_offset,
        'total':total,
        'processed':processed,
        'created':created,
        'updated':updated,
        'assigned':assigned,
        'collisions':collisions,
        'invalid':invalid,
        'failed':failed,
        'issues':issues,
    }
    if wants_json:
        return jsonify(result)

    parts=[
        f'{created} login'+('' if created==1 else 's')+' created',
        f'{updated} updated',
        f'{assigned} batch assignment'+('' if assigned==1 else 's')+' updated',
    ]
    if collisions:parts.append(f'{collisions} login/data collision'+('' if collisions==1 else 's')+' skipped')
    if invalid:parts.append(f'{invalid} invalid/duplicate row'+('' if invalid==1 else 's')+' skipped')
    if failed:parts.append(f'{failed} problem row'+('' if failed==1 else 's')+' skipped')
    flash('Practical list synced: '+', '.join(parts)+'.')
    return redirect(url_for('students'))

@app.route('/admin/students/import',methods=['POST'])
@staff_required
def import_students():
    upload=request.files.get('student_file')
    if not upload or not upload.filename: flash('Choose a CSV or Excel (.xlsx) file.','error'); return redirect(url_for('students'))
    try: headers,rows=_rows_from_upload(upload)
    except ValueError as exc: flash(str(exc),'error'); return redirect(url_for('students'))
    if 'name' not in set(headers) or not ({'roll_no','registration_no'} & set(headers)):
        flash('Required columns are name and either roll_no or registration_no.','error'); return redirect(url_for('students'))
    s=DB(); master=s.scalars(select(Student)).all()
    existing_logins={(x.roll_no or '').strip().casefold():x for x in master}
    existing_regs={_canonical_registration_no(getattr(x,'registration_no','')):x for x in master if _canonical_registration_no(getattr(x,'registration_no',''))}
    seen_logins=set();seen_regs=set();added=duplicates=invalid=collisions=0
    for row in rows:
        registration=(row.get('registration_no','') or '').strip(); raw_roll=(row.get('roll_no','') or '').strip(); name=(row.get('name','') or '').strip(); password=row.get('password','') or ''
        login=raw_roll or (_student_login_from_registration(registration) if registration else '')
        if not name or not login:
            invalid+=1;continue
        if registration and not password:password=_student_login_from_registration(registration) or login
        elif not password:password=login
        reg_key=_canonical_registration_no(registration) if registration else ''
        login_key=login.casefold()
        if login_key in seen_logins or (reg_key and reg_key in seen_regs):duplicates+=1;continue
        existing=existing_regs.get(reg_key) if reg_key else None
        login_owner=existing_logins.get(login_key)
        if existing and login_owner and existing.id!=login_owner.id:
            collisions+=1;continue
        if existing or login_owner:
            duplicates+=1;continue
        st=Student(roll_no=login,registration_no=registration,name=name,password_hash=generate_password_hash(password),created_at=now_iso());s.add(st);s.flush()
        group=find_or_create_group(s,row.get('department',''),row.get('program',''),row.get('semester',''),row.get('section',''),row.get('academic_year',''))
        if group:assign_student_group(s,st.id,group.id)
        seen_logins.add(login_key);existing_logins[login_key]=st
        if reg_key:seen_regs.add(reg_key);existing_regs[reg_key]=st
        added+=1
    try:
        audit_event(s,'students_bulk_import','student','',f'added={added}, duplicates={duplicates}, collisions={collisions}, invalid={invalid}'); s.commit()
    except IntegrityError:s.rollback();flash('Import could not be completed because one or more student login IDs conflict.','error');return redirect(url_for('students'))
    parts=[f'Imported {added} student login'+('' if added==1 else 's')+'.']
    if duplicates:parts.append(f'Skipped {duplicates} existing/duplicate student'+('' if duplicates==1 else 's')+'.')
    if collisions:parts.append(f'Skipped {collisions} login collision'+('' if collisions==1 else 's')+'.')
    if invalid:parts.append(f'Skipped {invalid} incomplete row'+('' if invalid==1 else 's')+'.')
    flash(' '.join(parts)); return redirect(url_for('students'))

@app.route('/admin/students/template/<fmt>')
@staff_required
def student_import_template(fmt):
    headers=['roll_no','registration_no','name','password','department','program','semester','section','academic_year']
    if fmt=='csv':
        out=io.StringIO(newline=''); writer=csv.writer(out); writer.writerow(headers); data=io.BytesIO(out.getvalue().encode('utf-8-sig')); return send_file(data,mimetype='text/csv',as_attachment=True,download_name='student_import_template.csv')
    if fmt=='xlsx':
        wb=Workbook(); ws=wb.active; ws.title='Students'; ws.append(headers)
        for cell in ws[1]:cell.font=Font(bold=True)
        ws.column_dimensions['A'].width=18;ws.column_dimensions['B'].width=28;ws.column_dimensions['C'].width=22
        data=io.BytesIO();wb.save(data);data.seek(0);return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name='student_import_template.xlsx')
    abort(404)


@app.route('/admin/groups',methods=['GET','POST'])
@staff_required
def academic_groups():
    s=DB()
    if request.method=='POST':
        action=request.form.get('action','create')
        if action=='create':
            group=find_or_create_group(s,request.form.get('department',''),request.form.get('program',''),request.form.get('semester',''),request.form.get('section',''),request.form.get('academic_year',''))
            if not group:flash('Program, semester and section are required.','error')
            else:audit_event(s,'academic_group_created','group',group.id,group_label(group));s.commit();flash('Batch / section created.')
        elif action=='assign':
            try:student_id=int(request.form.get('student_id','0'));group_id=int(request.form.get('group_id','0'))
            except ValueError:student_id=group_id=0
            if s.get(Student,student_id) and s.get(AcademicGroup,group_id):assign_student_group(s,student_id,group_id);audit_event(s,'student_group_assigned','student',student_id,f'group={group_id}');s.commit();flash('Student batch / section updated.')
        return redirect(url_for('academic_groups'))
    groups=s.scalars(select(AcademicGroup).order_by(AcademicGroup.academic_year.desc(),AcademicGroup.program,AcademicGroup.semester,AcademicGroup.section)).all()
    counts=dict(s.execute(select(StudentGroup.group_id,func.count()).group_by(StudentGroup.group_id)).all())
    students_list=s.scalars(select(Student).order_by(Student.roll_no)).all();memberships=dict(s.execute(select(StudentGroup.student_id,StudentGroup.group_id)).all())
    return render_template('groups.html',groups=groups,counts=counts,students=students_list,memberships=memberships,group_label=group_label,staff_role=current_staff_role(s))

@app.route('/admin/groups/<int:group_id>/toggle',methods=['POST'])
@staff_required
def toggle_group(group_id):
    s=DB();row=s.get(AcademicGroup,group_id)
    if row:row.is_active=not row.is_active;audit_event(s,'academic_group_enabled' if row.is_active else 'academic_group_disabled','group',row.id,group_label(row));s.commit()
    return redirect(url_for('academic_groups'))

@app.route('/admin/groups/<int:group_id>/delete',methods=['POST'])
@admin_required
def delete_group(group_id):
    """Delete only the exam-system batch/section mapping, never practical-register data."""
    s=DB();row=s.get(AcademicGroup,group_id)
    if not row:abort(404)
    label=group_label(row)
    member_count=s.scalar(select(func.count()).select_from(StudentGroup).where(StudentGroup.group_id==row.id)) or 0
    session_count=s.scalar(select(func.count()).select_from(ExamSession).where(ExamSession.group_id==row.id)) or 0
    # Student accounts remain intact and become Unassigned. Exam records remain
    # intact; only sessions that explicitly target this deleted group are removed.
    # PracticalRegister / PracticalStudent / PracticalMark are intentionally not
    # referenced here, so lab lists and marks are preserved unchanged.
    s.execute(delete(StudentGroup).where(StudentGroup.group_id==row.id))
    s.execute(delete(ExamSession).where(ExamSession.group_id==row.id))
    audit_event(s,'academic_group_deleted','group',row.id,f'{label}; students_unassigned={member_count}; exam_sessions_removed={session_count}; practical_data_preserved=1')
    s.delete(row);s.commit()
    flash(f'{label} deleted. {member_count} student account'+(' was' if member_count==1 else 's were')+' left unassigned; practical lists and practical marks were not changed.')
    return redirect(url_for('academic_groups'))

@app.route('/admin/question-bank',methods=['GET','POST'])
@staff_required
def question_bank():
    s=DB()
    seed_subject_catalog(s)
    if request.method=='POST':
        question=request.form.get('question','').strip();qdef=question_definition_from_form(request.form)
        subject_name=request.form.get('subject','').strip()
        catalog_subject=s.scalar(select(SubjectCatalog).where(SubjectCatalog.name==subject_name,SubjectCatalog.is_active==True)) if subject_name else None
        duplicate=s.scalar(select(BankQuestion).where(BankQuestion.subject==subject_name,func.lower(BankQuestion.question)==question.lower())) if question and subject_name else None
        if qdef['error']:flash(qdef['error'],'error')
        elif not catalog_subject:flash('Choose a subject from the Subject Catalog, or add your custom subject first.','error')
        elif duplicate:flash(f'A duplicate question already exists in this subject (Question #{duplicate.id}).','error')
        else:
            try:marks=max(1,int(request.form.get('marks','1')))
            except ValueError:marks=1
            status='approved' if can_approve_content(s) and request.form.get('status')=='approved' else 'draft'
            course_semester=request.form.get('course_semester','').strip() or catalog_subject.course_semester
            visibility=normalize_practice_visibility(request.form.get('practice_visibility'))
            practical_experiment_no=normalize_practical_exam_no(request.form.get('practical_experiment_no')) if visibility=='practical_exam' else ''
            if visibility=='practical_exam' and not practical_experiment_no:
                flash('Enter the Practical Experiment No. before saving a Practical Exam question.','error')
                return redirect(url_for('question_bank',subject=catalog_subject.name)+'#add-bank-question')
            opts=qdef['options']
            bq=BankQuestion(subject=catalog_subject.name,course_semester=course_semester,unit=request.form.get('unit','').strip(),topic=request.form.get('topic','').strip(),question_type=qdef['question_type'],question=question,
                option_a=opts['A'],option_b=opts['B'],option_c=opts['C'],option_d=opts['D'],correct_answer=qdef['legacy_correct_answer'],answer_key=qdef['answer_key'],answer_tolerance=qdef['answer_tolerance'],answer_case_sensitive=qdef['answer_case_sensitive'],marks=marks,
                difficulty=canonical_difficulty(request.form.get('difficulty')),bloom_level=canonical_bloom(request.form.get('bloom_level')),co_mapping=request.form.get('co_mapping','').strip(),po_mapping=request.form.get('po_mapping','').strip(),pso_mapping=request.form.get('pso_mapping','').strip(),tags=request.form.get('tags','').strip(),practice_visibility=visibility,practical_experiment_no=practical_experiment_no,explanation=request.form.get('explanation','').strip(),status=status,version=1,created_by=actor_label(s),created_at=now_iso(),updated_at=now_iso())
            s.add(bq);s.flush();audit_event(s,'bank_question_created','bank_question',bq.id,f'status={status}, type={qdef["question_type"]}, subject={catalog_subject.name}, category={catalog_subject.category}, visibility={visibility}, practical_experiment={practical_experiment_no}');s.commit();flash('Question added to the bank.')
            return redirect(url_for('question_bank',subject=catalog_subject.name)+'#subject-workspace')
    q=(request.args.get('q') or '').strip(); category=(request.args.get('category') or '').strip(); subject=(request.args.get('subject') or '').strip(); unit=(request.args.get('unit') or '').strip(); difficulty=(request.args.get('difficulty') or '').strip(); status=(request.args.get('status') or '').strip(); practice_visibility=normalize_practice_visibility(request.args.get('practice_visibility')) if request.args.get('practice_visibility') else ''
    catalog_rows=s.scalars(select(SubjectCatalog).where(SubjectCatalog.is_active==True).order_by(SubjectCatalog.category,SubjectCatalog.name)).all()
    catalog_map={row.name:row for row in catalog_rows}
    stmt=select(BankQuestion)
    if q: stmt=stmt.where(or_(BankQuestion.question.ilike(f'%{q}%'),BankQuestion.topic.ilike(f'%{q}%'),BankQuestion.tags.ilike(f'%{q}%')))
    if category:
        category_subjects=[row.name for row in catalog_rows if row.category==category]
        stmt=stmt.where(BankQuestion.subject.in_(category_subjects)) if category_subjects else stmt.where(BankQuestion.id==-1)
    if subject: stmt=stmt.where(BankQuestion.subject==subject)
    if unit: stmt=stmt.where(BankQuestion.unit==unit)
    if difficulty: stmt=stmt.where(BankQuestion.difficulty==canonical_difficulty(difficulty))
    if status: stmt=stmt.where(BankQuestion.status==status)
    if practice_visibility: stmt=stmt.where(BankQuestion.practice_visibility==practice_visibility)
    rows=s.scalars(stmt.order_by(BankQuestion.id.desc())).all()
    units=s.scalars(select(BankQuestion.unit).where(BankQuestion.unit!='').distinct().order_by(BankQuestion.unit)).all(); exams_list=s.scalars(select(Exam).order_by(Exam.id.desc())).all()
    usage=dict(s.execute(select(ExamBankMap.bank_question_id,func.count(func.distinct(ExamBankMap.exam_id))).group_by(ExamBankMap.bank_question_id)).all())
    question_counts=dict(s.execute(select(BankQuestion.subject,func.count()).group_by(BankQuestion.subject)).all())
    preloaded_packs=preloaded_pack_statuses(s)
    preloaded_categories=sorted({p.get('category','General') for p in preloaded_packs})
    catalog_groups=subject_catalog_groups(s)
    catalog_categories=sorted({row.category for row in catalog_rows})
    selected_catalog_subject=catalog_map.get(subject) if subject else None
    selected_subject_stats=None
    selected_subject_units=[]
    if selected_catalog_subject:
        approved_count=s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.subject==selected_catalog_subject.name,BankQuestion.status=='approved')) or 0
        official_count=s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.subject==selected_catalog_subject.name,BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['official_only','both']))) or 0
        practice_count=s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.subject==selected_catalog_subject.name,BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both']))) or 0
        draft_count=s.scalar(select(func.count()).select_from(BankQuestion).where(BankQuestion.subject==selected_catalog_subject.name,BankQuestion.status=='draft')) or 0
        selected_subject_stats={'total':approved_count+draft_count,'approved':approved_count,'official':official_count,'practice':practice_count,'draft':draft_count}
        selected_subject_units=[str(value).strip() for value in s.scalars(
            select(BankQuestion.unit).where(
                BankQuestion.subject==selected_catalog_subject.name,
                BankQuestion.status=='approved',
                BankQuestion.practice_visibility.in_(['official_only','both']),
                BankQuestion.unit!=''
            ).distinct()
        ).all() if str(value or '').strip()]
        selected_subject_units.sort(key=lambda value:(int(value) if value.isdigit() else 10**9,value.casefold()))
    target_exam_id=request.args.get('target_exam_id',type=int)
    return render_template('question_bank.html',questions=rows,subjects=catalog_rows,subject_groups=catalog_groups,catalog_categories=catalog_categories,catalog_map=catalog_map,question_counts=question_counts,selected_catalog_subject=selected_catalog_subject,selected_subject_stats=selected_subject_stats,selected_subject_units=selected_subject_units,units=units,exams=exams_list,usage=usage,target_exam_id=target_exam_id,filters={'q':q,'category':category,'subject':subject,'unit':unit,'difficulty':difficulty,'status':status,'practice_visibility':practice_visibility},preloaded_packs=preloaded_packs,preloaded_categories=preloaded_categories)

@app.route('/admin/question-bank/subjects',methods=['POST'])
@staff_required
def add_subject_catalog():
    s=DB()
    name=(request.form.get('subject_name') or '').strip()
    category=(request.form.get('category') or '').strip()
    course=(request.form.get('course_semester') or '').strip()
    if not name:
        flash('Subject name is required.','error'); return redirect(url_for('question_bank'))
    if not category:
        flash('Assign a category before adding the subject.','error'); return redirect(url_for('question_bank'))
    existing=s.scalar(select(SubjectCatalog).where(func.lower(SubjectCatalog.name)==name.lower()))
    if existing:
        existing.category=category; existing.course_semester=course or existing.course_semester; existing.is_active=True
        audit_event(s,'subject_catalog_updated','subject',existing.id,f'name={existing.name}, category={category}')
        s.commit(); flash(f'Updated subject “{existing.name}” under {category}.')
        return redirect(url_for('question_bank',subject=existing.name)+'#subject-workspace')
    row=ensure_subject_catalog_entry(s,name,category,course,actor_label(s)); s.flush()
    audit_event(s,'subject_catalog_created','subject',row.id,f'name={row.name}, category={row.category}')
    s.commit(); flash(f'Added “{row.name}” under {row.category}. It is now available in the question form.')
    return redirect(url_for('question_bank',subject=row.name)+'#subject-workspace')

@app.route('/admin/question-bank/subjects/<int:subject_id>/toggle',methods=['POST'])
@approver_required
def toggle_subject_catalog(subject_id):
    s=DB(); row=s.get(SubjectCatalog,subject_id)
    if not row: abort(404)
    row.is_active=not row.is_active
    audit_event(s,'subject_catalog_activated' if row.is_active else 'subject_catalog_deactivated','subject',row.id,f'name={row.name}, category={row.category}')
    s.commit(); flash(f'{row.name} is now '+('active.' if row.is_active else 'hidden from new-question selection.'))
    return redirect(url_for('question_bank'))

@app.route('/admin/question-bank/subjects/<int:subject_id>/create-exam',methods=['POST'])
@staff_required
def create_exam_from_catalog_subject(subject_id):
    """Create an inactive draft exam from the approved questions of a catalog subject."""
    s=DB(); subject=s.get(SubjectCatalog,subject_id)
    if not subject: abort(404)
    if not subject.is_active:
        flash('This subject is not active in the Subject Catalog.','error')
        return redirect(url_for('question_bank'))

    selected_unit=(request.form.get('unit') or '').strip()
    bank_stmt=select(BankQuestion).where(
        BankQuestion.subject==subject.name,
        BankQuestion.status=='approved',
        BankQuestion.practice_visibility.in_(['official_only','both'])
    )
    if selected_unit:
        bank_stmt=bank_stmt.where(BankQuestion.unit==selected_unit)
    bank_rows=s.scalars(bank_stmt.order_by(BankQuestion.unit,BankQuestion.id)).all()
    if not bank_rows:
        scope=f' Unit {selected_unit}' if selected_unit else ''
        flash(f'Add and approve at least one question for {subject.name}{scope} before creating an exam.','error')
        return redirect(url_for('question_bank',subject=subject.name)+'#subject-workspace')

    try:
        duration=max(1,int(request.form.get('duration','20')))
    except ValueError:
        duration=20
    try:
        per_student=max(1,int(request.form.get('question_count','10')))
    except ValueError:
        per_student=10
    per_student=min(per_student,len(bank_rows))

    requested_title=(request.form.get('exam_title') or '').strip()
    if requested_title:
        subject_prefix=re.compile(r'^'+re.escape(subject.name)+r'\s*(?:[-–—:|]\s*)+',re.IGNORECASE)
        short_title=subject_prefix.sub('',requested_title,count=1).strip() or requested_title
    elif selected_unit:
        short_title=f'Unit {selected_unit} - Ready Exam'
    else:
        short_title='Ready Exam'
    exam=Exam(title=f"{subject.name} - {short_title}",duration_minutes=duration,is_active=False,created_at=now_iso())
    s.add(exam); s.flush()
    cfg=get_exam_config(s,exam.id,create=True)
    cfg.subject=subject.name
    cfg.course_semester=subject.course_semester or ''
    cfg.question_count=per_student
    cfg.pool_size=len(bank_rows)

    easy=sum(1 for q in bank_rows if canonical_difficulty(q.difficulty)=='Easy')
    medium=sum(1 for q in bank_rows if canonical_difficulty(q.difficulty)=='Medium')
    total=max(1,len(bank_rows))
    cfg.easy_pct=round(easy*100/total)
    cfg.medium_pct=round(medium*100/total)
    cfg.hard_pct=max(0,100-cfg.easy_pct-cfg.medium_pct)
    units=sorted({(q.unit or '').strip() for q in bank_rows if (q.unit or '').strip()})
    cfg.unit_weights=json.dumps({u:1 for u in units},ensure_ascii=False)
    cfg.randomize_questions=True
    cfg.shuffle_options=True
    cfg.require_fullscreen=False
    cfg.tab_switch_limit=3
    source_label=f'{subject.name} / Unit {selected_unit}' if selected_unit else subject.name
    cfg.last_generation_summary=f'Created from {source_label}: {len(bank_rows)} approved bank questions; each student receives {per_student}.'
    cfg.updated_at=now_iso()
    get_exam_approval(s,exam.id,create=True)

    for bq in bank_rows:
        copy_bank_question_to_exam(s,bq,exam.id)

    audit_event(s,'catalog_subject_exam_created','exam',exam.id,f'subject={subject.name}, unit={selected_unit or "all"}, pool={len(bank_rows)}, per_student={per_student}, title={short_title}')
    s.commit()
    flash(f'Created “{short_title}” with {len(bank_rows)} approved question(s) in the pool and {per_student} question(s) per student.')
    return redirect(url_for('exam_builder', exam_id=exam.id))


@app.route('/admin/question-bank/preloaded/export/<fmt>')
@staff_required
def export_preloaded_question_banks(fmt):
    headers=['category','subject','course_semester','unit','topic','question','option_a','option_b','option_c','option_d','correct_answer','marks','difficulty','bloom_level','co_mapping','tags']
    rows=[]
    for pack in load_preloaded_question_banks().values():
        for q in pack.get('questions',[]):
            row=dict(q)
            row['category']=pack.get('category','Engineering')
            row['subject']=row.get('subject') or pack.get('subject','')
            row['course_semester']=row.get('course_semester') or pack.get('course_semester','')
            rows.append([row.get(h,'') for h in headers])
    if fmt=='csv':
        out=io.StringIO(newline=''); writer=csv.writer(out); writer.writerow(headers); writer.writerows(rows)
        data=io.BytesIO(out.getvalue().encode('utf-8-sig'))
        return send_file(data,mimetype='text/csv',as_attachment=True,download_name='preloaded_engineering_question_banks.csv')
    if fmt=='xlsx':
        wb=Workbook(); ws=wb.active; ws.title='Question Banks'; ws.append(headers)
        for cell in ws[1]: cell.font=Font(bold=True)
        for row in rows: ws.append(row)
        widths={'A':22,'B':34,'C':24,'D':8,'E':28,'F':58,'G':45,'H':45,'I':45,'J':45,'K':14,'L':9,'M':12,'N':16,'O':12,'P':34}
        for col,width in widths.items(): ws.column_dimensions[col].width=width
        data=io.BytesIO(); wb.save(data); data.seek(0)
        return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name='preloaded_engineering_question_banks.xlsx')
    abort(404)

@app.route('/admin/question-bank/preloaded/<slug>/activate',methods=['POST'])
@staff_required
def activate_preloaded_question_bank(slug):
    pack=load_preloaded_question_banks().get(slug)
    if not pack: abort(404)
    s=DB()
    try:
        added,existing=activate_preloaded_pack(s,pack)
        audit_event(s,'preloaded_question_bank_activated','question_bank',slug,f'subject={pack["subject"]}, added={added}, existing={existing}')
        s.commit()
    except Exception as exc:
        s.rollback(); flash(f'Could not activate the preloaded bank: {exc}','error')
        return redirect(url_for('question_bank'))
    if added:
        flash(f'Activated {pack["subject"]}: {added} approved questions across {pack.get("unit_count",5)} units.')
    else:
        flash(f'{pack["subject"]} is already activated.')
    return redirect(url_for('question_bank',subject=pack['subject'],status='approved'))

@app.route('/admin/question-bank/preloaded/category/activate',methods=['POST'])
@staff_required
def activate_preloaded_category():
    category=(request.form.get('category') or '').strip()
    packs=[p for p in load_preloaded_question_banks().values() if p.get('category')==category]
    if not packs:
        flash('No preloaded subject banks were found for that category.','error')
        return redirect(url_for('question_bank'))
    s=DB(); total=0
    try:
        for pack in packs:
            added,_=activate_preloaded_pack(s,pack); total+=added
        audit_event(s,'preloaded_category_activated','question_bank','',f'category={category}, added={total}')
        s.commit()
    except Exception as exc:
        s.rollback(); flash(f'Could not activate the category: {exc}','error')
        return redirect(url_for('question_bank'))
    flash(f'Activated {category}: {total} new approved questions from {len(packs)} subject bank(s).')
    return redirect(url_for('question_bank'))

@app.route('/admin/question-bank/preloaded/<slug>/quick-exam',methods=['POST'])
@staff_required
def create_exam_from_preloaded_bank(slug):
    pack=load_preloaded_question_banks().get(slug)
    if not pack: abort(404)
    s=DB()
    try:
        added,_=activate_preloaded_pack(s,pack)
        if added: s.flush()
        marker=f'preloaded:{slug}'
        bank_rows=s.scalars(select(BankQuestion).where(BankQuestion.created_by==marker,BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['official_only','both'])).order_by(BankQuestion.unit,BankQuestion.id)).all()
        if not bank_rows:
            flash('This subject bank has no approved questions.','error'); s.rollback()
            return redirect(url_for('question_bank'))
        title=f"{pack['subject']} - Ready Exam"
        duration=20
        exam=Exam(title=title,duration_minutes=duration,is_active=False,created_at=now_iso())
        s.add(exam); s.flush()
        cfg=get_exam_config(s,exam.id,create=True)
        cfg.subject=pack['subject']; cfg.course_semester=pack.get('course_semester','')
        cfg.question_count=min(10,len(bank_rows)); cfg.pool_size=len(bank_rows)
        cfg.easy_pct=30; cfg.medium_pct=40; cfg.hard_pct=30
        unit_numbers=sorted({(q.unit or '').strip() for q in bank_rows if (q.unit or '').strip()})
        cfg.unit_weights=json.dumps({u:1 for u in unit_numbers},ensure_ascii=False)
        cfg.randomize_questions=True; cfg.shuffle_options=True; cfg.require_fullscreen=False; cfg.tab_switch_limit=3
        cfg.last_generation_summary=f"Ready-made pool from {pack['subject']}: {len(bank_rows)} approved questions; each student receives {cfg.question_count}."
        cfg.updated_at=now_iso()
        for bq in bank_rows:
            copy_bank_question_to_exam(s,bq,exam.id)
        audit_event(s,'ready_exam_created','exam',exam.id,f'preloaded={slug}, pool={len(bank_rows)}, per_student={cfg.question_count}')
        s.commit()
        flash(f'Created "{title}" with {len(bank_rows)}-question pool. Each student will receive {cfg.question_count} randomized questions. Review it, then click Activate.')
        return redirect(url_for('exams'))
    except Exception as exc:
        s.rollback(); flash(f'Could not create the ready exam: {exc}','error')
        return redirect(url_for('question_bank'))

@app.route('/admin/question-bank/preloaded/<slug>/unit-set/<unit>/<set_code>/quick-exam',methods=['POST'])
@staff_required
def create_unit_set_exam_from_preloaded_bank(slug,unit,set_code):
    pack=load_preloaded_question_banks().get(slug)
    if not pack: abort(404)
    set_code=str(set_code or '').upper()
    unit=str(unit or '').strip()
    configured=next((p for p in pack.get('paper_sets',[]) if str(p.get('unit'))==unit and str(p.get('set_code','')).upper()==set_code),None)
    if not configured:
        abort(404)
    s=DB()
    try:
        added,_=activate_preloaded_pack(s,pack)
        if added: s.flush()
        marker=f'preloaded:{slug}'
        set_marker=f'paper-set-{set_code.lower()}'
        bank_rows=s.scalars(select(BankQuestion).where(
            BankQuestion.created_by==marker,
            BankQuestion.status=='approved',
            BankQuestion.practice_visibility.in_(['official_only','both']),
            BankQuestion.unit==unit,
            BankQuestion.tags.contains(set_marker)
        ).order_by(BankQuestion.id)).all()
        expected=int(configured.get('question_count') or 0)
        if expected and len(bank_rows)!=expected:
            flash(f'Unit {unit} Set {set_code} is incomplete: expected {expected} questions but found {len(bank_rows)}.','error')
            s.rollback(); return redirect(url_for('question_bank'))
        if not bank_rows:
            flash(f'No approved questions were found for Unit {unit} Set {set_code}.','error')
            s.rollback(); return redirect(url_for('question_bank'))
        title=f"{pack['subject']} - Unit {unit} - Set {set_code}"
        duration=max(5,int(configured.get('duration_minutes') or 15))
        exam=Exam(title=title,duration_minutes=duration,is_active=False,created_at=now_iso())
        s.add(exam); s.flush()
        cfg=get_exam_config(s,exam.id,create=True)
        cfg.subject=pack['subject']; cfg.course_semester=pack.get('course_semester','')
        cfg.question_count=len(bank_rows); cfg.pool_size=len(bank_rows)
        easy=sum(1 for x in bank_rows if canonical_difficulty(x.difficulty)=='Easy')
        medium=sum(1 for x in bank_rows if canonical_difficulty(x.difficulty)=='Medium')
        hard=sum(1 for x in bank_rows if canonical_difficulty(x.difficulty)=='Hard')
        total=max(1,len(bank_rows))
        cfg.easy_pct=round(easy*100/total); cfg.medium_pct=round(medium*100/total); cfg.hard_pct=max(0,100-cfg.easy_pct-cfg.medium_pct)
        cfg.unit_weights=json.dumps({unit:1},ensure_ascii=False)
        cfg.randomize_questions=False; cfg.shuffle_options=False; cfg.require_fullscreen=False; cfg.tab_switch_limit=3
        cfg.last_generation_summary=f"Prepared Unit {unit} Set {set_code}: {len(bank_rows)} fixed syllabus-aligned MCQs."
        cfg.updated_at=now_iso()
        for bq in bank_rows:
            copy_bank_question_to_exam(s,bq,exam.id)
        audit_event(s,'unit_set_exam_created','exam',exam.id,f'preloaded={slug}, unit={unit}, set={set_code}, questions={len(bank_rows)}')
        s.commit()
        flash(f'Created "{title}" with {len(bank_rows)} fixed questions. Review it, then click Activate.')
        return redirect(url_for('exams'))
    except Exception as exc:
        s.rollback(); flash(f'Could not create Unit {unit} Set {set_code}: {exc}','error')
        return redirect(url_for('question_bank'))

@app.route('/admin/question-bank/import',methods=['POST'])
@staff_required
def import_question_bank():
    upload=request.files.get('question_file')
    target_subject=(request.form.get('target_subject') or '').strip()
    redirect_url=url_for('question_bank',subject=target_subject)+'#subject-workspace' if target_subject else url_for('question_bank')
    if not upload or not upload.filename: flash('Choose a CSV or Excel (.xlsx) file.','error');return redirect(redirect_url)
    try: headers,rows=_rows_from_upload(upload)
    except ValueError as exc: flash(str(exc),'error');return redirect(redirect_url)
    s=DB()
    target_catalog=None
    if target_subject:
        target_catalog=s.scalar(select(SubjectCatalog).where(SubjectCatalog.name==target_subject,SubjectCatalog.is_active==True))
        if not target_catalog:
            flash('The selected Subject Catalog entry is not active.','error');return redirect(url_for('question_bank'))
    required={'question','marks'}
    if not target_catalog:required.add('subject')
    if not required.issubset(set(headers)):
        flash('Question bank file is missing required columns. Download the new template and try again.','error');return redirect(redirect_url)
    added=invalid=duplicates=0
    for r in rows:
        question=(r.get('question') or '').strip();qtype=canonical_question_type(r.get('question_type') or 'single_choice')
        options={key:(r.get(f'option_{key.lower()}') or '').strip() for key in 'ABCD'}
        if qtype=='single_choice':answer_key=(r.get('answer_key') or r.get('correct_answer') or '').strip()
        elif qtype=='multiple_select':answer_key=(r.get('answer_key') or r.get('correct_answer') or '').strip()
        elif qtype=='true_false':answer_key=(r.get('answer_key') or r.get('correct_answer') or '').strip()
        else:answer_key=(r.get('answer_key') or '').strip()
        tolerance=(r.get('answer_tolerance') or '').strip()
        error=validate_question_definition(qtype,question,options,answer_key,tolerance)
        if error:invalid+=1;continue
        try:marks=max(1,int(r.get('marks') or 1))
        except ValueError:marks=1
        requested_status=(r.get('status') or 'draft').lower();status='approved' if can_approve_content(s) and requested_status=='approved' else 'draft'
        if target_catalog:
            subject_name=target_catalog.name;course=(r.get('course_semester') or '').strip() or target_catalog.course_semester;category=target_catalog.category
        else:
            subject_name=(r.get('subject') or 'General').strip() or 'General';course=(r.get('course_semester') or '').strip();category=(r.get('category') or '').strip() or 'Imported / Other';ensure_subject_catalog_entry(s,subject_name,category,course,actor_label(s))
        if s.scalar(select(BankQuestion.id).where(BankQuestion.subject==subject_name,func.lower(BankQuestion.question)==question.lower())):
            duplicates+=1;continue
        legacy=answer_key.strip().upper() if qtype=='single_choice' else 'A';legacy=legacy if legacy in {'A','B','C','D'} else 'A'
        case_sensitive=str(r.get('answer_case_sensitive') or '').strip().lower() in {'1','true','yes','y','on'}
        visibility=normalize_practice_visibility(r.get('practice_visibility'));practical_experiment_no=normalize_practical_exam_no(r.get('practical_experiment_no')) if visibility=='practical_exam' else ''
        if visibility=='practical_exam' and not practical_experiment_no:
            invalid+=1;continue
        s.add(BankQuestion(subject=subject_name,course_semester=course,unit=(r.get('unit') or '').strip(),topic=(r.get('topic') or '').strip(),question_type=qtype,question=question,option_a=options['A'],option_b=options['B'],option_c=options['C'],option_d=options['D'],correct_answer=legacy,answer_key=answer_key,answer_tolerance=tolerance,answer_case_sensitive=case_sensitive,marks=marks,difficulty=canonical_difficulty(r.get('difficulty')),bloom_level=canonical_bloom(r.get('bloom_level')),co_mapping=(r.get('co_mapping') or '').strip(),po_mapping=(r.get('po_mapping') or '').strip(),pso_mapping=(r.get('pso_mapping') or '').strip(),tags=(r.get('tags') or '').strip(),practice_visibility=visibility,practical_experiment_no=practical_experiment_no,explanation=(r.get('explanation') or '').strip(),status=status,version=1,created_by=actor_label(s),created_at=now_iso(),updated_at=now_iso()));added+=1
    audit_event(s,'question_bank_bulk_import','bank_question','',f'added={added}, invalid={invalid}, duplicates={duplicates}, target_subject={target_subject or "mixed"}')
    s.commit();flash(f'Imported {added} question(s).'+(f' Skipped {invalid} invalid row(s).' if invalid else '')+(f' Skipped {duplicates} duplicate(s).' if duplicates else ''));return redirect(redirect_url)

@app.route('/admin/question-bank/template/<fmt>')
@staff_required
def question_bank_template(fmt):
    headers=['category','subject','course_semester','unit','topic','question_type','question','option_a','option_b','option_c','option_d','correct_answer','answer_key','answer_tolerance','answer_case_sensitive','marks','difficulty','bloom_level','co_mapping','po_mapping','pso_mapping','tags','practice_visibility','practical_experiment_no','explanation','status']
    subject_name=(request.args.get('subject') or '').strip()
    category='Computer Science'; course='B.Tech CSE / Sem 5'
    if subject_name:
        s=DB(); row=s.scalar(select(SubjectCatalog).where(SubjectCatalog.name==subject_name,SubjectCatalog.is_active==True))
        if row: category=row.category; course=row.course_semester; subject_name=row.name
        else: subject_name=''
    example=[category,subject_name or 'Mobile Application Development',course or 'B.Tech CSE / Sem 5','1','Introduction','single_choice','Replace this sample with your question','Option A','Option B','Option C','Option D','A','A','','false','1','Medium','Understand','CO1','PO1','PSO1','custom','official_only','','Explain why option A is correct.','approved']
    safe_name=''.join(ch if ch.isalnum() else '_' for ch in (subject_name or 'question_bank')).strip('_') or 'question_bank'
    if fmt=='csv':
        out=io.StringIO(newline='');writer=csv.writer(out);writer.writerow(headers);writer.writerow(example);data=io.BytesIO(out.getvalue().encode('utf-8-sig'));return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'{safe_name}_question_bank_template.csv')
    if fmt=='xlsx':
        wb=Workbook();ws=wb.active;ws.title='Question Bank';ws.append(headers);ws.append(example)
        for cell in ws[1]:cell.font=Font(bold=True)
        for idx,width in enumerate([22,28,24,10,24,18,58,26,26,26,26,15,20,16,18,10,14,16,14,14,14,26,18,46,12],1):ws.column_dimensions[chr(64+idx) if idx<=26 else 'A'].width=width
        data=io.BytesIO();wb.save(data);data.seek(0);return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=f'{safe_name}_question_bank_template.xlsx')
    abort(404)

@app.route('/admin/question-bank/<int:question_id>/edit',methods=['GET','POST'])
@staff_required
def edit_bank_question(question_id):
    s=DB();q=s.get(BankQuestion,question_id)
    if not q:abort(404)
    revisions=s.scalars(select(QuestionRevision).where(QuestionRevision.bank_question_id==q.id).order_by(QuestionRevision.id.desc())).all()
    if request.method=='POST':
        snapshot={c.name:getattr(q,c.name) for c in BankQuestion.__table__.columns if c.name not in {'id','updated_at'}}
        s.add(QuestionRevision(bank_question_id=q.id,version=q.version,snapshot_json=json.dumps(snapshot,ensure_ascii=False),changed_by=actor_label(s),changed_at=now_iso()))
        qdef=question_definition_from_form(request.form)
        if qdef['error']:flash(qdef['error'],'error');return redirect(url_for('edit_bank_question',question_id=q.id))
        subject_name=request.form.get('subject','').strip();catalog_subject=s.scalar(select(SubjectCatalog).where(SubjectCatalog.name==subject_name,SubjectCatalog.is_active==True));q.subject=catalog_subject.name if catalog_subject else q.subject;q.course_semester=request.form.get('course_semester','').strip() or (catalog_subject.course_semester if catalog_subject else q.course_semester);q.unit=request.form.get('unit','').strip();q.topic=request.form.get('topic','').strip();q.question=request.form.get('question','').strip();q.question_type=qdef['question_type'];q.option_a=qdef['options']['A'];q.option_b=qdef['options']['B'];q.option_c=qdef['options']['C'];q.option_d=qdef['options']['D'];q.correct_answer=qdef['legacy_correct_answer'];q.answer_key=qdef['answer_key'];q.answer_tolerance=qdef['answer_tolerance'];q.answer_case_sensitive=qdef['answer_case_sensitive']
        try:q.marks=max(1,int(request.form.get('marks','1')))
        except ValueError:q.marks=1
        visibility=normalize_practice_visibility(request.form.get('practice_visibility'));practical_experiment_no=normalize_practical_exam_no(request.form.get('practical_experiment_no')) if visibility=='practical_exam' else ''
        if visibility=='practical_exam' and not practical_experiment_no:
            s.rollback();flash('Enter the Practical Experiment No. before saving a Practical Exam question.','error');return redirect(url_for('edit_bank_question',question_id=q.id))
        q.difficulty=canonical_difficulty(request.form.get('difficulty'));q.bloom_level=canonical_bloom(request.form.get('bloom_level'));q.co_mapping=request.form.get('co_mapping','').strip();q.po_mapping=request.form.get('po_mapping','').strip();q.pso_mapping=request.form.get('pso_mapping','').strip();q.tags=request.form.get('tags','').strip();q.practice_visibility=visibility;q.practical_experiment_no=practical_experiment_no;q.explanation=request.form.get('explanation','').strip();q.version+=1;q.updated_at=now_iso()
        if can_approve_content(s):q.status=request.form.get('status','draft') if request.form.get('status') in {'draft','approved'} else 'draft'
        else:q.status='draft'
        audit_event(s,'bank_question_edited','bank_question',q.id,f'version={q.version}, status={q.status}');s.commit();flash('Question updated. Previous version was preserved.');return redirect(url_for('edit_bank_question',question_id=q.id))
    seed_subject_catalog(s); return render_template('question_bank_edit.html',question=q,question_type=canonical_question_type(q.question_type),revisions=revisions,subject_groups=subject_catalog_groups(s))

@app.route('/admin/question-bank/<int:question_id>/approve',methods=['POST'])
@approver_required
def approve_bank_question(question_id):
    s=DB();q=s.get(BankQuestion,question_id)
    if not q:abort(404)
    q.status='approved';q.updated_at=now_iso();audit_event(s,'bank_question_approved','bank_question',q.id,q.subject);s.commit();flash('Question approved.');return redirect(request.referrer or url_for('question_bank'))

@app.route('/admin/question-bank/add-to-exam',methods=['POST'])
@staff_required
def bank_add_to_exam():
    try:exam_id=int(request.form.get('exam_id','0'))
    except ValueError:exam_id=0
    ids=[]
    for v in request.form.getlist('question_ids'):
        try:ids.append(int(v))
        except ValueError:pass
    s=DB();exam=s.get(Exam,exam_id)
    if not exam or not ids:
        flash('Select an exam and at least one bank question.','error')
        return redirect(url_for('question_bank'))

    cfg=get_exam_config(s,exam_id,create=False)
    stmt=select(BankQuestion).where(
        BankQuestion.id.in_(ids),
        BankQuestion.status=='approved',
        BankQuestion.practice_visibility.in_(['official_only','both','practical_exam'])
    )
    # If the exam belongs to an existing subject, prevent accidental
    # cross-subject question mixing. Questions are never copied to any other exam.
    if cfg and (cfg.subject or '').strip():
        stmt=stmt.where(BankQuestion.subject==cfg.subject.strip())
    rows=s.scalars(stmt).all();added=0
    incoming_practical=[q for q in rows if normalize_practice_visibility(q.practice_visibility)=='practical_exam']
    incoming_regular=[q for q in rows if normalize_practice_visibility(q.practice_visibility)!='practical_exam']
    existing_questions=s.scalars(select(Question).where(Question.exam_id==exam_id)).all()
    existing_practical=[q for q in existing_questions if normalize_practical_exam_no(q.practical_experiment_no)]
    existing_regular=[q for q in existing_questions if not normalize_practical_exam_no(q.practical_experiment_no)]
    if incoming_practical:
        experiment_numbers={normalize_practical_exam_no(q.practical_experiment_no) for q in incoming_practical if normalize_practical_exam_no(q.practical_experiment_no)}
        missing_mapping=any(not normalize_practical_exam_no(q.practical_experiment_no) for q in incoming_practical)
        existing_numbers={normalize_practical_exam_no(q.practical_experiment_no) for q in existing_practical if normalize_practical_exam_no(q.practical_experiment_no)}
        if incoming_regular or existing_regular or missing_mapping or len(experiment_numbers)!=1 or (existing_numbers and existing_numbers!=experiment_numbers):
            flash('Practical Exam questions must be the only questions in the exam and must all use the same Practical Experiment No.','error')
            return redirect(url_for('question_bank',subject=(cfg.subject if cfg else ''),target_exam_id=exam_id)+'#bank-questions')
    elif existing_practical and incoming_regular:
        flash('This exam is already mapped as a Practical Exam. Add only Practical Exam questions for the same experiment number.','error')
        return redirect(url_for('question_bank',subject=(cfg.subject if cfg else ''),target_exam_id=exam_id)+'#bank-questions')
    for q in rows:
        if copy_bank_question_to_exam(s,q,exam_id):added+=1

    pool_count=sync_manual_exam_question_count(s,exam_id)
    audit_event(s,'bank_questions_added_to_exam','exam',exam_id,f'added={added}, pool={pool_count}, per_student={pool_count}')
    s.commit()
    if added:
        flash(f'Added {added} approved question(s) to {exam.title}. New attempts will receive all {pool_count} question(s) in this exam.')
    else:
        flash('No new questions were added. The selected questions may already be in this exam or belong to another subject.','error')
    return redirect(url_for('questions',exam_id=exam_id))

@app.route('/admin/question-bank/practice-visibility',methods=['POST'])
@staff_required
def set_question_practice_visibility():
    visibility=normalize_practice_visibility(request.form.get('practice_visibility'))
    practical_experiment_no=normalize_practical_exam_no(request.form.get('practical_experiment_no')) if visibility=='practical_exam' else ''
    if visibility=='practical_exam' and not practical_experiment_no:
        flash('Enter the Practical Experiment No. before marking questions as Practical Exam.','error');return redirect(request.referrer or url_for('question_bank'))
    ids=[]
    for value in request.form.getlist('question_ids'):
        try:ids.append(int(value))
        except ValueError:pass
    if not ids:
        flash('Select at least one question to update its student-practice visibility.','error');return redirect(request.referrer or url_for('question_bank'))
    s=DB();rows=s.scalars(select(BankQuestion).where(BankQuestion.id.in_(ids))).all();updated=0;draft_count=0
    for row in rows:
        row.practice_visibility=visibility;row.practical_experiment_no=practical_experiment_no if visibility=='practical_exam' else '';row.updated_at=now_iso();updated+=1
        if row.status!='approved' and visibility in {'practice_only','both'}:draft_count+=1
    audit_event(s,'practice_visibility_bulk_updated','bank_question','',f'visibility={visibility}, practical_experiment={practical_experiment_no}, updated={updated}');s.commit()
    message=f'Updated {updated} question(s) to {PRACTICE_VISIBILITY_LABELS[visibility]}.'
    if draft_count:message+=f' {draft_count} draft question(s) remain hidden from students until approved.'
    flash(message);return redirect(request.referrer or url_for('question_bank'))


@app.route('/admin/exams/from-subject',methods=['POST'])
@staff_required
def create_exam_for_existing_subject():
    """Create a separate, empty exam under an existing subject.

    Question Bank rows are copied only when the admin explicitly selects them and
    uses Add Selected to Exam. This keeps multiple exams under one subject fully
    isolated from each other.
    """
    s=DB()
    try:
        subject_id=int(request.form.get('subject_id','0'))
    except ValueError:
        subject_id=0
    subject=s.get(SubjectCatalog,subject_id)
    if not subject or not subject.is_active:
        flash('Choose a valid existing subject.','error')
        return redirect(url_for('exams'))

    exam_title=(request.form.get('exam_title') or '').strip()
    if not exam_title:
        flash('Enter an exam title.','error')
        return redirect(url_for('exams'))

    selected_unit=(request.form.get('unit') or '').strip()
    try:
        duration=max(1,int(request.form.get('duration','20')))
    except ValueError:
        duration=20

    exam=Exam(title=exam_title,duration_minutes=duration,is_active=False,created_at=now_iso())
    s.add(exam);s.flush()
    cfg=get_exam_config(s,exam.id,create=True)
    cfg.subject=subject.name
    cfg.course_semester=subject.course_semester or ''
    cfg.question_count=0
    cfg.pool_size=0
    cfg.unit_weights=json.dumps({selected_unit:1},ensure_ascii=False) if selected_unit else ''
    cfg.randomize_questions=True
    cfg.shuffle_options=True
    cfg.require_fullscreen=False
    cfg.tab_switch_limit=3
    cfg.last_generation_summary=f'Manual exam created for {subject.name}' + (f' / Unit {selected_unit}' if selected_unit else '') + '; no questions copied automatically.'
    cfg.updated_at=now_iso()
    get_exam_approval(s,exam.id,create=True)
    audit_event(
        s,'catalog_subject_exam_created','exam',exam.id,
        f'manual_subject_exam=1, subject={subject.name}, unit={selected_unit or "all"}, pool=0, title={exam_title}'
    )
    s.commit()

    flash(f'Created “{exam_title}” under {subject.name}. Select only the questions you want and add them to this exam.')
    return redirect(url_for('question_bank',subject=subject.name,unit=selected_unit,target_exam_id=exam.id)+'#bank-questions')


@app.route('/admin/exams',methods=['GET','POST'])
@staff_required
def exams():
    s=DB()
    if request.method=='POST':
        try:duration=max(1,int(request.form.get('duration','30')))
        except ValueError:duration=30
        title=request.form.get('title','').strip()
        if title:
            e=Exam(title=title,duration_minutes=duration,is_active=False,created_at=now_iso());s.add(e);s.flush();get_exam_config(s,e.id,create=True);get_exam_approval(s,e.id,create=True);audit_event(s,'exam_created','exam',e.id,title);s.commit();flash('Exam created as draft.')
    subject_exam_options=[]
    catalog_subjects=s.scalars(select(SubjectCatalog).where(SubjectCatalog.is_active==True).order_by(SubjectCatalog.name)).all()
    for catalog_subject in catalog_subjects:
        unit_rows=s.execute(
            select(BankQuestion.unit,func.count(BankQuestion.id))
            .where(
                BankQuestion.subject==catalog_subject.name,
                BankQuestion.status=='approved',
                BankQuestion.practice_visibility.in_(['official_only','both','practical_exam'])
            )
            .group_by(BankQuestion.unit)
            .order_by(BankQuestion.unit)
        ).all()
        approved_count=sum(int(count or 0) for _unit,count in unit_rows)
        if not approved_count:
            continue
        units=[(unit or '').strip() for unit,count in unit_rows if (unit or '').strip() and int(count or 0)>0]
        subject_exam_options.append({
            'id':catalog_subject.id,
            'name':catalog_subject.name,
            'course_semester':catalog_subject.course_semester or '',
            'approved_count':approved_count,
            'units':units,
        })

    raw=s.execute(select(Exam,func.count(Question.id)).outerjoin(Question,Question.exam_id==Exam.id).group_by(Exam.id).order_by(Exam.id.desc())).all();rows=[]
    for e,count in raw:
        cfg=get_exam_config(s,e.id);target=(cfg.question_count if cfg and cfg.question_count else count)
        approval=get_exam_approval(s,e.id,create=True);session_count=s.scalar(select(func.count()).select_from(ExamSession).where(ExamSession.exam_id==e.id)) or 0
        policy=exam_approval_policy(s,e)
        subject,unit_label=student_exam_subject_unit(s,e,cfg)
        resolved_meta=practical_exam_metadata_for_exam(s,e.id)
        exam_type=('practical_exam' if resolved_meta.get('is_practical') else ((getattr(cfg,'exam_type','') or 'regular') if cfg else 'regular'))
        practical_experiment_no=(resolved_meta.get('experiment_no') or (getattr(cfg,'practical_experiment_no','') if cfg else '') or '')
        practical_code_start_at=(getattr(cfg,'practical_code_start_at','') or '') if cfg else ''
        practical_code_end_at=(getattr(cfg,'practical_code_end_at','') or '') if cfg else ''
        rows.append(type('ExamRow',(),{'id':e.id,'title':e.title,'duration_minutes':e.duration_minutes,'is_active':e.is_active,'question_count':count,'student_question_count':min(target,count) if count else 0,'approval_status':approval.status,'session_count':session_count,'self_approval_allowed':policy['self_approval_allowed'],'external_approval_required':policy['external_approval_required'],'approval_policy_message':policy['message'],'daily_exam_count':policy['daily_exam_count'],'subject':subject,'unit_label':unit_label,'exam_type':exam_type,'practical_experiment_no':practical_experiment_no,'practical_code_start_at':practical_code_start_at,'practical_code_end_at':practical_code_end_at})())

    grouped={}
    for row in rows:
        grouped.setdefault(row.subject,[]).append(row)
    exam_groups=[{'subject':subject,'exams':grouped[subject]} for subject in sorted(grouped,key=lambda name:(name in {'General','Mixed Subjects'},name.casefold()))]
    return render_template(
        'exams.html',exams=rows,exam_groups=exam_groups,
        subject_exam_options=subject_exam_options,catalog_subjects=catalog_subjects
    )

@app.route('/admin/exam/<int:exam_id>/edit-metadata',methods=['POST'])
@admin_required
def edit_exam_metadata(exam_id):
    s=DB()
    if current_staff_role(s)!='super_admin':
        abort(403)
    exam=s.get(Exam,exam_id)
    if not exam:
        abort(404)

    title=(request.form.get('title') or '').strip()
    if not title:
        flash('Exam title is required.','error')
        return redirect(url_for('exams'))

    exam_type=(request.form.get('exam_type') or 'regular').strip().lower()
    if exam_type not in {'regular','practical_exam'}:
        flash('Choose a valid exam type.','error')
        return redirect(url_for('exams'))
    practical_experiment_no=normalize_practical_exam_no(request.form.get('practical_experiment_no')) if exam_type=='practical_exam' else ''
    if exam_type=='practical_exam' and not practical_experiment_no:
        flash('Practical Experiment No. is required for a Practical Exam.','error')
        return redirect(url_for('exams'))

    practical_code_start_at=''
    practical_code_end_at=''
    if exam_type=='practical_exam':
        try:
            practical_code_start_at=parse_local_schedule(request.form.get('practical_code_start_at',''))
            practical_code_end_at=parse_local_schedule(request.form.get('practical_code_end_at',''))
        except ValueError:
            flash('Practical Code start/end time is invalid.','error')
            return redirect(url_for('exams'))
        if bool(practical_code_start_at)!=bool(practical_code_end_at):
            flash('Set both Practical Code start and end time, or leave both blank.','error')
            return redirect(url_for('exams'))
        if practical_code_start_at and practical_code_end_at:
            if datetime.fromisoformat(practical_code_end_at)<=datetime.fromisoformat(practical_code_start_at):
                flash('Practical Code end time must be after the start time.','error')
                return redirect(url_for('exams'))

    subject_value=(request.form.get('subject_id') or '').strip()
    if subject_value=='__general__':
        subject_name='General'
        course_semester=''
    else:
        try:
            subject_id=int(subject_value)
        except (TypeError,ValueError):
            flash('Choose a valid subject section.','error')
            return redirect(url_for('exams'))
        catalog_subject=s.get(SubjectCatalog,subject_id)
        if not catalog_subject or not catalog_subject.is_active:
            flash('The selected subject is not available.','error')
            return redirect(url_for('exams'))
        subject_name=catalog_subject.name
        course_semester=catalog_subject.course_semester or ''

    cfg=get_exam_config(s,exam.id,create=True)
    old_title=exam.title or ''
    old_subject,_old_unit=student_exam_subject_unit(s,exam,cfg)
    old_exam_type=(getattr(cfg,'exam_type','') or 'regular')
    old_experiment_no=(getattr(cfg,'practical_experiment_no','') or '')
    old_code_start=(getattr(cfg,'practical_code_start_at','') or '')
    old_code_end=(getattr(cfg,'practical_code_end_at','') or '')
    exam.title=title
    cfg.subject=subject_name
    if course_semester:
        cfg.course_semester=course_semester
    cfg.exam_type=exam_type
    cfg.practical_experiment_no=practical_experiment_no
    cfg.practical_code_start_at=practical_code_start_at if exam_type=='practical_exam' else ''
    cfg.practical_code_end_at=practical_code_end_at if exam_type=='practical_exam' else ''

    # Exam Question rows are exam-specific snapshots, so keep their practical
    # serial aligned with the explicit exam-level setting.  This makes later
    # additions/mixing checks safe and prevents a different experiment from
    # receiving Viva marks by mistake.
    exam_questions=s.scalars(select(Question).where(Question.exam_id==exam.id)).all()
    for question in exam_questions:
        question.practical_experiment_no=practical_experiment_no if exam_type=='practical_exam' else ''
    summary=(cfg.last_generation_summary or '').strip()
    if 'manual_subject_override=1' not in summary.lower():
        summary=(summary+'; ' if summary else '')+'manual_subject_override=1'
    cfg.last_generation_summary=summary
    cfg.updated_at=now_iso()
    audit_event(
        s,'exam_metadata_edited','exam',exam.id,
        f'title={old_title} -> {title}, subject={old_subject} -> {subject_name}, '
        f'type={old_exam_type} -> {exam_type}, experiment={old_experiment_no} -> {practical_experiment_no}, '
        f'code_window={old_code_start or "-"}..{old_code_end or "-"} -> '
        f'{cfg.practical_code_start_at or "-"}..{cfg.practical_code_end_at or "-"}'
    )
    s.commit()
    if exam_type=='practical_exam':
        # If this exam had already been submitted before it was converted to a
        # Practical Exam, repair those Viva entries immediately.
        resync_submitted_practical_attempts(s,exam.id)
        flash(f'Updated “{title}” as Practical Exam · Experiment {practical_experiment_no}.')
    else:
        flash(f'Updated “{title}” as Regular Exam.')
    return redirect(url_for('exams')+f'#exam-{exam.id}')

@app.route('/admin/exam/<int:exam_id>/delete',methods=['POST'])
@admin_required
def delete_exam(exam_id):
    # Permanently delete one exam and only data owned by that exam.
    # The reusable BankQuestion library is intentionally not touched.
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:
        abort(404)

    title=exam.title or f'Exam {exam.id}'
    question_ids=list(s.scalars(select(Question.id).where(Question.exam_id==exam.id)).all())
    attempt_ids=list(s.scalars(select(Attempt.id).where(Attempt.exam_id==exam.id)).all())
    edge_result_receipt_ids=list(s.scalars(
        select(EdgeResultReceipt.id).where(EdgeResultReceipt.target_exam_id==exam.id)
    ).all())

    answer_count=0
    if attempt_ids:
        answer_count=s.scalar(
            select(func.count()).select_from(Answer).where(Answer.attempt_id.in_(attempt_ids))
        ) or 0

        s.execute(delete(AttemptHeartbeat).where(AttemptHeartbeat.attempt_id.in_(attempt_ids)))
        s.execute(delete(IntegrityEvent).where(IntegrityEvent.attempt_id.in_(attempt_ids)))
        s.execute(delete(Answer).where(Answer.attempt_id.in_(attempt_ids)))
        s.execute(delete(AttemptQuestion).where(AttemptQuestion.attempt_id.in_(attempt_ids)))
        s.execute(delete(Attempt).where(Attempt.id.in_(attempt_ids)))

    if edge_result_receipt_ids:
        s.execute(delete(EdgeResultAttempt).where(
            EdgeResultAttempt.receipt_id.in_(edge_result_receipt_ids)
        ))
        s.execute(delete(EdgeResultReceipt).where(
            EdgeResultReceipt.id.in_(edge_result_receipt_ids)
        ))

    s.execute(delete(ExamBankMap).where(ExamBankMap.exam_id==exam.id))
    if question_ids:
        s.execute(delete(Question).where(Question.id.in_(question_ids)))

    s.execute(delete(PracticalCodeSubmission).where(PracticalCodeSubmission.exam_id==exam.id))
    s.execute(delete(PracticeAttempt).where(PracticeAttempt.exam_id==exam.id))
    s.execute(delete(ExamSession).where(ExamSession.exam_id==exam.id))
    s.execute(delete(ExamApproval).where(ExamApproval.exam_id==exam.id))
    s.execute(delete(ExamPracticeRelease).where(ExamPracticeRelease.exam_id==exam.id))
    s.execute(delete(ExamSecurityPolicy).where(ExamSecurityPolicy.exam_id==exam.id))
    s.execute(delete(ExamStudentAccess).where(ExamStudentAccess.exam_id==exam.id))
    s.execute(delete(ExamDeviceLock).where(ExamDeviceLock.exam_id==exam.id))
    s.execute(delete(ExamCandidateCheckin).where(ExamCandidateCheckin.exam_id==exam.id))
    s.execute(delete(EdgePackageReceipt).where(EdgePackageReceipt.exam_id==exam.id))
    s.execute(delete(ExamConfig).where(ExamConfig.exam_id==exam.id))

    audit_event(
        s,'exam_deleted','exam',exam.id,
        f'title={title}, questions={len(question_ids)}, attempts={len(attempt_ids)}, answers={answer_count}'
    )
    s.delete(exam)
    s.commit()
    flash(f'Exam “{title}” and its related results/marks were deleted.')
    return redirect(url_for('exams'))


@app.route('/admin/exam/<int:exam_id>/toggle',methods=['POST'])
@staff_required
def toggle_exam(exam_id):
    s=DB();e=s.get(Exam,exam_id)
    if e:
        if not e.is_active and (s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==exam_id)) or 0)==0:flash('Add questions before activating this exam.','error');return redirect(url_for('exams'))
        if not e.is_active:
            approval=get_exam_approval(s,exam_id,create=True)
            if approval.status!='approved':
                if can_approve_exams(s):
                    approval.status='approved';approval.reviewed_by=actor_label(s);approval.reviewed_at=now_iso();approval.comments='Approved during activation';audit_event(s,'exam_approved','exam',e.id,'approved during activation')
                elif current_staff_role(s)=='faculty':
                    policy=exam_approval_policy(s,e)
                    if policy['self_approval_allowed']:
                        approval.status='approved';approval.reviewed_by=actor_label(s);approval.reviewed_at=now_iso();approval.comments=f'Faculty self-approved under <= {FACULTY_DAILY_SELF_APPROVAL_LIMIT} exams/day policy';audit_event(s,'exam_self_approved','exam',e.id,policy['message'])
                    else:
                        flash(policy['message'] or 'This exam requires HOD / Exam Controller approval before activation. Use Request Approval first.','error');return redirect(request.referrer or url_for('exams'))
                else:
                    flash('This exam requires HOD / Exam Controller approval before activation. Use Request Approval first.','error');return redirect(request.referrer or url_for('exams'))
        e.is_active=not bool(e.is_active)
        audit_event(s,'exam_activated' if e.is_active else 'exam_deactivated','exam',e.id,e.title);s.commit()
    return redirect(url_for('exams'))


@app.route('/admin/exam/<int:exam_id>/approval/request',methods=['POST'])
@staff_required
def request_exam_approval(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    row=get_exam_approval(s,exam_id,create=True);row.status='pending';row.requested_by=actor_label(s);row.requested_at=now_iso();row.comments=request.form.get('comments','').strip()[:500];row.reviewed_by='';row.reviewed_at='';audit_event(s,'exam_approval_requested','exam',exam_id,row.comments);s.commit();flash('Exam sent for HOD / Exam Controller approval.');return redirect(request.referrer or url_for('exams'))

@app.route('/admin/exam/<int:exam_id>/approval/<decision>',methods=['POST'])
@approver_required
def review_exam_approval(exam_id,decision):
    if decision not in {'approve','reject'}:abort(404)
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    row=get_exam_approval(s,exam_id,create=True);row.status='approved' if decision=='approve' else 'rejected';row.reviewed_by=actor_label(s);row.reviewed_at=now_iso();row.comments=request.form.get('comments','').strip()[:500];audit_event(s,'exam_approved' if decision=='approve' else 'exam_rejected','exam',exam_id,row.comments);s.commit();flash('Exam approved.' if decision=='approve' else 'Exam returned for changes.');return redirect(request.referrer or url_for('exams'))

@app.route('/admin/exam/<int:exam_id>/session',methods=['POST'])
@staff_required
def save_exam_session(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    try:group_id=int(request.form.get('group_id','0'))
    except ValueError:group_id=0
    group=s.get(AcademicGroup,group_id)
    if not group:flash('Choose a valid batch / section.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
    try:start=parse_local_schedule(request.form.get('scheduled_start',''));end=parse_local_schedule(request.form.get('scheduled_end',''))
    except ValueError as exc:flash(str(exc),'error');return redirect(url_for('exam_builder',exam_id=exam_id))
    if start and end and datetime.fromisoformat(end)<=datetime.fromisoformat(start):flash('Session end time must be after the start time.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
    row=s.scalar(select(ExamSession).where(ExamSession.exam_id==exam_id,ExamSession.group_id==group_id))
    if row:row.scheduled_start=start;row.scheduled_end=end;row.venue=request.form.get('venue','').strip()
    else:s.add(ExamSession(exam_id=exam_id,group_id=group_id,scheduled_start=start,scheduled_end=end,venue=request.form.get('venue','').strip(),created_at=now_iso()))
    audit_event(s,'exam_session_saved','exam',exam_id,f'{group_label(group)}, {start or "any time"}, venue={request.form.get("venue","").strip()}');s.commit();flash('Exam batch / section session saved.');return redirect(url_for('exam_builder',exam_id=exam_id))

@app.route('/admin/exam/<int:exam_id>/session/<int:session_id>/delete',methods=['POST'])
@staff_required
def delete_exam_session(exam_id,session_id):
    s=DB();row=s.get(ExamSession,session_id)
    if row and row.exam_id==exam_id:s.delete(row);audit_event(s,'exam_session_deleted','exam',exam_id,f'session={session_id}');s.commit();flash('Exam session removed.')
    return redirect(url_for('exam_builder',exam_id=exam_id))

@app.route('/admin/exam/<int:exam_id>/edge-package')
@staff_required
def export_edge_exam_package(exam_id):
    if len(EXAM_PACKAGE_SIGNING_KEY.encode('utf-8'))<32:abort(503,'Set EXAM_PACKAGE_SIGNING_KEY (32+ characters) before using Edge packages.')
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    payload=edge_exam_payload(s,exam)
    if not payload['questions']:flash('Add questions before exporting an Edge package.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
    envelope=seal_envelope(payload,EXAM_PACKAGE_SIGNING_KEY);raw=json.dumps(envelope,ensure_ascii=False,separators=(',',':')).encode('utf-8');audit_event(s,'edge_exam_package_exported','exam',exam.id,f'package={envelope["package_id"]}');s.commit();data=io.BytesIO(raw)
    return send_file(data,mimetype='application/octet-stream',as_attachment=True,download_name=f'{_edge_filename(exam.title)}_{envelope["package_id"]}.lwhexam')


@app.route('/admin/exam/<int:exam_id>/edge-results-package')
@staff_required
def export_edge_results_package(exam_id):
    if len(EXAM_PACKAGE_SIGNING_KEY.encode('utf-8'))<32:abort(503,'Set EXAM_PACKAGE_SIGNING_KEY (32+ characters) before using Edge packages.')
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    payload=edge_results_payload(s,exam);envelope=seal_envelope(payload,EXAM_PACKAGE_SIGNING_KEY);raw=json.dumps(envelope,ensure_ascii=False,separators=(',',':')).encode('utf-8');audit_event(s,'edge_results_package_exported','exam',exam.id,f'package={envelope["package_id"]}');s.commit();data=io.BytesIO(raw)
    return send_file(data,mimetype='application/octet-stream',as_attachment=True,download_name=f'{_edge_filename(exam.title)}_results_{envelope["package_id"]}.lwhresults')


@app.route('/admin/edge-package/import',methods=['POST'])
@approver_required
def import_edge_exam_package():
    if len(EXAM_PACKAGE_SIGNING_KEY.encode('utf-8'))<32:abort(503,'Set EXAM_PACKAGE_SIGNING_KEY (32+ characters) before importing Edge packages.')
    upload=request.files.get('edge_package')
    if not upload or not upload.filename:flash('Choose an encrypted .lwhexam Edge package.','error');return redirect(url_for('exam_centre'))
    s=DB()
    try:
        raw=upload.read()
        if not raw or len(raw)>10*1024*1024:raise ValueError('The Edge package is empty or exceeds the 10 MB import limit.')
        envelope=json.loads(raw.decode('utf-8'));payload=open_sealed_envelope(envelope,EXAM_PACKAGE_SIGNING_KEY);pid=str(envelope.get('package_id') or '')
        if payload.get('kind')!='exam' or int(payload.get('schema_version') or 0)!=1:raise ValueError('This file is not a supported exam Edge package.')
        existing=s.scalar(select(EdgePackageReceipt).where(EdgePackageReceipt.package_id==pid))
        if existing:
            flash(f'This Edge package was already imported as Exam #{existing.exam_id}.');return redirect(url_for('exam_builder',exam_id=existing.exam_id))
        meta=payload.get('exam') or {};title=str(meta.get('title') or '').strip();duration=int(meta.get('duration_minutes') or 0);qrows=payload.get('questions') or []
        if not title or duration<=0 or not isinstance(qrows,list) or not qrows:raise ValueError('The Edge package is missing required exam details or questions.')
        exam=Exam(title=title,duration_minutes=max(1,duration),is_active=False,created_at=now_iso());s.add(exam);s.flush()
        cfg_data=payload.get('config') or {};cfg=get_exam_config(s,exam.id,create=True);cfg.subject=str(cfg_data.get('subject') or '')[:200];cfg.course_semester=str(cfg_data.get('course_semester') or '')[:200];cfg.question_count=max(1,min(len(qrows),int(cfg_data.get('question_count') or len(qrows))));cfg.pool_size=max(cfg.question_count,min(len(qrows),int(cfg_data.get('pool_size') or len(qrows))))
        easy=int(cfg_data.get('easy_pct') or 30);medium=int(cfg_data.get('medium_pct') or 50);hard=int(cfg_data.get('hard_pct') or 20)
        if easy+medium+hard!=100:easy,medium,hard=30,50,20
        cfg.easy_pct=max(0,easy);cfg.medium_pct=max(0,medium);cfg.hard_pct=max(0,hard);cfg.unit_weights=str(cfg_data.get('unit_weights') or '')[:2000];cfg.randomize_questions=bool(cfg_data.get('randomize_questions',True));cfg.shuffle_options=bool(cfg_data.get('shuffle_options',True));cfg.require_fullscreen=bool(cfg_data.get('require_fullscreen',False));cfg.tab_switch_limit=max(0,min(100,int(cfg_data.get('tab_switch_limit') or 3)));cfg.last_generation_summary=f'Imported from encrypted Edge package {pid}.';cfg.updated_at=now_iso()
        security_data=payload.get('security') or {};security=get_exam_security_policy(s,exam.id,create=True);security.require_candidate_checkin=bool(security_data.get('require_candidate_checkin',False));security.require_exam_pin=bool(security_data.get('require_exam_pin',False));security.heartbeat_seconds=max(10,min(60,int(security_data.get('heartbeat_seconds') or 15)));security.updated_at=now_iso()
        for idx,item in enumerate(qrows,1):
            if not isinstance(item,dict):raise ValueError(f'Question {idx} is malformed.')
            qtype=canonical_question_type(item.get('question_type'));question_text=str(item.get('question') or '').strip();options={key:str(item.get(f'option_{key.lower()}') or '').strip() for key in 'ABCD'};answer_key=str(item.get('answer_key') or item.get('correct_answer') or '').strip() if qtype!='essay' else '';tolerance=str(item.get('answer_tolerance') or '').strip();error=validate_question_definition(qtype,question_text,options,answer_key,tolerance)
            if error:raise ValueError(f'Question {idx}: {error}')
            try:marks=max(1,int(item.get('marks') or 1))
            except Exception:marks=1
            legacy=str(item.get('correct_answer') or '').upper();legacy=legacy if legacy in {'A','B','C','D'} else (answer_key[:1].upper() if qtype=='single_choice' and answer_key[:1].upper() in {'A','B','C','D'} else 'A')
            s.add(Question(exam_id=exam.id,question=question_text,option_a=options['A'],option_b=options['B'],option_c=options['C'],option_d=options['D'],correct_answer=legacy,question_type=qtype,answer_key=answer_key,answer_tolerance=tolerance,answer_case_sensitive=bool(item.get('answer_case_sensitive',False)),marks=marks,practical_experiment_no=normalize_practical_exam_no(item.get('practical_experiment_no'))))
        source=payload.get('source') or {};s.add(EdgePackageReceipt(package_id=pid,exam_id=exam.id,source_mode=str(source.get('mode') or '')[:20],source_exam_id=str(source.get('exam_id') or '')[:80],imported_by=actor_label(s),imported_at=now_iso()));audit_event(s,'edge_exam_package_imported','exam',exam.id,f'package={pid}, source={source.get("mode","")}:{source.get("exam_id","")}');s.commit();flash(f'Encrypted Edge package verified and imported as Draft Exam #{exam.id}. Assign the local batch/session, then approve and activate it.');return redirect(url_for('exam_builder',exam_id=exam.id))
    except (ValueError,TypeError,json.JSONDecodeError,UnicodeDecodeError) as exc:
        s.rollback();flash(f'Edge package rejected: {exc}','error');return redirect(url_for('exam_centre'))
    except Exception:
        s.rollback();flash('Edge package import failed safely. The database was not changed.','error');return redirect(url_for('exam_centre'))


@app.route('/admin/edge-results/import',methods=['POST'])
@approver_required
def import_edge_results_package():
    if len(EXAM_PACKAGE_SIGNING_KEY.encode('utf-8'))<32:abort(503,'Set EXAM_PACKAGE_SIGNING_KEY (32+ characters) before importing Edge result packages.')
    upload=request.files.get('edge_results_package')
    if not upload or not upload.filename:flash('Choose an encrypted .lwhresults package.','error');return redirect(url_for('exam_centre'))
    s=DB()
    try:
        raw=upload.read()
        if not raw or len(raw)>10*1024*1024:raise ValueError('The results package is empty or exceeds the 10 MB import limit.')
        envelope=json.loads(raw.decode('utf-8'));payload=open_sealed_envelope(envelope,EXAM_PACKAGE_SIGNING_KEY);pid=str(envelope.get('package_id') or '')
        if payload.get('kind')!='results' or int(payload.get('schema_version') or 0)!=1:raise ValueError('This file is not a supported Edge results package.')
        existing=s.scalar(select(EdgeResultReceipt).where(EdgeResultReceipt.package_id==pid))
        if existing:flash('This results package was already reconciled.');return redirect(url_for('edge_results_reconciliation',receipt_id=existing.id))
        attempts=payload.get('attempts') or []
        if not isinstance(attempts,list) or len(attempts)>50000:raise ValueError('The results package contains an invalid attempt collection.')
        source=payload.get('source') or {};origin=payload.get('origin') or {};exam_meta=payload.get('exam') or {}
        origin_exam_id=str(origin.get('exam_id') or '')[:80];target_exam=None
        if origin_exam_id.isdigit():target_exam=s.get(Exam,int(origin_exam_id))
        receipt=EdgeResultReceipt(package_id=pid,source_mode=str(source.get('mode') or '')[:20],source_exam_id=str(source.get('exam_id') or '')[:80],origin_exam_id=origin_exam_id,target_exam_id=target_exam.id if target_exam else None,exam_title=str(exam_meta.get('title') or '')[:250],attempts_count=len(attempts),submitted_count=sum(1 for a in attempts if isinstance(a,dict) and a.get('status')=='submitted'),imported_by=actor_label(s),imported_at=now_iso());s.add(receipt);s.flush()
        seen=set()
        for idx,item in enumerate(attempts,1):
            if not isinstance(item,dict):raise ValueError(f'Attempt {idx} is malformed.')
            roll=str(item.get('roll_no') or '').strip()
            if not roll:raise ValueError(f'Attempt {idx} has no roll number.')
            if roll in seen:raise ValueError(f'Duplicate roll number in package: {roll}.')
            seen.add(roll)
            score=item.get('score');total=item.get('total_marks')
            try:score=None if score in {None,''} else int(score)
            except Exception:raise ValueError(f'Attempt {idx} has an invalid score.')
            try:total=None if total in {None,''} else int(total)
            except Exception:raise ValueError(f'Attempt {idx} has invalid total marks.')
            integrity=item.get('integrity') or []
            if not isinstance(integrity,list):integrity=[]
            s.add(EdgeResultAttempt(receipt_id=receipt.id,roll_no=roll[:100],name=str(item.get('name') or '')[:250],status=str(item.get('status') or '')[:30],grading_status=str(item.get('grading_status') or '')[:30],score=score,total_marks=total,integrity_count=len(integrity),payload_json=json.dumps(item,ensure_ascii=False,separators=(',',':'))))
        audit_event(s,'edge_results_package_imported','edge_result_receipt',receipt.id,f'package={pid}, attempts={len(attempts)}, target_exam={target_exam.id if target_exam else "unmapped"}');s.commit();flash(f'Encrypted Edge results verified. {len(attempts)} candidate record(s) added to reconciliation without overwriting primary exam data.');return redirect(url_for('edge_results_reconciliation',receipt_id=receipt.id))
    except (ValueError,TypeError,json.JSONDecodeError,UnicodeDecodeError) as exc:
        s.rollback();flash(f'Edge results package rejected: {exc}','error');return redirect(url_for('exam_centre'))
    except Exception:
        s.rollback();flash('Edge results import failed safely. Primary exam data was not changed.','error');return redirect(url_for('exam_centre'))


@app.route('/admin/edge-results')
@staff_required
def edge_results_reconciliation():
    s=DB();receipts=s.scalars(select(EdgeResultReceipt).order_by(EdgeResultReceipt.id.desc()).limit(100)).all();selected=None;attempts=[]
    rid=request.args.get('receipt_id','').strip()
    if rid.isdigit():selected=s.get(EdgeResultReceipt,int(rid))
    elif receipts:selected=receipts[0]
    if selected:attempts=s.scalars(select(EdgeResultAttempt).where(EdgeResultAttempt.receipt_id==selected.id).order_by(EdgeResultAttempt.roll_no)).all()
    return render_template('edge_results.html',receipts=receipts,selected=selected,attempts=attempts)


@app.route('/admin/edge-results/<int:receipt_id>/csv')
@staff_required
def edge_results_reconciliation_csv(receipt_id):
    s=DB();receipt=s.get(EdgeResultReceipt,receipt_id)
    if not receipt:abort(404)
    rows=s.scalars(select(EdgeResultAttempt).where(EdgeResultAttempt.receipt_id==receipt.id).order_by(EdgeResultAttempt.roll_no)).all();out=io.StringIO(newline='');writer=csv.writer(out);writer.writerow(['roll_no','name','status','grading_status','score','total_marks','integrity_events']);
    for row in rows:writer.writerow([row.roll_no,row.name,row.status,row.grading_status,row.score if row.score is not None else '',row.total_marks if row.total_marks is not None else '',row.integrity_count])
    data=io.BytesIO(out.getvalue().encode('utf-8-sig'));return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'edge_reconciliation_{receipt.package_id}.csv')


@app.route('/admin/exam/<int:exam_id>/builder',methods=['GET','POST'])
@staff_required
def exam_builder(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    cfg=get_exam_config(s,exam_id,create=True);security=get_exam_security_policy(s,exam_id,create=True)
    if request.method=='POST':
        action=request.form.get('action','save')
        try:
            qcount=max(1,int(request.form.get('question_count','20')));pool_size=max(qcount,int(request.form.get('pool_size',str(qcount))))
            easy=max(0,int(request.form.get('easy_pct','30')));medium=max(0,int(request.form.get('medium_pct','50')));hard=max(0,int(request.form.get('hard_pct','20')));tab_limit=max(0,int(request.form.get('tab_switch_limit','3')));heartbeat_seconds=max(10,min(60,int(request.form.get('heartbeat_seconds','15') or 15)))
        except ValueError:flash('Blueprint numeric values are invalid.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
        if easy+medium+hard!=100:flash('Difficulty distribution must total 100%.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
        try:unit_weights=parse_unit_weights(request.form.get('unit_weights',''))
        except ValueError as exc:flash(str(exc),'error');return redirect(url_for('exam_builder',exam_id=exam_id))
        cfg.subject=request.form.get('subject','').strip();cfg.course_semester=request.form.get('course_semester','').strip();cfg.question_count=qcount;cfg.pool_size=pool_size;cfg.easy_pct=easy;cfg.medium_pct=medium;cfg.hard_pct=hard;cfg.unit_weights=json.dumps(unit_weights,ensure_ascii=False);cfg.randomize_questions=request.form.get('randomize_questions')=='on';cfg.shuffle_options=request.form.get('shuffle_options')=='on';cfg.require_fullscreen=request.form.get('require_fullscreen')=='on';cfg.tab_switch_limit=tab_limit;cfg.updated_at=now_iso();security.require_candidate_checkin=request.form.get('require_candidate_checkin')=='on';security.require_exam_pin=request.form.get('require_exam_pin')=='on';security.heartbeat_seconds=heartbeat_seconds;security.updated_at=now_iso();s.flush()
        if action=='generate':
            if (s.scalar(select(func.count()).select_from(Attempt).where(Attempt.exam_id==exam_id)) or 0)>0:flash('This exam already has attempts. The question pool is locked to protect result integrity.','error');s.rollback();return redirect(url_for('exam_builder',exam_id=exam_id))
            stmt=select(BankQuestion).where(BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['official_only','both']))
            if cfg.subject:stmt=stmt.where(BankQuestion.subject==cfg.subject)
            if cfg.course_semester:stmt=stmt.where(BankQuestion.course_semester==cfg.course_semester)
            candidates=s.scalars(stmt).all()
            if unit_weights:candidates=[q for q in candidates if (q.unit or '').strip() in unit_weights]
            selected,ut,dt,uc,dc=choose_blueprint_questions(candidates,pool_size,unit_weights,{'Easy':easy,'Medium':medium,'Hard':hard})
            maps=s.scalars(select(ExamBankMap).where(ExamBankMap.exam_id==exam_id)).all(); mapped_qids=[m.exam_question_id for m in maps]
            for m in maps:s.delete(m)
            s.flush()
            for qid in mapped_qids:
                q=s.get(Question,qid)
                if q:s.delete(q)
            s.flush();added=0
            for bq in selected:
                if copy_bank_question_to_exam(s,bq,exam_id):added+=1
            summary=f'Generated pool: {added}/{pool_size}. Difficulty actual: Easy {dc.get("Easy",0)}, Medium {dc.get("Medium",0)}, Hard {dc.get("Hard",0)}.'
            if unit_weights:summary+=' Units: '+', '.join(f'{k}={uc.get(k,0)}' for k in unit_weights)+'.'
            if added<pool_size:summary+=' Bank does not yet contain enough approved questions for the requested blueprint.'
            cfg.last_generation_summary=summary;audit_event(s,'exam_pool_generated','exam',exam_id,summary);s.commit();flash(summary);return redirect(url_for('exam_builder',exam_id=exam_id))
        audit_event(s,'exam_blueprint_saved','exam',exam_id,f'questions={qcount}, pool={pool_size}, rotating_pin={security.require_exam_pin}');s.commit();flash('Exam blueprint and integrity settings saved.');return redirect(url_for('exam_builder',exam_id=exam_id))
    subjects=s.scalars(select(BankQuestion.subject).where(BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['official_only','both'])).distinct().order_by(BankQuestion.subject)).all();pool_count=s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==exam_id)) or 0;attempt_count=s.scalar(select(func.count()).select_from(Attempt).where(Attempt.exam_id==exam_id)) or 0
    try:unit_weights_display=', '.join(f'{k}:{v}' for k,v in json.loads(cfg.unit_weights or '{}').items())
    except Exception:unit_weights_display=''
    groups=s.scalars(select(AcademicGroup).where(AcademicGroup.is_active==True).order_by(AcademicGroup.program,AcademicGroup.semester,AcademicGroup.section)).all();sessions=s.execute(select(ExamSession,AcademicGroup).join(AcademicGroup,AcademicGroup.id==ExamSession.group_id).where(ExamSession.exam_id==exam_id).order_by(ExamSession.scheduled_start)).all();approval=get_exam_approval(s,exam_id,create=True);approval_policy=exam_approval_policy(s,exam);practice_release=get_exam_practice_release(s,exam_id,create=True);s.commit();return render_template('exam_builder.html',exam=exam,cfg=cfg,security=security,subjects=subjects,pool_count=pool_count,attempt_count=attempt_count,unit_weights_display=unit_weights_display,groups=groups,sessions=sessions,approval=approval,approval_policy=approval_policy,practice_release=practice_release,group_label=group_label,can_approve=can_approve_exams(s))

@app.route('/admin/exam/<int:exam_id>/student-access')
@staff_required
def exam_student_access_admin(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    security=get_exam_security_policy(s,exam_id,create=True);students=eligible_students_for_exam(s,exam_id)
    student_ids=[st.id for st in students]
    lock_rows=s.scalars(select(ExamDeviceLock).where(ExamDeviceLock.exam_id==exam_id,ExamDeviceLock.student_id.in_(student_ids))).all() if student_ids else []
    lock_map={row.student_id:row for row in lock_rows};rows=[]
    for st in students:
        lock=lock_map.get(st.id)
        rows.append(type('StudentAccessView',(),{'student':st,'locked':bool(lock),'last_seen_at':lock.last_seen_at if lock else ''})())
    s.commit();return render_template('exam_student_access.html',exam=exam,security=security,rows=rows)


@app.route('/admin/exam/<int:exam_id>/student-access/security',methods=['POST'])
@staff_required
def update_exam_student_access_security(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    security=get_exam_security_policy(s,exam_id,create=True);security.require_exam_pin=request.form.get('require_exam_pin')=='on';security.updated_at=now_iso()
    audit_event(s,'exam_rotating_pin_security_updated','exam',exam_id,f'enabled={security.require_exam_pin}');s.commit();flash('Rotating Exam PIN security enabled.' if security.require_exam_pin else 'Rotating Exam PIN security disabled.')
    return redirect(url_for('exam_student_access_admin',exam_id=exam_id))


@app.route('/admin/exam/<int:exam_id>/student-access/generate',methods=['POST'])
@staff_required
def generate_exam_student_access(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    flash('The Exam PIN is generated automatically and changes every 60 seconds.')
    return redirect(url_for('exam_student_access_admin',exam_id=exam_id))


@app.route('/admin/exam/<int:exam_id>/student-access.csv')
@staff_required
def exam_student_access_csv(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    students=eligible_students_for_exam(s,exam_id);out=io.StringIO(newline='');writer=csv.writer(out);writer.writerow(['roll_no','name','exam_title','exam_pin'])
    for st in students:
        access=exam_pin_record(s,exam_id,st.id);writer.writerow([st.roll_no,st.name,exam.title,decrypt_exam_pin(access.pin_ciphertext) if access else ''])
    data=io.BytesIO(out.getvalue().encode('utf-8-sig'));audit_event(s,'exam_pin_list_exported','exam',exam_id,f'students={len(students)}');s.commit()
    return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'exam_{exam_id}_student_pins.csv')


@app.route('/admin/exam/<int:exam_id>/student-access/<int:student_id>/reset-device',methods=['POST'])
@staff_required
def reset_exam_student_device(exam_id,student_id):
    s=DB();exam=s.get(Exam,exam_id);student=s.get(Student,student_id)
    if not exam or not student:abort(404)
    row=s.scalar(select(ExamDeviceLock).where(ExamDeviceLock.exam_id==exam_id,ExamDeviceLock.student_id==student_id))
    if row:s.delete(row)
    audit_event(s,'exam_device_lock_reset','exam',exam_id,f'student={student.roll_no}, by={actor_label(s)}');s.commit();flash(f'Device lock reset for {student.roll_no}.')
    return redirect(url_for('exam_student_access_admin',exam_id=exam_id))


@app.route('/admin/exam/<int:exam_id>/practice-release',methods=['POST'])
@staff_required
def update_exam_practice_release(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    role=current_staff_role(s);owner=exam_owner_actor(s,exam_id)
    if role=='faculty' and owner!=actor_label(s):abort(403)
    row=get_exam_practice_release(s,exam_id,create=True)
    wants_release=request.form.get('is_released')=='on';release_after=(request.form.get('release_after') or '').strip()
    if wants_release and exam.is_active and not release_after:
        flash('An active official exam cannot be released immediately as a practice paper. Set a future release time or deactivate the exam first.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
    if release_after:
        try:parse_dt(release_after)
        except Exception:flash('The practice-paper release date/time is invalid.','error');return redirect(url_for('exam_builder',exam_id=exam_id))
    row.is_released=wants_release;row.release_after=release_after;row.show_solutions=request.form.get('show_solutions')=='on';row.allow_mock=request.form.get('allow_mock')=='on';row.updated_by=actor_label(s);row.updated_at=now_iso()
    audit_event(s,'exam_practice_release_updated','exam',exam.id,f'released={row.is_released}, release_after={row.release_after}, solutions={row.show_solutions}, mock={row.allow_mock}');s.commit();flash('Previous-paper practice settings updated.');return redirect(url_for('exam_builder',exam_id=exam_id))

@app.route('/admin/exam/<int:exam_id>/questions',methods=['GET','POST'])
@staff_required
def questions(exam_id):
    s=DB();exam=s.get(Exam,exam_id)
    if not exam:abort(404)
    if request.method=='POST':
        qdef=question_definition_from_form(request.form)
        if qdef['error']:flash(qdef['error'],'error');return redirect(url_for('questions',exam_id=exam_id))
        try:marks=max(1,int(request.form.get('marks','1')))
        except ValueError:marks=1
        opts=qdef['options'];q=Question(exam_id=exam_id,question=request.form.get('question','').strip(),option_a=opts['A'],option_b=opts['B'],option_c=opts['C'],option_d=opts['D'],correct_answer=qdef['legacy_correct_answer'],question_type=qdef['question_type'],answer_key=qdef['answer_key'],answer_tolerance=qdef['answer_tolerance'],answer_case_sensitive=qdef['answer_case_sensitive'],marks=marks);s.add(q);s.flush();pool_count=sync_manual_exam_question_count(s,exam_id);audit_event(s,'exam_question_added','exam',exam_id,f'question_id={q.id}, type={qdef["question_type"]}, pool={pool_count}');s.commit();flash(f'Question added. New attempts will receive all {pool_count} question(s) in this exam.')
    qs=s.scalars(select(Question).where(Question.exam_id==exam_id).order_by(Question.id)).all()
    cfg=normalize_legacy_manual_subject_exam(s,exam_id,get_exam_config(s,exam_id,create=False))
    mapped=set(s.scalars(select(ExamBankMap.exam_question_id).where(ExamBankMap.exam_id==exam_id)).all())
    return render_template('questions.html',exam=exam,questions=qs,mapped=mapped)

@app.route('/admin/exam/<int:exam_id>/import',methods=['POST'])
@staff_required
def import_questions(exam_id):
    f=request.files.get('csv_file')
    if not f:flash('Choose a CSV file.','error');return redirect(url_for('questions',exam_id=exam_id))
    try:text=f.stream.read().decode('utf-8-sig')
    except UnicodeDecodeError:flash('CSV must be UTF-8 encoded.','error');return redirect(url_for('questions',exam_id=exam_id))
    reader=csv.DictReader(io.StringIO(text));required={'question','option_a','option_b','option_c','option_d','correct_answer','marks'}
    if not required.issubset(set(reader.fieldnames or [])):flash('CSV columns are incorrect. Use sample_questions.csv.','error');return redirect(url_for('questions',exam_id=exam_id))
    s=DB();count=0
    for r in reader:
        ans=(r.get('correct_answer') or '').strip().upper()
        if ans not in {'A','B','C','D'}:continue
        try:marks=max(1,int(r.get('marks') or 1))
        except ValueError:marks=1
        if not (r.get('question') or '').strip():continue
        s.add(Question(exam_id=exam_id,question=r['question'].strip(),option_a=(r.get('option_a') or '').strip(),option_b=(r.get('option_b') or '').strip(),option_c=(r.get('option_c') or '').strip(),option_d=(r.get('option_d') or '').strip(),correct_answer=ans,question_type='single_choice',answer_key=ans,answer_tolerance='',answer_case_sensitive=False,marks=marks));count+=1
    pool_count=sync_manual_exam_question_count(s,exam_id)
    audit_event(s,'exam_questions_csv_import','exam',exam_id,f'count={count}, pool={pool_count}');s.commit();flash(f'Imported {count} questions. New attempts will receive all {pool_count} question(s) in this exam.');return redirect(url_for('questions',exam_id=exam_id))

def result_rows(s,exam_id=None):
    stmt=select(Attempt,Student,Exam).join(Student,Student.id==Attempt.student_id).join(Exam,Exam.id==Attempt.exam_id)
    if exam_id:stmt=stmt.where(Exam.id==exam_id)
    raw=s.execute(stmt.order_by(Attempt.id.desc())).all();rows=[]

    # Load integrity activity once so the Results page can show a useful breakdown
    # without exposing those details on the student's result screen.
    integrity_by_attempt={}
    for ev in s.scalars(select(IntegrityEvent).order_by(IntegrityEvent.created_at,IntegrityEvent.id)).all():
        integrity_by_attempt.setdefault(ev.attempt_id,[]).append(ev)

    for a,st,e in raw:
        events=integrity_by_attempt.get(a.id,[])
        tab_switches=sum(1 for ev in events if ev.event_type=='tab_hidden')
        fullscreen_exits=sum(1 for ev in events if ev.event_type=='fullscreen_exit')
        pct,grade,_grade_class=result_performance(a.score,a.total_marks) if a.status=='submitted' else (0,'-','')
        grading_status=getattr(a,'grading_status','complete') or 'complete'
        if a.status=='submitted' and grading_status=='pending':grade='Pending grading'
        grp=student_group(s,st.id)
        rows.append(type('ResultRow',(),{
            'attempt_id':a.id,
            'roll_no':st.roll_no,
            'name':st.name,
            'group_label':group_label(grp) if grp else 'Unassigned',
            'title':e.title,
            'exam_id':e.id,
            'status':a.status,
            'grading_status':grading_status,
            'score':a.score,
            'total_marks':a.total_marks,
            'percentage':pct,
            'grade':grade,
            'started_at':a.started_at,
            'submitted_at':a.submitted_at,
            'violations':len(events),
            'tab_switches':tab_switches,
            'fullscreen_exits':fullscreen_exits,
            'integrity_events':events,
        })())
    return rows

@app.route('/admin/results')
@staff_required
def results():
    s=DB();exam_id=request.args.get('exam_id',type=int);rows=result_rows(s,exam_id);exams_list=s.scalars(select(Exam).order_by(Exam.title)).all();return render_template('results.html',rows=rows,exams=exams_list,selected_exam_id=exam_id)

@app.route('/admin/attempt/<int:attempt_id>/grade',methods=['GET','POST'])
@staff_required
def manual_grade_attempt(attempt_id):
    s=DB();attempt=s.get(Attempt,attempt_id)
    if not attempt or attempt.status!='submitted':abort(404)
    student=s.get(Student,attempt.student_id);exam=s.get(Exam,attempt.exam_id);qids=attempt_question_ids(s,attempt)
    questions=s.scalars(select(Question).where(Question.id.in_(qids))).all() if qids else [];qmap={q.id:q for q in questions};essay_questions=[qmap[qid] for qid in qids if qid in qmap and canonical_question_type(qmap[qid].question_type)=='essay']
    answers=s.scalars(select(Answer).where(Answer.attempt_id==attempt.id)).all();amap={a.question_id:a for a in answers}
    if not essay_questions:flash('This attempt has no descriptive questions.');return redirect(url_for('results',exam_id=attempt.exam_id))
    if request.method=='POST':
        errors=[]
        for q in essay_questions:
            ans=amap.get(q.id);value=answer_record_value(ans) if ans else ''
            if not ans or not value:continue
            raw=(request.form.get(f'score_{q.id}') or '').strip();comment=(request.form.get(f'comment_{q.id}') or '').strip()[:1500]
            if raw=='':ans.manual_score=None;ans.grader_comment=comment;ans.graded_by='';ans.graded_at='';continue
            try:score=int(raw)
            except ValueError:errors.append(f'Q{q.id}: marks must be a whole number.');continue
            if score<0 or score>q.marks:errors.append(f'Q{q.id}: marks must be between 0 and {q.marks}.');continue
            ans.manual_score=score;ans.grader_comment=comment;ans.graded_by=actor_label(s);ans.graded_at=now_iso()
        if errors:
            s.rollback();flash(' '.join(errors),'error');return redirect(url_for('manual_grade_attempt',attempt_id=attempt.id))
        recalculate_attempt_score(s,attempt,questions,answers);audit_event(s,'attempt_manual_graded','attempt',attempt.id,f'status={attempt.grading_status}, score={attempt.score}/{attempt.total_marks}');s.commit()
        try:
            sync_result=sync_practical_viva_from_attempt(s,attempt)
            if sync_result.get('updated'):s.commit()
        except Exception:
            s.rollback()
            try:app.logger.exception('Practical Exam viva re-sync failed after manual grading for attempt %s',attempt.id)
            except Exception:pass
        flash('Manual grading saved.');return redirect(url_for('manual_grade_attempt',attempt_id=attempt.id))
    return render_template('manual_grade.html',attempt=attempt,student=student,exam=exam,questions=essay_questions,answers=amap)


@app.route('/admin/results/export/<fmt>')
@staff_required
def export_results(fmt):
    s=DB();exam_id=request.args.get('exam_id',type=int);rows=result_rows(s,exam_id);headers=['roll_no','name','batch_section','exam','status','grading_status','score','total_marks','percentage','grade','tab_switches','fullscreen_exits','total_integrity_events','started','submitted']
    matrix=[[r.roll_no,r.name,r.group_label,r.title,r.status,r.grading_status,r.score if r.score is not None else '',r.total_marks if r.total_marks is not None else '',r.percentage if r.status=='submitted' else '',r.grade if r.status=='submitted' else '',r.tab_switches,r.fullscreen_exits,r.violations,r.started_at,r.submitted_at or ''] for r in rows]
    suffix=f'_exam_{exam_id}' if exam_id else '_all'
    if fmt=='csv':
        out=io.StringIO(newline='');w=csv.writer(out);w.writerow(headers);w.writerows(matrix);data=io.BytesIO(out.getvalue().encode('utf-8-sig'));return send_file(data,mimetype='text/csv',as_attachment=True,download_name=f'exam_results{suffix}.csv')
    if fmt=='xlsx':
        wb=Workbook();ws=wb.active;ws.title='Results';ws.append(headers)
        for row in matrix:ws.append(row)
        for cell in ws[1]:cell.font=Font(bold=True)
        widths=[16,28,34,30,14,18,10,12,12,16,14,16,20,24,24]
        for idx,width in enumerate(widths,1):ws.column_dimensions[chr(64+idx)].width=width
        data=io.BytesIO();wb.save(data);data.seek(0);return send_file(data,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',as_attachment=True,download_name=f'exam_results{suffix}.xlsx')
    abort(404)


def _score_summary(percentages):
    if not percentages:return {'count':0,'average':0,'highest':0,'lowest':0,'pass_pct':0,'outstanding':0,'very_good':0,'good':0,'average_grade':0,'poor':0}
    vals=[float(x) for x in percentages];n=len(vals)
    return {'count':n,'average':round(sum(vals)/n,1),'highest':round(max(vals),1),'lowest':round(min(vals),1),'pass_pct':round(sum(1 for x in vals if x>=40)*100/n,1),'outstanding':sum(1 for x in vals if x>=90),'very_good':sum(1 for x in vals if 75<=x<90),'good':sum(1 for x in vals if 60<=x<75),'average_grade':sum(1 for x in vals if 40<=x<60),'poor':sum(1 for x in vals if x<40)}


def outcome_mapping_tokens(value):
    return [x.strip().upper() for x in str(value or '').replace(';',',').split(',') if x.strip()]


def institutional_analytics_data(s):
    raw=s.execute(select(Attempt,Student,Exam).join(Student,Student.id==Attempt.student_id).join(Exam,Exam.id==Attempt.exam_id).where(Attempt.status=='submitted',Attempt.grading_status=='complete')).all()
    percentages=[];exam_pcts={};student_pcts={}
    for a,st,e in raw:
        pct=((a.score or 0)/(a.total_marks or 1))*100;percentages.append(pct);exam_pcts.setdefault(e.id,[]).append(pct);student_pcts.setdefault(st.id,[]).append(pct)
    overall=_score_summary(percentages)
    exam_rows=[]
    exams_all=s.scalars(select(Exam).order_by(Exam.id.desc())).all();total_students=s.scalar(select(func.count()).select_from(Student)) or 0
    for e in exams_all:
        sessions=s.scalars(select(ExamSession).where(ExamSession.exam_id==e.id)).all();group_ids={x.group_id for x in sessions}
        if group_ids:
            strength=s.scalar(select(func.count(func.distinct(StudentGroup.student_id))).where(StudentGroup.group_id.in_(group_ids))) or 0
        else:strength=total_students
        summary=_score_summary(exam_pcts.get(e.id,[]));appeared=summary['count'];exam_rows.append({'id':e.id,'title':e.title,'strength':int(strength),'appeared':appeared,'absent':max(0,int(strength)-appeared),'average':summary['average'],'highest':summary['highest'],'lowest':summary['lowest'],'pass_pct':summary['pass_pct']})
    group_rows=[];groups=s.scalars(select(AcademicGroup).order_by(AcademicGroup.program,AcademicGroup.semester,AcademicGroup.section)).all();membership=dict(s.execute(select(StudentGroup.student_id,StudentGroup.group_id)).all())
    for g in groups:
        student_ids=[sid for sid,gid in membership.items() if gid==g.id];vals=[]
        for sid in student_ids:vals.extend(student_pcts.get(sid,[]))
        summary=_score_summary(vals);group_rows.append({'label':group_label(g),'students':len(student_ids),'attempts':summary['count'],'average':summary['average'],'pass_pct':summary['pass_pct']})
    # Academic attainment from reusable bank metadata.
    maps=s.scalars(select(ExamBankMap)).all();bank_by_exam_q={m.exam_question_id:m.bank_question_id for m in maps};bank_ids=set(bank_by_exam_q.values());banks={b.id:b for b in (s.scalars(select(BankQuestion).where(BankQuestion.id.in_(bank_ids))).all() if bank_ids else [])};questions={q.id:q for q in (s.scalars(select(Question).where(Question.id.in_(list(bank_by_exam_q)))).all() if bank_by_exam_q else [])};submitted_ids={a.id for a,_,_ in raw};unit_stats={};co_stats={};po_stats={};pso_stats={}
    if submitted_ids and bank_by_exam_q:
        answers=s.scalars(select(Answer).where(Answer.attempt_id.in_(submitted_ids),Answer.question_id.in_(list(bank_by_exam_q)))).all()
        for ans in answers:
            bank=banks.get(bank_by_exam_q.get(ans.question_id));q=questions.get(ans.question_id)
            if not bank or not q:continue
            if canonical_question_type(q.question_type)=='essay':
                if ans.manual_score is None:continue
                ok=(max(0,min(ans.manual_score,q.marks))/q.marks) if q.marks else 0
            else:ok=1 if is_answer_correct(q,answer_record_value(ans)) else 0
            if bank.unit:
                key=f'{bank.subject} · Unit {bank.unit}';st=unit_stats.setdefault(key,[0,0]);st[0]+=1;st[1]+=ok
            if bank.co_mapping:
                key=f'{bank.subject} · {bank.co_mapping}';st=co_stats.setdefault(key,[0,0]);st[0]+=1;st[1]+=ok
            for token in outcome_mapping_tokens(getattr(bank,'po_mapping','')):
                key=f'{bank.subject} · {token}';st=po_stats.setdefault(key,[0,0]);st[0]+=1;st[1]+=ok
            for token in outcome_mapping_tokens(getattr(bank,'pso_mapping','')):
                key=f'{bank.subject} · {token}';st=pso_stats.setdefault(key,[0,0]);st[0]+=1;st[1]+=ok
    unit_rows=[{'label':k,'responses':v[0],'attainment':round(v[1]*100/v[0],1) if v[0] else 0} for k,v in sorted(unit_stats.items())]
    co_rows=[{'label':k,'responses':v[0],'attainment':round(v[1]*100/v[0],1) if v[0] else 0} for k,v in sorted(co_stats.items())]
    po_rows=[{'label':k,'responses':v[0],'attainment':round(v[1]*100/v[0],1) if v[0] else 0} for k,v in sorted(po_stats.items())]
    pso_rows=[{'label':k,'responses':v[0],'attainment':round(v[1]*100/v[0],1) if v[0] else 0} for k,v in sorted(pso_stats.items())]
    return overall,exam_rows,group_rows,unit_rows,co_rows,po_rows,pso_rows

@app.route('/admin/analytics')
@staff_required
def analytics():
    s=DB();bank=s.scalars(select(BankQuestion).order_by(BankQuestion.subject,BankQuestion.unit,BankQuestion.id)).all();maps=s.scalars(select(ExamBankMap)).all();bank_by_exam_q={m.exam_question_id:m.bank_question_id for m in maps};exam_q_ids=list(bank_by_exam_q)
    qmap={q.id:q for q in (s.scalars(select(Question).where(Question.id.in_(exam_q_ids))).all() if exam_q_ids else [])};attempts=s.scalars(select(Attempt).where(Attempt.status=='submitted',Attempt.grading_status=='complete')).all();attempt_pct={a.id:((a.score or 0)/(a.total_marks or 1)) for a in attempts};submitted_ids=set(attempt_pct)
    stats={q.id:{'responses':0,'correct':0,'samples':[]} for q in bank}
    if exam_q_ids and submitted_ids:
        answers=s.scalars(select(Answer).where(Answer.question_id.in_(exam_q_ids),Answer.attempt_id.in_(submitted_ids))).all()
        for a in answers:
            bid=bank_by_exam_q.get(a.question_id)
            if bid not in stats:continue
            question=qmap.get(a.question_id)
            if not question:continue
            if canonical_question_type(question.question_type)=='essay':
                if a.manual_score is None:continue
                ok=(max(0,min(a.manual_score,question.marks))/question.marks) if question.marks else 0
            else:ok=1 if is_answer_correct(question,answer_record_value(a)) else 0
            stats[bid]['responses']+=1;stats[bid]['correct']+=ok;stats[bid]['samples'].append((attempt_pct.get(a.attempt_id,0),ok))
    usage=dict(s.execute(select(ExamBankMap.bank_question_id,func.count(func.distinct(ExamBankMap.exam_id))).group_by(ExamBankMap.bank_question_id)).all());rows=[]
    for q in bank:
        st=stats[q.id];rate=(st['correct']/st['responses']) if st['responses'] else None;disc=None
        samples=sorted(st['samples'],key=lambda x:x[0])
        if len(samples)>=6:
            n=max(1,round(len(samples)*0.27));bottom=samples[:n];top=samples[-n:];disc=(sum(x[1] for x in top)/n)-(sum(x[1] for x in bottom)/n)
        observed='-' if rate is None else ('Easy' if rate>=0.75 else 'Medium' if rate>=0.4 else 'Hard')
        rows.append(type('AnalyticsRow',(),{'id':q.id,'subject':q.subject,'unit':q.unit,'topic':q.topic,'question':q.question,'declared_difficulty':q.difficulty,'times_used':usage.get(q.id,0),'responses':st['responses'],'correct_rate':rate,'observed_difficulty':observed,'discrimination':disc})())
    overall,exam_rows,group_rows,unit_rows,co_rows,po_rows,pso_rows=institutional_analytics_data(s);return render_template('analytics.html',rows=rows,overall=overall,exam_rows=exam_rows,group_rows=group_rows,unit_rows=unit_rows,co_rows=co_rows,po_rows=po_rows,pso_rows=pso_rows)


@app.route('/admin/institution',methods=['GET','POST'])
@admin_required
def institution_settings():
    s=DB();row=get_institution(s,create=True)
    if request.method=='POST':
        row.institution_name=request.form.get('institution_name','').strip() or 'Learn with Hemant';row.short_name=request.form.get('short_name','').strip() or row.institution_name;row.system_name=request.form.get('system_name','').strip() or 'Examination Management System';row.department=request.form.get('department','').strip();row.academic_year=request.form.get('academic_year','').strip();row.admin_email=request.form.get('admin_email','').strip();row.exam_controller=request.form.get('exam_controller','').strip();row.contact_phone=request.form.get('contact_phone','').strip()
        logo=request.files.get('logo_file')
        if logo and logo.filename:
            raw=logo.read();mime=(logo.mimetype or '').lower()
            if mime not in {'image/png','image/jpeg','image/webp'}:flash('Logo must be PNG, JPG or WebP.','error');return redirect(url_for('institution_settings'))
            if len(raw)>1024*1024:flash('Logo must be smaller than 1 MB.','error');return redirect(url_for('institution_settings'))
            row.logo_data=f'data:{mime};base64,'+base64.b64encode(raw).decode('ascii')
        if request.form.get('remove_logo')=='1':row.logo_data=''
        row.updated_at=now_iso();audit_event(s,'institution_profile_updated','system',row.id,row.institution_name);s.commit();flash('Institution branding and profile updated.');return redirect(url_for('institution_settings'))
    return render_template('institution.html',profile=row)

def exam_centre_network_details():
    port=int(os.getenv('PORT','8080'))
    lan_ip=local_lan_ip()
    student_url=request.url_root.rstrip('/') if APP_MODE=='online' else f'http://{lan_ip}:{port}'
    return {'mode':APP_MODE,'lan_ip':lan_ip,'port':port,'student_url':student_url,'qr_uri':qr_data_uri(student_url)}


def exam_centre_live_payload(s):
    raw=s.execute(select(Attempt,Student,Exam).join(Student,Student.id==Attempt.student_id).join(Exam,Exam.id==Attempt.exam_id).where(Attempt.status=='in_progress').order_by(Attempt.id.desc()).limit(300)).all()
    attempt_ids=[a.id for a,_,_ in raw]
    heartbeats={h.attempt_id:h for h in (s.scalars(select(AttemptHeartbeat).where(AttemptHeartbeat.attempt_id.in_(attempt_ids))).all() if attempt_ids else [])}
    events=dict(s.execute(select(IntegrityEvent.attempt_id,func.count()).where(IntegrityEvent.attempt_id.in_(attempt_ids)).group_by(IntegrityEvent.attempt_id)).all()) if attempt_ids else {}
    answers=dict(s.execute(select(Answer.attempt_id,func.count()).where(Answer.attempt_id.in_(attempt_ids)).group_by(Answer.attempt_id)).all()) if attempt_ids else {}
    rows=[]
    now=now_dt()
    for attempt,student,exam in raw:
        hb=heartbeats.get(attempt.id);state=heartbeat_status(hb.last_seen_at if hb else '')
        try:remaining=max(0,int((parse_dt(attempt.end_at)-now).total_seconds()))
        except Exception:remaining=0
        rows.append({'attempt_id':attempt.id,'roll_no':student.roll_no,'name':student.name,'exam_id':exam.id,'exam':exam.title,'connection':state,'last_seen':hb.last_seen_at if hb else '','answers':int(answers.get(attempt.id,0)),'integrity':int(events.get(attempt.id,0)),'remaining_seconds':remaining})
    counts={'online':sum(1 for x in rows if x['connection']=='online'),'stale':sum(1 for x in rows if x['connection']=='stale'),'offline':sum(1 for x in rows if x['connection'] in {'offline','unknown'})}
    return {'generated_at':now_iso(),'rows':rows,'counts':counts}


@app.route('/admin/exam-centre')
@staff_required
def exam_centre():
    s=DB();network=exam_centre_network_details();active_exams=s.scalars(select(Exam).where(Exam.is_active==True).order_by(Exam.title)).all();live=exam_centre_live_payload(s)
    stats={'students':s.scalar(select(func.count()).select_from(Student)) or 0,'active_exams':len(active_exams),'in_progress':len(live['rows']),'online_now':live['counts']['online']}
    return render_template('exam_centre.html',mode=network['mode'],db_name='PostgreSQL' if DATABASE_URL.startswith('postgresql') else 'SQLite',student_url=network['student_url'],qr_uri=network['qr_uri'],lan_ip=network['lan_ip'],port=network['port'],stats=stats,active_exams=active_exams,live=live)


@app.route('/admin/exam-centre/live-data')
@staff_required
def exam_centre_live_data():
    response=jsonify(exam_centre_live_payload(DB()));response.headers['Cache-Control']='no-store';return response


@app.route('/admin/exam-centre/check-in',methods=['POST'])
@staff_required
def exam_centre_check_in():
    s=DB();exam_id=request.form.get('exam_id',type=int);roll_no=(request.form.get('roll_no') or '').strip();exam=s.get(Exam,exam_id) if exam_id else None;student=s.scalar(select(Student).where(Student.roll_no==roll_no)) if roll_no else None
    if not exam or not student:flash('Choose a valid exam and student roll number.','error');return redirect(url_for('exam_centre'))
    row=s.scalar(select(ExamCandidateCheckin).where(ExamCandidateCheckin.exam_id==exam.id,ExamCandidateCheckin.student_id==student.id))
    if not row:row=ExamCandidateCheckin(exam_id=exam.id,student_id=student.id,status='verified',verified_by=actor_label(s),verified_at=now_iso(),notes=(request.form.get('notes') or '').strip()[:250]);s.add(row)
    else:row.status='verified';row.verified_by=actor_label(s);row.verified_at=now_iso();row.notes=(request.form.get('notes') or '').strip()[:250]
    audit_event(s,'candidate_identity_verified','student',student.id,f'exam={exam.id}, roll={student.roll_no}');s.commit();flash(f'{student.roll_no} verified for {exam.title}.');return redirect(url_for('exam_centre'))


@app.route('/admin/exam-centre/check-in/revoke',methods=['POST'])
@staff_required
def exam_centre_revoke_checkin():
    s=DB();exam_id=request.form.get('exam_id',type=int);roll_no=(request.form.get('roll_no') or '').strip();student=s.scalar(select(Student).where(Student.roll_no==roll_no)) if roll_no else None
    row=s.scalar(select(ExamCandidateCheckin).where(ExamCandidateCheckin.exam_id==exam_id,ExamCandidateCheckin.student_id==student.id)) if exam_id and student else None
    if row:row.status='revoked';audit_event(s,'candidate_identity_revoked','student',student.id,f'exam={exam_id}');s.commit();flash(f'Check-in revoked for {student.roll_no}.')
    else:flash('No verified check-in was found.','error')
    return redirect(url_for('exam_centre'))


@app.route('/admin/exam-centre/network-info')
@staff_required
def exam_centre_network_info():
    response=jsonify(exam_centre_network_details())
    response.headers['Cache-Control']='no-store, no-cache, must-revalidate, max-age=0'
    return response

@app.route('/admin/faculty',methods=['GET','POST'])
@staff_required
def faculty_users():
    s=DB();viewer_role=current_staff_role(s)
    if viewer_role not in {'super_admin','hod'}:
        abort(403)
    if request.method=='POST':
        # Only the Super Admin may create staff accounts or assign privileged roles.
        if viewer_role!='super_admin':
            abort(403)
        username=request.form.get('username','').strip();name=request.form.get('name','').strip();password=request.form.get('password','')
        if len(username)<3 or not name or len(password)<10:flash('Faculty name, a 3+ character username and a 10+ character password are required.','error')
        elif username.casefold()==super_admin_username.casefold():flash('That username is reserved for the Super Admin account.','error')
        else:
            try:
                role=request.form.get('staff_role','faculty') if request.form.get('staff_role') in {'faculty','hod','exam_controller'} else 'faculty'
                row=Faculty(username=username,name=name,password_hash=generate_password_hash(password),is_active=True,created_at=now_iso());s.add(row);s.flush();s.add(FacultyRole(faculty_id=row.id,role=role,department=request.form.get('department','').strip(),updated_at=now_iso()));audit_event(s,'faculty_created','faculty',row.id,f'{username}, role={role}');s.commit();flash('Staff login created.')
            except IntegrityError:s.rollback();flash('That faculty username already exists.','error')
    rows=s.scalars(select(Faculty).order_by(Faculty.username)).all()
    role_map={r.faculty_id:r for r in s.scalars(select(FacultyRole)).all()}
    # HOD users are deliberately shown Faculty accounts only. They cannot view or
    # manage another HOD, an Exam Controller, or the Super Admin account here.
    if viewer_role=='hod':
        rows=[f for f in rows if (role_map.get(f.id).role if role_map.get(f.id) else 'faculty')=='faculty']
    return render_template('faculty.html',faculty=rows,role_map=role_map,role_labels=ROLE_LABELS,viewer_role=viewer_role)

@app.route('/admin/faculty/<int:faculty_id>/password',methods=['POST'])
@staff_required
def reset_faculty_password(faculty_id):
    s=DB();actor_role=current_staff_role(s)
    if actor_role not in {'super_admin','hod'}:
        abort(403)
    target=s.get(Faculty,faculty_id)
    if not target:
        abort(404)
    target_role_row=s.scalar(select(FacultyRole).where(FacultyRole.faculty_id==faculty_id))
    target_role=target_role_row.role if target_role_row and target_role_row.role in ROLE_LABELS else 'faculty'
    if actor_role=='hod':
        # HOD may reset only ordinary Faculty accounts, never HOD/Controller/Super Admin.
        if target_role!='faculty':
            abort(403)
        if web_session.get('role')=='faculty' and int(web_session.get('user_id') or 0)==faculty_id:
            abort(403)
    password=request.form.get('new_password','')
    confirm=request.form.get('confirm_password','')
    if len(password)<10:
        flash('New password must contain at least 10 characters.','error')
        return redirect(url_for('faculty_users'))
    if password!=confirm:
        flash('New password and confirmation do not match.','error')
        return redirect(url_for('faculty_users'))
    target.password_hash=generate_password_hash(password)
    audit_event(s,'staff_password_reset','faculty',target.id,f'target={target.username}, role={target_role}')
    s.commit()
    flash(f'Password updated for {target.username}.')
    return redirect(url_for('faculty_users'))

@app.route('/admin/faculty/<int:faculty_id>/mfa-reset',methods=['POST'])
@staff_required
def reset_faculty_mfa(faculty_id):
    s=DB();actor_role=current_staff_role(s)
    if actor_role not in {'super_admin','hod'}:abort(403)
    target=s.get(Faculty,faculty_id)
    if not target:abort(404)
    role_row=s.scalar(select(FacultyRole).where(FacultyRole.faculty_id==faculty_id));target_role=role_row.role if role_row and role_row.role in ROLE_LABELS else 'faculty'
    if actor_role=='hod' and target_role!='faculty':abort(403)
    target.mfa_enabled=False;target.mfa_secret='';audit_event(s,'staff_mfa_admin_reset','faculty',target.id,f'target={target.username}, role={target_role}');s.commit();flash(f'MFA reset for {target.username}. They can sign in with their password and enroll again.');return redirect(url_for('faculty_users'))


@app.route('/admin/faculty/<int:faculty_id>/toggle',methods=['POST'])
@admin_required
def toggle_faculty(faculty_id):
    s=DB();row=s.get(Faculty,faculty_id)
    if row:row.is_active=not row.is_active;audit_event(s,'faculty_enabled' if row.is_active else 'faculty_disabled','faculty',row.id,row.username);s.commit()
    return redirect(url_for('faculty_users'))


@app.route('/admin/faculty/<int:faculty_id>/role',methods=['POST'])
@admin_required
def update_faculty_role(faculty_id):
    s=DB();faculty=s.get(Faculty,faculty_id)
    if not faculty:abort(404)
    role=request.form.get('staff_role','faculty')
    if role not in {'faculty','hod','exam_controller'}:flash('Invalid staff role.','error');return redirect(url_for('faculty_users'))
    row=s.scalar(select(FacultyRole).where(FacultyRole.faculty_id==faculty_id))
    if row:row.role=role;row.department=request.form.get('department','').strip();row.updated_at=now_iso()
    else:s.add(FacultyRole(faculty_id=faculty_id,role=role,department=request.form.get('department','').strip(),updated_at=now_iso()))
    audit_event(s,'faculty_role_updated','faculty',faculty_id,f'role={role}');s.commit();flash('Staff role updated.');return redirect(url_for('faculty_users'))

@app.route('/admin/audit')
@admin_required
def audit_logs():
    s=DB();rows=s.scalars(select(AuditLog).order_by(AuditLog.id.desc()).limit(300)).all();chain_status=audit_chain_status(s);return render_template('audit.html',rows=rows,chain_status=chain_status)

@app.route('/admin/system')
@admin_required
def system_tools():
    s=DB();return render_template('system.html',offline=(APP_MODE=='offline'),institution=get_institution(s,create=True))

@app.route('/admin/system/backup')
@admin_required
def system_backup():
    if APP_MODE!='offline':abort(400,'Direct database backup is available only in offline mode.')
    db_path=DATA_DIR/'exam.db'
    if not db_path.exists():abort(404)
    tmp=tempfile.NamedTemporaryFile(prefix='lwh_exam_backup_',suffix='.db',delete=False);tmp.close()
    src=sqlite3.connect(str(db_path));dst=sqlite3.connect(tmp.name)
    try:src.backup(dst)
    finally:dst.close();src.close()
    @after_this_request
    def cleanup_backup(response):
        try:os.unlink(tmp.name)
        except OSError:pass
        return response
    stamp=now_dt().strftime('%Y%m%d_%H%M');return send_file(tmp.name,as_attachment=True,download_name=f'LearnWithHemant_Exam_Backup_{stamp}.db',mimetype='application/octet-stream')


@app.route('/admin/system/backup/encrypted',methods=['POST'])
@admin_required
def system_backup_encrypted():
    if APP_MODE!='offline':abort(400,'Encrypted database backup is available only in offline mode.')
    password=request.form.get('backup_password','')
    if len(password)<10:flash('Use a backup password with at least 10 characters.','error');return redirect(url_for('system_tools'))
    db_path=DATA_DIR/'exam.db'
    if not db_path.exists():abort(404)
    tmp=tempfile.NamedTemporaryFile(prefix='lwh_exam_backup_',suffix='.db',delete=False);tmp.close();src=sqlite3.connect(str(db_path));dst=sqlite3.connect(tmp.name)
    try:src.backup(dst)
    finally:dst.close();src.close()
    try:raw=Path(tmp.name).read_bytes();payload=encrypt_backup_bytes(raw,password)
    finally:
        try:os.unlink(tmp.name)
        except OSError:pass
    s=DB();audit_event(s,'encrypted_backup_downloaded','system','','offline database');s.commit();data=io.BytesIO(payload);stamp=now_dt().strftime('%Y%m%d_%H%M');return send_file(data,as_attachment=True,download_name=f'ExamSystem_Encrypted_Backup_{stamp}.lwhbackup',mimetype='application/octet-stream')

@app.route('/admin/system/restore',methods=['POST'])
@admin_required
def system_restore():
    if APP_MODE!='offline':abort(400,'Direct database restore is available only in offline mode.')
    upload=request.files.get('backup_file')
    if not upload or not upload.filename:flash('Choose a .db or .lwhbackup backup file.','error');return redirect(url_for('system_tools'))
    fd,temp_name=tempfile.mkstemp(prefix='lwh_restore_',suffix='.db'); os.close(fd); temp=Path(temp_name)
    try:
        raw=upload.read()
        if upload.filename.lower().endswith('.lwhbackup') or raw.startswith(b'LWHBK1'):
            password=request.form.get('backup_password','')
            if not password:raise ValueError('Enter the backup password used when the encrypted backup was created.')
            raw=decrypt_backup_bytes(raw,password)
        temp.write_bytes(raw)
        conn=sqlite3.connect(str(temp));tables={r[0] for r in conn.execute("select name from sqlite_master where type='table'").fetchall()};conn.close()
        if not {'admins','students','exams','questions','attempts','answers'}.issubset(tables):raise ValueError('This is not a valid Learn with Hemant exam backup.')
        DB.remove();engine.dispose();shutil.copy2(temp,DATA_DIR/'exam.db');init_db();rs=DB();audit_event(rs,'offline_backup_restored','system','',upload.filename or 'backup.db');rs.commit();flash('Backup restored successfully. Refresh the page and verify the data before conducting an exam.')
    except Exception as exc:flash(f'Restore failed: {exc}','error')
    finally:
        try:temp.unlink()
        except OSError:pass
    return redirect(url_for('system_tools'))

# ------------------------- Student Practice & Learning Centre -------------------------
def _practice_ref_key(ref):
    return f"{ref.get('source','bank')}:{int(ref.get('id') or 0)}"


def _practice_ref_field(ref):
    return f"q_{ref.get('source','bank')}_{int(ref.get('id') or 0)}"


def _practice_question_view(s,ref):
    source=str(ref.get('source') or 'bank')
    source_id=int(ref.get('id') or 0)
    bank=None
    if source=='bank':
        question=s.get(BankQuestion,source_id);bank=question
    elif source=='exam':
        question=s.get(Question,source_id)
        map_row=s.scalar(select(ExamBankMap).where(ExamBankMap.exam_question_id==source_id))
        bank=s.get(BankQuestion,map_row.bank_question_id) if map_row else None
    else:
        question=None
    if not question:return None
    qtype=canonical_question_type(question.question_type)
    order=str(ref.get('option_order') or 'ABCD')
    if len(order)!=4 or set(order)!=set('ABCD'):order='ABCD'
    text_map={'A':question.option_a,'B':question.option_b,'C':question.option_c,'D':question.option_d}
    options=[]
    if qtype in {'single_choice','multiple_select'}:
        options=[{'label':chr(65+i),'key':key,'text':text_map.get(key,'')} for i,key in enumerate(order)]
    subject=(bank.subject if bank else '')
    unit=(bank.unit if bank else '')
    topic=(bank.topic if bank else '')
    return {
        'source':source,'source_id':source_id,'ref_key':_practice_ref_key(ref),'field_name':_practice_ref_field(ref),
        'bank_id':bank.id if bank else None,'question':question.question,'question_type':qtype,
        'question_type_label':QUESTION_TYPE_LABELS.get(qtype,qtype.replace('_',' ').title()),'display_options':options,
        'marks':int(question.marks or 1),'subject':subject,'unit':unit,'topic':topic,
        'difficulty':bank.difficulty if bank else '','co_mapping':bank.co_mapping if bank else '',
        'explanation':(bank.explanation or '').strip() if bank else '',
        'question_obj':question,'option_order':order,
    }


def _practice_answer_display(view,value):
    value=(value or '').strip();qtype=view['question_type']
    if not value:return 'Not answered'
    if qtype in {'single_choice','multiple_select'}:
        selected={part.strip().upper() for part in value.split(',') if part.strip()}
        return '; '.join(f"{opt['label']}. {opt['text']}" for opt in view['display_options'] if opt['key'] in selected) or value
    if qtype=='true_false':return value.title()
    return value


def _practice_correct_display(view):
    q=view['question_obj'];qtype=view['question_type'];value=normalized_key(q)
    if qtype=='essay':return 'Self-review question (not automatically graded)'
    return _practice_answer_display(view,value)


def _practice_bank_refs(rows):
    refs=[]
    for row in rows:
        keys=list('ABCD')
        if canonical_question_type(row.question_type) in {'single_choice','multiple_select'}:random.shuffle(keys)
        refs.append({'source':'bank','id':row.id,'bank_id':row.id,'option_order':''.join(keys)})
    return refs


def _create_practice_attempt(s,student_id,refs,mode,subject='',unit_filter='',difficulty_filter='',exam_id=None,duration_minutes=0):
    started=now_dt();duration=max(0,int(duration_minutes or 0));ends=(started+timedelta(minutes=duration)).isoformat(timespec='seconds') if duration else ''
    attempt=PracticeAttempt(student_id=student_id,mode=mode,subject=subject or 'Mixed Practice',unit_filter=unit_filter or '',difficulty_filter=difficulty_filter or '',exam_id=exam_id,duration_minutes=duration,started_at=started.isoformat(timespec='seconds'),ends_at=ends,submitted_at='',status='in_progress',score=0,total_marks=0,question_refs_json=json.dumps(refs,separators=(',',':')),answers_json='[]',incorrect_bank_ids_json='[]')
    s.add(attempt);s.flush();return attempt


def _practice_wrong_bank_ids(s,student_id):
    ids=set()
    rows=s.scalars(select(PracticeAttempt).where(PracticeAttempt.student_id==student_id,PracticeAttempt.status=='submitted')).all()
    for row in rows:
        for value in safe_json_load(row.incorrect_bank_ids_json,[]):
            try:ids.add(int(value))
            except Exception:pass
    return ids


def _practice_student_metrics(s,student_id):
    rows=s.scalars(select(PracticeAttempt).where(PracticeAttempt.student_id==student_id,PracticeAttempt.status=='submitted').order_by(PracticeAttempt.id.desc())).all()
    subject_stats={};unit_stats={};total_answered=total_correct=0
    for attempt in rows:
        for item in safe_json_load(attempt.answers_json,[]):
            if item.get('is_manual'):continue
            marks=max(1,int(item.get('marks') or 1));earned=max(0,int(item.get('earned') or 0));total_answered+=1
            if earned>=marks:total_correct+=1
            subject=(item.get('subject') or attempt.subject or 'General').strip() or 'General'
            unit=(item.get('unit') or '').strip()
            for key,bucket in [(subject,subject_stats),(f'{subject} · Unit {unit}' if unit else subject,unit_stats)]:
                stat=bucket.setdefault(key,{'label':key,'earned':0,'marks':0,'questions':0})
                stat['earned']+=earned;stat['marks']+=marks;stat['questions']+=1
    def finish(mapping):
        out=[]
        for stat in mapping.values():
            stat['percentage']=round(100*stat['earned']/stat['marks']) if stat['marks'] else 0;out.append(stat)
        return out
    subjects=sorted(finish(subject_stats),key=lambda x:(x['percentage'],x['label']))
    units=sorted(finish(unit_stats),key=lambda x:(x['percentage'],x['label']))
    return {'attempts':len(rows),'answered':total_answered,'correct':total_correct,'accuracy':round(100*total_correct/total_answered) if total_answered else 0,'subjects':subjects,'units':units,'weak_areas':units[:5]}


def _practice_released_exams(s):
    releases=s.scalars(select(ExamPracticeRelease).where(ExamPracticeRelease.is_released==True).order_by(ExamPracticeRelease.id.desc())).all();rows=[]
    for release in releases:
        if not practice_release_is_available(release):continue
        exam=s.get(Exam,release.exam_id)
        if not exam or exam.is_active:continue
        qcount=s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==exam.id)) or 0
        rows.append({'exam':exam,'release':release,'question_count':qcount})
    return rows


@app.route('/student/practice')
@student_required
def practice_centre():
    s=DB();student=s.get(Student,web_session['user_id'])
    eligible=s.scalars(select(BankQuestion).where(BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both'])).order_by(BankQuestion.subject,BankQuestion.unit,BankQuestion.id)).all()
    eligible=[q for q in eligible if canonical_question_type(q.question_type)!='essay']
    subject_counts={};units_by_subject={}
    for q in eligible:
        subject_counts[q.subject]=subject_counts.get(q.subject,0)+1
        if q.unit:units_by_subject.setdefault(q.subject,set()).add(q.unit)
    subjects=[{'name':name,'count':count,'units':sorted(units_by_subject.get(name,set()))} for name,count in sorted(subject_counts.items())]
    recent=s.scalars(select(PracticeAttempt).where(PracticeAttempt.student_id==student.id).order_by(PracticeAttempt.id.desc()).limit(10)).all()
    metrics=_practice_student_metrics(s,student.id)
    bookmarks=s.scalar(select(func.count()).select_from(PracticeBookmark).where(PracticeBookmark.student_id==student.id)) or 0
    wrong_ids=_practice_wrong_bank_ids(s,student.id)
    published_ids={q.id for q in eligible};wrong_available=len(wrong_ids & published_ids)
    return render_template('practice_centre.html',student=student,subjects=subjects,published_count=len(eligible),recent=recent,metrics=metrics,bookmark_count=bookmarks,wrong_count=wrong_available,released_exams=_practice_released_exams(s))


@app.route('/student/practice/start',methods=['POST'])
@student_required
def start_practice():
    s=DB();subject=(request.form.get('subject') or '').strip();unit=(request.form.get('unit') or '').strip();difficulty=(request.form.get('difficulty') or '').strip();mode=(request.form.get('mode') or 'practice').strip().lower()
    if mode not in {'practice','mock'}:mode='practice'
    try:count=max(1,min(50,int(request.form.get('question_count') or 10)))
    except ValueError:count=10
    stmt=select(BankQuestion).where(BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both']))
    if subject:stmt=stmt.where(BankQuestion.subject==subject)
    if unit:stmt=stmt.where(BankQuestion.unit==unit)
    if difficulty:stmt=stmt.where(BankQuestion.difficulty==canonical_difficulty(difficulty))
    rows=[q for q in s.scalars(stmt).all() if canonical_question_type(q.question_type)!='essay']
    if not rows:flash('No approved questions are currently published for that practice selection. Ask your faculty to publish practice questions.','error');return redirect(url_for('practice_centre'))
    selected=random.sample(rows,min(count,len(rows)));refs=_practice_bank_refs(selected)
    duration=0
    if mode=='mock':
        try:duration=max(5,min(180,int(request.form.get('duration_minutes') or max(10,len(refs)))))
        except ValueError:duration=max(10,len(refs))
    attempt=_create_practice_attempt(s,web_session['user_id'],refs,mode,subject or 'Mixed Practice',unit,difficulty,None,duration);s.commit();return redirect(url_for('practice_attempt',attempt_id=attempt.id))


@app.route('/student/practice/previous/<int:exam_id>/start',methods=['POST'])
@student_required
def start_previous_paper_practice(exam_id):
    s=DB();exam=s.get(Exam,exam_id);release=get_exam_practice_release(s,exam_id,create=False)
    if not exam or exam.is_active or not practice_release_is_available(release):abort(404)
    questions=s.scalars(select(Question).where(Question.exam_id==exam.id).order_by(Question.id)).all()
    if not questions:flash('This released paper does not contain any questions.','error');return redirect(url_for('practice_centre'))
    maps={m.exam_question_id:m.bank_question_id for m in s.scalars(select(ExamBankMap).where(ExamBankMap.exam_id==exam.id)).all()}
    refs=[]
    for q in questions:
        refs.append({'source':'exam','id':q.id,'bank_id':maps.get(q.id),'option_order':'ABCD'})
    timed=request.form.get('timed')=='1' and bool(release.allow_mock);duration=exam.duration_minutes if timed else 0
    attempt=_create_practice_attempt(s,web_session['user_id'],refs,'previous',exam.title,'','',exam.id,duration);s.commit();return redirect(url_for('practice_attempt',attempt_id=attempt.id))


@app.route('/student/practice/wrong/start',methods=['POST'])
@student_required
def start_wrong_answer_practice():
    s=DB();ids=_practice_wrong_bank_ids(s,web_session['user_id'])
    if not ids:flash('You do not have any incorrect practice questions to retry yet.');return redirect(url_for('practice_centre'))
    rows=[q for q in s.scalars(select(BankQuestion).where(BankQuestion.id.in_(ids),BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both']))).all() if canonical_question_type(q.question_type)!='essay']
    if not rows:flash('Your previously missed questions are no longer published for practice.','error');return redirect(url_for('practice_centre'))
    random.shuffle(rows);rows=rows[:min(30,len(rows))];attempt=_create_practice_attempt(s,web_session['user_id'],_practice_bank_refs(rows),'wrong','My Wrong Answers');s.commit();return redirect(url_for('practice_attempt',attempt_id=attempt.id))


@app.route('/student/practice/bookmarks/start',methods=['POST'])
@student_required
def start_bookmarked_practice():
    s=DB();ids=s.scalars(select(PracticeBookmark.bank_question_id).where(PracticeBookmark.student_id==web_session['user_id'])).all()
    if not ids:flash('You have not bookmarked any practice questions yet.');return redirect(url_for('practice_centre'))
    rows=[q for q in s.scalars(select(BankQuestion).where(BankQuestion.id.in_(list(ids)),BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both']))).all() if canonical_question_type(q.question_type)!='essay']
    if not rows:flash('Your bookmarked questions are not currently published for practice.','error');return redirect(url_for('practice_centre'))
    random.shuffle(rows);rows=rows[:min(30,len(rows))];attempt=_create_practice_attempt(s,web_session['user_id'],_practice_bank_refs(rows),'bookmarks','Bookmarked Questions');s.commit();return redirect(url_for('practice_attempt',attempt_id=attempt.id))


@app.route('/student/practice/<int:attempt_id>')
@student_required
def practice_attempt(attempt_id):
    s=DB();attempt=s.get(PracticeAttempt,attempt_id)
    if not attempt or attempt.student_id!=web_session['user_id']:abort(404)
    if attempt.status=='submitted':return redirect(url_for('practice_result',attempt_id=attempt.id))
    refs=safe_json_load(attempt.question_refs_json,[]);views=[]
    for ref in refs:
        view=_practice_question_view(s,ref)
        if view:views.append(view)
    if not views:flash('This practice set is no longer available.','error');return redirect(url_for('practice_centre'))
    feedback_enabled=attempt.mode in {'practice','wrong','bookmarks'}
    end_epoch=parse_dt(attempt.ends_at).timestamp() if attempt.ends_at else 0
    return render_template('practice_attempt.html',attempt=attempt,questions=views,feedback_enabled=feedback_enabled,end_epoch=end_epoch)


@app.route('/student/practice/check-answer',methods=['POST'])
@student_required
def practice_check_answer():
    data=request.get_json(silent=True) or {}
    try:attempt_id=int(data.get('attempt_id') or 0)
    except Exception:return jsonify(error='Invalid practice attempt'),400
    ref_key=str(data.get('ref_key') or '');value=str(data.get('answer') or '')[:MAX_ANSWER_LENGTH]
    s=DB();attempt=s.get(PracticeAttempt,attempt_id)
    if not attempt or attempt.student_id!=web_session['user_id'] or attempt.status!='in_progress':return jsonify(error='Practice attempt not available'),404
    if attempt.mode not in {'practice','wrong','bookmarks'}:return jsonify(error='Immediate feedback is disabled in timed/previous-paper mode'),403
    ref=next((r for r in safe_json_load(attempt.question_refs_json,[]) if _practice_ref_key(r)==ref_key),None)
    if not ref:return jsonify(error='Question is not part of this practice set'),400
    view=_practice_question_view(s,ref)
    if not view:return jsonify(error='Question unavailable'),404
    correct=is_answer_correct(view['question_obj'],value)
    return jsonify(correct=bool(correct),correct_answer=_practice_correct_display(view),explanation=view['explanation'] or 'No explanation has been added by the faculty yet.')


@app.route('/student/practice/<int:attempt_id>/submit',methods=['POST'])
@student_required
def submit_practice_attempt(attempt_id):
    s=DB();attempt=s.get(PracticeAttempt,attempt_id)
    if not attempt or attempt.student_id!=web_session['user_id']:abort(404)
    if attempt.status=='submitted':return redirect(url_for('practice_result',attempt_id=attempt.id))
    refs=safe_json_load(attempt.question_refs_json,[]);answers=[];score=total=0;incorrect=[]
    late_submission=False
    if attempt.ends_at:
        try:late_submission=now_dt()>parse_dt(attempt.ends_at)+timedelta(seconds=60)
        except Exception:late_submission=False
    for ref in refs:
        view=_practice_question_view(s,ref)
        if not view:continue
        values=[] if late_submission else request.form.getlist(view['field_name']);qtype=view['question_type']
        value=','.join(values) if qtype=='multiple_select' else (values[0] if values else '')
        is_manual=qtype=='essay';correct=False if is_manual else is_answer_correct(view['question_obj'],value)
        marks=int(view['marks'] or 1);earned=0 if is_manual else (marks if correct else 0)
        if not is_manual:total+=marks;score+=earned
        bank_id=view['bank_id']
        if bank_id and not is_manual and not correct:incorrect.append(int(bank_id))
        answers.append({'ref_key':view['ref_key'],'source':view['source'],'source_id':view['source_id'],'bank_id':bank_id,'question':view['question'],'question_type':qtype,'question_type_label':view['question_type_label'],'student_answer':_practice_answer_display(view,value),'correct_answer':_practice_correct_display(view),'raw_answer':value,'is_correct':bool(correct),'is_manual':bool(is_manual),'marks':marks,'earned':earned,'subject':view['subject'],'unit':view['unit'],'topic':view['topic'],'difficulty':view['difficulty'],'co_mapping':view['co_mapping'],'explanation':view['explanation'],'options':view['display_options']})
    attempt.score=score;attempt.total_marks=total;attempt.answers_json=json.dumps(answers,ensure_ascii=False,separators=(',',':'));attempt.incorrect_bank_ids_json=json.dumps(sorted(set(incorrect)),separators=(',',':'));attempt.status='submitted';attempt.submitted_at=now_iso();s.commit()
    if late_submission:flash('The timed practice window had already expired, so late answer changes were not accepted.','error')
    return redirect(url_for('practice_result',attempt_id=attempt.id))


@app.route('/student/practice/<int:attempt_id>/result')
@student_required
def practice_result(attempt_id):
    s=DB();attempt=s.get(PracticeAttempt,attempt_id)
    if not attempt or attempt.student_id!=web_session['user_id'] or attempt.status!='submitted':abort(404)
    answers=safe_json_load(attempt.answers_json,[]);bookmarked=set(s.scalars(select(PracticeBookmark.bank_question_id).where(PracticeBookmark.student_id==web_session['user_id'])).all())
    bank_ids={int(item.get('bank_id')) for item in answers if item.get('bank_id')}
    bookmarkable_ids=set(s.scalars(select(BankQuestion.id).where(BankQuestion.id.in_(bank_ids),BankQuestion.status=='approved',BankQuestion.practice_visibility.in_(['practice_only','both']))).all()) if bank_ids else set()
    show_solutions=True
    if attempt.mode=='previous' and attempt.exam_id:
        release=get_exam_practice_release(s,attempt.exam_id,create=False);show_solutions=bool(release and release.show_solutions)
    percentage=round(100*attempt.score/attempt.total_marks) if attempt.total_marks else 0
    return render_template('practice_result.html',attempt=attempt,answers=answers,percentage=percentage,show_solutions=show_solutions,bookmarked=bookmarked,bookmarkable_ids=bookmarkable_ids)


@app.route('/student/practice/bookmark/<int:bank_question_id>',methods=['POST'])
@student_required
def toggle_practice_bookmark(bank_question_id):
    s=DB();question=s.get(BankQuestion,bank_question_id)
    if not question or not question_is_practice_eligible(question):abort(404)
    row=s.scalar(select(PracticeBookmark).where(PracticeBookmark.student_id==web_session['user_id'],PracticeBookmark.bank_question_id==bank_question_id))
    if row:s.delete(row);flash('Bookmark removed.')
    else:s.add(PracticeBookmark(student_id=web_session['user_id'],bank_question_id=bank_question_id,created_at=now_iso()));flash('Question bookmarked for later practice.')
    s.commit();return redirect(request.referrer or url_for('practice_centre'))

# ----------------------- End Student Practice & Learning Centre -----------------------

def student_exam_subject_unit(s, exam, cfg):
    """Return stable subject/unit labels for the student exam dashboard.

    Prefer the mapped Question Bank subject, then fall back to ExamConfig metadata.
    No schema change is required, so older production exams remain compatible.
    """
    configured_subject=((cfg.subject if cfg else '') or '').strip()
    mapped=s.execute(
        select(BankQuestion.subject,BankQuestion.unit)
        .join(ExamBankMap,ExamBankMap.bank_question_id==BankQuestion.id)
        .where(ExamBankMap.exam_id==exam.id)
    ).all()
    mapped_subjects=sorted({(row[0] or '').strip() for row in mapped if (row[0] or '').strip()},key=str.casefold)
    mapped_units=sorted({(row[1] or '').strip() for row in mapped if (row[1] or '').strip()},key=lambda value:(int(value) if str(value).isdigit() else 10**9,str(value).casefold()))

    # A Super Admin can explicitly repair legacy exam metadata from the Exam
    # List. That intentional override must win over old Question Bank mappings;
    # otherwise a corrected exam can jump back into its previous/General group.
    manual_subject_override=bool(
        cfg and 'manual_subject_override=1' in ((cfg.last_generation_summary or '').lower())
    )

    # For normal Ready Exams the mapped Question Bank remains the strongest
    # source of truth. This protects newer exams from stale ExamConfig metadata.
    if manual_subject_override and configured_subject:
        subject=configured_subject
    elif len(mapped_subjects)==1:
        subject=mapped_subjects[0]
    elif configured_subject:
        subject=configured_subject
    elif len(mapped_subjects)>1:
        subject='Mixed Subjects'
    else:
        subject=''

    units=list(mapped_units)
    if not units and cfg and cfg.unit_weights:
        weights=safe_json_load(cfg.unit_weights,{})
        if isinstance(weights,dict):
            units=[str(key).strip() for key,value in weights.items() if str(key).strip() and value not in (0,'0',None,'')]

    def unit_display(value):
        value=str(value or '').strip()
        if not value:return ''
        return value if value.lower().startswith('unit ') else f'Unit {value}'

    if len(units)==1:
        unit_label=unit_display(units[0])
    elif len(units)>1:
        unit_label='All Units'
    else:
        unit_label='General'

    if not subject:
        subject='General'
    return subject,unit_label

def student_exam_unit_sort_key(label):
    text=(label or '').strip()
    lower=text.lower()
    if lower.startswith('unit '):
        value=text[5:].strip()
        try:return (0,int(value),'')
        except ValueError:return (0,10**9,value.casefold())
    if lower=='all units':return (1,0,'')
    if lower=='general':return (2,0,'')
    return (3,0,lower)

@app.route('/student')
@student_required
def student_dashboard():
    s=DB();st=s.get(Student,web_session['user_id']);exams_list=s.scalars(select(Exam).where(Exam.is_active==True).order_by(Exam.id.desc())).all();rows=[]
    for e in exams_list:
        allowed,access_label,session_row=exam_access_for_student(s,st.id,e)
        if access_label=='Not assigned to your batch/section':continue
        pool_count=s.scalar(select(func.count()).select_from(Question).where(Question.exam_id==e.id)) or 0;cfg=normalize_legacy_manual_subject_exam(s,e.id,get_exam_config(s,e.id));display_count=min(cfg.question_count,pool_count) if cfg and cfg.question_count else pool_count;att=get_attempt(s,st.id,e.id)
        subject,unit_label=student_exam_subject_unit(s,e,cfg);security=get_exam_security_policy(s,e.id,create=False)
        rows.append(type('StudentExamRow',(),{'id':e.id,'title':e.title,'display_title':student_grouped_exam_display_title(s,e,subject),'duration_minutes':e.duration_minutes,'question_count':display_count,'attempt_status':att.status if att else None,'can_start':allowed,'access_label':access_label,'venue':session_row.venue if session_row else '','subject':subject,'unit_label':unit_label,'pin_required':bool(security and security.require_exam_pin)})())

    grouped={}
    for row in rows:
        grouped.setdefault(row.subject,{}).setdefault(row.unit_label,[]).append(row)
    exam_groups=[]
    for subject in sorted(grouped,key=str.casefold):
        units=[]
        for unit_label in sorted(grouped[subject],key=student_exam_unit_sort_key):
            units.append({'label':unit_label,'exams':grouped[subject][unit_label]})
        exam_groups.append({'subject':subject,'units':units})
    return render_template('student_dashboard.html',student=st,exams=rows,exam_groups=exam_groups)


@app.route('/student/practical-code',methods=['GET','POST'])
@student_required
def student_practical_code():
    s=DB();student=s.get(Student,web_session['user_id'])
    if not student:abort(404)
    rows=practical_code_exam_rows_for_student(s,student)
    if not rows:
        flash('Practical Marks are available only when a Practical Exam is assigned.','error')
        return redirect(url_for('student_dashboard'))

    selected_exam_id=request.args.get('exam_id',type=int)
    if request.method=='POST':
        try:exam_id=int(request.form.get('exam_id','0'))
        except (TypeError,ValueError):exam_id=0
        selected=next((row for row in rows if row['exam'].id==exam_id),None)
        if not selected:
            flash('This Practical Exam is not currently available.','error');return redirect(url_for('student_practical_code'))
        if not selected.get('can_submit'):
            flash(selected.get('access_label') or 'This Practical Exam is not open yet.','error');return redirect(url_for('student_practical_code',exam_id=selected['exam'].id))
        target=selected['target'];meta=selected['meta'];exam=selected['exam']
        if not target.get('ok'):
            flash('Your practical register mapping could not be resolved safely. Ask the faculty member to verify your Roll No. and Experiment No.','error')
            return redirect(url_for('student_practical_code',exam_id=exam.id))
        register=target['register'];practical_student=target['practical_student'];experiment=target['experiment']
        reference=(experiment.reference_code or '').strip()
        if not reference:
            flash(f'Reference code has not been configured for Experiment {experiment.experiment_no}.','error')
            return redirect(url_for('student_practical_code',exam_id=exam.id))
        source=(request.form.get('source_code') or '').strip()
        if not source:
            flash('Type your practical code before submitting.','error');return redirect(url_for('student_practical_code',exam_id=exam.id))
        if len(source)>250000:
            flash('Submitted code is too large.','error');return redirect(url_for('student_practical_code',exam_id=exam.id))

        similarity=evaluate_practical_code(reference,source)
        performance_max=float(practical_marks_maxima(register)['performance'])
        base_performance_value=round((similarity*performance_max)*2.0)/2.0
        base_performance_value=max(0.0,min(performance_max,base_performance_value))
        penalty=practical_code_exact_penalty(source,getattr(experiment,'penalty_rules','') or '',1.0)
        penalty_deduction=min(base_performance_value,float(penalty.get('deduction',0.0) or 0.0))
        performance_value=max(0.0,base_performance_value-penalty_deduction)

        # A correct solution may naturally resemble the faculty reference.  Copy
        # prevention therefore compares this submission only with OTHER students
        # in the same practical exam and flags only near-clones at a very high
        # structural similarity threshold.
        peer_submission,peer_similarity=find_practical_code_peer_clone(
            s,student.id,exam.id,experiment.id,source
        )

        submission=s.scalar(select(PracticalCodeSubmission).where(
            PracticalCodeSubmission.student_id==student.id,
            PracticalCodeSubmission.exam_id==exam.id
        ))
        if not submission:
            submission=PracticalCodeSubmission(
                student_id=student.id,exam_id=exam.id,register_id=register.id,
                practical_student_id=practical_student.id,practical_experiment_id=experiment.id,
                experiment_no=experiment.experiment_no,source_code=source,
                similarity_pct=round(similarity*100,2),performance_marks=0.0,submitted_at=now_iso()
            )
            s.add(submission);s.flush()
        else:
            submission.register_id=register.id
            submission.practical_student_id=practical_student.id
            submission.practical_experiment_id=experiment.id
            submission.experiment_no=experiment.experiment_no
            submission.source_code=source
            submission.similarity_pct=round(similarity*100,2)
            submission.submitted_at=now_iso()

        mark=s.scalar(select(PracticalMark).where(
            PracticalMark.practical_student_id==practical_student.id,
            PracticalMark.practical_experiment_id==experiment.id
        ))

        if peer_submission:
            # Do not auto-award Performance marks for a near-clone.  Preserve
            # faculty-entered marks, but remove an earlier automatic Practical
            # Code award from this student and flag both submissions for review.
            peer_student=s.get(Student,peer_submission.student_id)
            peer_roll=(getattr(peer_student,'registration_no','') or getattr(peer_student,'roll_no','') or str(peer_submission.student_id)).strip()
            current_roll=(practical_student.roll_no or student.registration_no or student.roll_no or str(student.id)).strip()
            review_text=f'Near-identical Practical Code detected with Roll No. {peer_roll} ({peer_similarity*100:.1f}% structural similarity).'

            if not mark:
                mark=PracticalMark(
                    register_id=register.id,practical_student_id=practical_student.id,
                    practical_experiment_id=experiment.id,attendance='',attendance_marks=None,
                    record_marks=None,performance_marks=None,viva_marks=None,marks=None,
                    remarks=append_code_review_remark('',review_text),updated_by='Practical Code Review',updated_at=now_iso()
                )
                s.add(mark);s.flush()
            else:
                if mark.updated_by=='Practical Code':
                    mark.performance_marks=None
                    components=[mark.attendance_marks,mark.record_marks,mark.performance_marks,mark.viva_marks]
                    mark.marks=sum(value or 0 for value in components) if any(value is not None for value in components) else None
                mark.remarks=append_code_review_remark(mark.remarks,review_text)
                mark.updated_by='Practical Code Review';mark.updated_at=now_iso()

            peer_mark=s.scalar(select(PracticalMark).where(
                PracticalMark.practical_student_id==peer_submission.practical_student_id,
                PracticalMark.practical_experiment_id==peer_submission.practical_experiment_id
            ))
            if peer_mark:
                peer_review=f'Near-identical Practical Code detected with Roll No. {current_roll} ({peer_similarity*100:.1f}% structural similarity).'
                peer_mark.remarks=append_code_review_remark(peer_mark.remarks,peer_review)
                peer_mark.updated_at=now_iso()

            submission.performance_marks=0.0
            register.updated_at=now_iso()
            audit_event(
                s,'practical_code_peer_clone_flagged','practical_code_submission',submission.id or '',
                f'exam={exam.id}, student={student.id}, roll={current_roll}, experiment={experiment.experiment_no}, '
                f'peer_student={peer_submission.student_id}, peer_roll={peer_roll}, peer_similarity={round(peer_similarity*100,2)}, '
                f'reference_similarity={round(similarity*100,2)}'
            )
            s.commit()
            flash('This code is almost identical to another student submission. Performance marks were not auto-awarded; faculty review is required.','error')
            return redirect(url_for('student_practical_code',exam_id=exam.id))

        if not mark:
            mark=PracticalMark(
                register_id=register.id,practical_student_id=practical_student.id,
                practical_experiment_id=experiment.id,attendance='',attendance_marks=None,
                record_marks=None,performance_marks=performance_value,viva_marks=None,marks=performance_value,
                remarks='',updated_by='Practical Code',updated_at=now_iso()
            )
            s.add(mark);s.flush()
        else:
            mark.performance_marks=performance_value
            mark.remarks=clear_code_review_remark(mark.remarks)
            if (mark.attendance or '').upper()=='A':
                mark.marks=0.0
            else:
                components=[mark.attendance_marks,mark.record_marks,mark.performance_marks,mark.viva_marks]
                mark.marks=sum(value or 0 for value in components) if any(value is not None for value in components) else performance_value
            mark.updated_by='Practical Code';mark.updated_at=now_iso()

        submission.performance_marks=performance_value
        register.updated_at=now_iso()
        audit_event(
            s,'practical_code_auto_evaluated','practical_mark',mark.id or '',
            f'exam={exam.id}, student={student.id}, roll={practical_student.roll_no}, experiment={experiment.experiment_no}, '
            f'similarity={round(similarity*100,2)}, base_performance={base_performance_value}/{performance_max}, '
            f'penalty={penalty_deduction}, penalty_units={penalty.get("units",0)}, performance={performance_value}/{performance_max}, match={target.get("match_basis","")}'
        )
        s.commit()
        # Keep faculty-configured exact-match penalties private. Students only
        # see their final Performance score; deduction details remain available
        # to staff through the audit trail.
        flash(f'Practical code evaluated. Performance marks: {performance_value:g} / {performance_max:g}.')
        return redirect(url_for('student_practical_code',exam_id=exam.id))

    if not selected_exam_id and len(rows)==1:selected_exam_id=rows[0]['exam'].id
    return render_template('practical_code.html',student=student,rows=rows,selected_exam_id=selected_exam_id)

@app.route('/student/exam/<int:exam_id>/current-pin')
@student_required
def current_student_exam_pin(exam_id):
    s=DB();exam=s.scalar(select(Exam).where(Exam.id==exam_id,Exam.is_active==True));student=s.get(Student,web_session['user_id'])
    if not exam or not student:abort(404)
    security=get_exam_security_policy(s,exam_id,create=False)
    if not security or not security.require_exam_pin:abort(404)
    allowed,access_label,_session=exam_access_for_student(s,student.id,exam)
    if not allowed:return jsonify({'error':access_label}),403
    response=jsonify({'pin':rotating_exam_pin(exam_id,student.id),'seconds_remaining':rotating_exam_pin_seconds_remaining()})
    response.headers['Cache-Control']='no-store, no-cache, must-revalidate, max-age=0'
    return response


@app.route('/student/exam/<int:exam_id>/verify-pin',methods=['GET','POST'])
@student_required
def verify_student_exam_pin(exam_id):
    s=DB();exam=s.scalar(select(Exam).where(Exam.id==exam_id,Exam.is_active==True));student=s.get(Student,web_session['user_id'])
    if not exam or not student:abort(404)
    security=get_exam_security_policy(s,exam_id,create=False)
    if not security or not security.require_exam_pin:return redirect(url_for('take_exam',exam_id=exam_id))
    allowed,access_label,_session=exam_access_for_student(s,student.id,exam)
    if not allowed:flash(access_label,'error');return redirect(url_for('student_dashboard'))
    ajax=request.headers.get('X-Requested-With')=='XMLHttpRequest' or request.accept_mimetypes.best=='application/json'
    if request.method=='POST':
        pin=(request.form.get('exam_pin') or '').strip()
        if rotating_exam_pin_matches(exam_id,student.id,pin):
            device_ok,_lock=ensure_exam_device_lock(s,exam_id,student.id)
            if not device_ok:
                s.commit()
                if ajax:return jsonify(ok=False,error='device_locked'),409
                flash('This exam is already locked to another browser/device. Ask the invigilator to reset your device lock.','error');return redirect(url_for('student_dashboard'))
            mark_exam_pin_verified(exam_id);launch_token=create_secure_exam_launch_token(exam_id);audit_event(s,'student_rotating_exam_pin_verified','exam',exam_id,f'student_id={student.id}');s.commit()
            exam_url=url_for('take_exam',exam_id=exam_id,secure_shell=1,launch=launch_token)
            if ajax:return jsonify(ok=True,exam_url=exam_url)
            return redirect(url_for('take_exam',exam_id=exam_id))
        audit_event(s,'student_rotating_exam_pin_failed','exam',exam_id,f'student_id={student.id}, ip={request.remote_addr or ""}');s.commit()
        if ajax:return jsonify(ok=False,error='invalid_pin',pin=rotating_exam_pin(exam_id,student.id),seconds_remaining=rotating_exam_pin_seconds_remaining()),400
        flash('The PIN changed or was entered incorrectly. Use the PIN currently shown on screen.','error')
    return render_template('exam_pin_verify.html',exam=exam,student=student,display_title=student_exam_display_title(s,exam),rotating_pin=rotating_exam_pin(exam_id,student.id),pin_seconds_remaining=rotating_exam_pin_seconds_remaining())


@app.route('/student/exam/<int:exam_id>')
@student_required
def take_exam(exam_id):
    s=DB();exam=s.scalar(select(Exam).where(Exam.id==exam_id,Exam.is_active==True))
    if not exam:flash('Exam is not active.','error');return redirect(url_for('student_dashboard'))
    allowed,access_label,_session=exam_access_for_student(s,web_session['user_id'],exam)
    if not allowed:flash(access_label,'error');return redirect(url_for('student_dashboard'))
    security=get_exam_security_policy(s,exam_id,create=False)
    if security and security.require_exam_pin:
        # PIN-secured exams must never be resumed directly from the dashboard.
        # Only the same-session URL issued after a successful PIN check may
        # render inside the silent secure shell.
        launch_token=(request.args.get('launch') or '').strip()
        if request.args.get('secure_shell')!='1' or not exam_pin_is_verified(exam_id) or not secure_exam_launch_token_valid(exam_id,launch_token):
            return redirect(url_for('verify_student_exam_pin',exam_id=exam_id))
        device_ok,_lock=ensure_exam_device_lock(s,exam_id,web_session['user_id'])
        if not device_ok:
            s.commit();flash('This exam is already locked to another browser/device. Ask the invigilator to reset your device lock.','error');return redirect(url_for('student_dashboard'))
        s.commit()
    cfg=normalize_legacy_manual_subject_exam(s,exam_id,get_exam_config(s,exam_id));attempt=get_attempt(s,web_session['user_id'],exam_id)
    if attempt and attempt.status=='submitted':return redirect(url_for('submitted',exam_id=exam_id))
    if not attempt:
        qids=list(s.scalars(select(Question.id).where(Question.exam_id==exam_id).order_by(Question.id)).all())
        if not qids:flash('This exam has no questions.','error');return redirect(url_for('student_dashboard'))
        target=min((cfg.question_count if cfg and cfg.question_count else len(qids)),len(qids))
        if cfg is None or cfg.randomize_questions:qids=random.sample(qids,target)
        else:qids=qids[:target]
        started=now_dt();end=started+timedelta(minutes=exam.duration_minutes)
        if _session and _session.scheduled_end:
            session_end=datetime.fromisoformat(_session.scheduled_end).astimezone() if datetime.fromisoformat(_session.scheduled_end).tzinfo else datetime.fromisoformat(_session.scheduled_end).replace(tzinfo=started.tzinfo)
            if session_end<end:end=session_end
        attempt=Attempt(student_id=web_session['user_id'],exam_id=exam_id,started_at=started.isoformat(timespec='seconds'),end_at=end.isoformat(timespec='seconds'),status='in_progress',question_order=','.join(map(str,qids)));s.add(attempt);s.flush();s.add(AttemptHeartbeat(attempt_id=attempt.id,last_seen_at=now_iso(),answer_count=0,client_state='active',client_fingerprint=''))
        for pos,qid in enumerate(qids,1):
            keys=list('ABCD')
            if cfg and cfg.shuffle_options:random.shuffle(keys)
            s.add(AttemptQuestion(attempt_id=attempt.id,question_id=qid,position=pos,option_order=''.join(keys)))
        s.commit()
    end_dt=parse_dt(attempt.end_at)
    if now_dt()>=end_dt:finalize_attempt(s,attempt);return redirect(url_for('submitted',exam_id=exam_id))
    aq_rows=s.scalars(select(AttemptQuestion).where(AttemptQuestion.attempt_id==attempt.id).order_by(AttemptQuestion.position)).all()
    if not aq_rows:
        qids=[int(x) for x in attempt.question_order.split(',') if x]
        for pos,qid in enumerate(qids,1):s.add(AttemptQuestion(attempt_id=attempt.id,question_id=qid,position=pos,option_order='ABCD'))
        s.commit();aq_rows=s.scalars(select(AttemptQuestion).where(AttemptQuestion.attempt_id==attempt.id).order_by(AttemptQuestion.position)).all()
    qids=[x.question_id for x in aq_rows];qrows=s.scalars(select(Question).where(Question.id.in_(qids))).all();qmap={q.id:q for q in qrows};views=[]
    for aq in aq_rows:
        q=qmap.get(aq.question_id)
        if not q:continue
        qtype=canonical_question_type(q.question_type);text={'A':q.option_a,'B':q.option_b,'C':q.option_c,'D':q.option_d};order=aq.option_order or 'ABCD';display=[(chr(65+i),key,text[key]) for i,key in enumerate(order)] if qtype in {'single_choice','multiple_select'} else []
        views.append(type('QuestionView',(),{'id':q.id,'question':q.question,'question_type':qtype,'question_type_label':QUESTION_TYPE_LABELS.get(qtype,qtype),'display_options':display,'marks':q.marks})())
    saved=s.scalars(select(Answer).where(Answer.attempt_id==attempt.id)).all();answers={a.question_id:answer_record_value(a) for a in saved}
    security=get_exam_security_policy(s,exam_id,create=False)
    secure_shell=bool(request.args.get('secure_shell')=='1' and security and security.require_exam_pin)
    return render_template('exam.html',exam=exam,display_title=student_exam_display_title(s,exam),questions=views,answers=answers,end_epoch=end_dt.timestamp(),cfg=cfg,security=security,secure_shell=secure_shell)

@app.route('/student/save-answer',methods=['POST'])
@student_required
def save_answer():
    data=request.get_json(silent=True) or {}
    try:exam_id=int(data.get('exam_id'));qid=int(data.get('question_id'))
    except Exception:return jsonify(error='Invalid request'),400
    ans=str(data.get('answer',''))[:MAX_ANSWER_LENGTH]
    s=DB();attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not attempt:return jsonify(error='Attempt not found'),404
    if attempt.status=='submitted':return jsonify(saved=False,submitted=True)
    if not secure_exam_device_allowed(s,attempt):return jsonify(error='Exam is locked to another device.',device_locked=True),409
    if now_dt()>=parse_dt(attempt.end_at):finalize_attempt(s,attempt);return jsonify(saved=False,submitted=True)
    if qid not in attempt_question_ids(s,attempt):return jsonify(error='Question not part of this attempt'),400
    question=s.get(Question,qid)
    try:save_answer_record(s,attempt.id,qid,ans,question)
    except ValueError as exc:return jsonify(error=str(exc)),400
    s.commit();return jsonify(saved=True)

@app.route('/student/integrity-event',methods=['POST'])
@student_required
def integrity_event():
    data=request.get_json(silent=True) or {}
    try:exam_id=int(data.get('exam_id'))
    except Exception:return jsonify(saved=False),400
    event_type=str(data.get('event_type',''))[:50]
    if event_type not in {'tab_hidden','fullscreen_exit','copy_attempt','paste_attempt','context_menu'}:return jsonify(saved=False),400
    s=DB();attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not attempt or attempt.status=='submitted':return jsonify(saved=False),404
    if not secure_exam_device_allowed(s,attempt):return jsonify(saved=False,device_locked=True),409
    s.add(IntegrityEvent(attempt_id=attempt.id,event_type=event_type,details=str(data.get('details',''))[:250],created_at=now_iso()));s.commit();count=s.scalar(select(func.count()).select_from(IntegrityEvent).where(IntegrityEvent.attempt_id==attempt.id)) or 0;return jsonify(saved=True,count=count)

@app.route('/student/heartbeat',methods=['POST'])
@student_required
def student_heartbeat():
    data=request.get_json(silent=True) or {}
    try:exam_id=int(data.get('exam_id'))
    except Exception:return jsonify(saved=False),400
    s=DB();attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not attempt or attempt.status=='submitted':return jsonify(saved=False),404
    if not secure_exam_device_allowed(s,attempt):return jsonify(saved=False,device_locked=True),409
    device_lock=s.scalar(select(ExamDeviceLock).where(ExamDeviceLock.exam_id==exam_id,ExamDeviceLock.student_id==attempt.student_id))
    if device_lock:device_lock.last_seen_at=now_iso()
    count=s.scalar(select(func.count()).select_from(Answer).where(Answer.attempt_id==attempt.id)) or 0
    row=s.scalar(select(AttemptHeartbeat).where(AttemptHeartbeat.attempt_id==attempt.id));fingerprint=hashlib.sha256(((request.headers.get('User-Agent') or '')+'|'+(request.remote_addr or '')).encode('utf-8')).hexdigest()[:24]
    if not row:row=AttemptHeartbeat(attempt_id=attempt.id,last_seen_at=now_iso(),answer_count=int(count),client_state=str(data.get('state') or 'active')[:30],client_fingerprint=fingerprint);s.add(row)
    else:row.last_seen_at=now_iso();row.answer_count=int(count);row.client_state=str(data.get('state') or 'active')[:30];row.client_fingerprint=fingerprint
    s.commit();return jsonify(saved=True,server_time=now_iso())

@app.route('/student/exam/<int:exam_id>/submit',methods=['POST'])
@student_required
def submit_exam(exam_id):
    s=DB();attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not attempt:flash('Attempt not found.','error');return redirect(url_for('student_dashboard'))
    if not secure_exam_device_allowed(s,attempt):flash('This exam is locked to another browser/device. Ask the invigilator to reset your device lock.','error');return redirect(url_for('student_dashboard'))
    allowed=set(attempt_question_ids(s,attempt))
    if attempt.status!='submitted':
        questions={q.id:q for q in (s.scalars(select(Question).where(Question.id.in_(allowed))).all() if allowed else [])}
        for qid,question in questions.items():
            field=f'q_{qid}';values=request.form.getlist(field)
            if not values:continue
            value=','.join(values) if canonical_question_type(question.question_type)=='multiple_select' else values[0]
            try:save_answer_record(s,attempt.id,qid,value[:MAX_ANSWER_LENGTH],question)
            except ValueError:continue
        s.commit();finalize_attempt(s,attempt)
    # The secure iframe launch token is single-session exam state. Once the
    # paper is submitted, discard it before showing the result page.
    clear_secure_exam_launch_token(exam_id)
    return redirect(url_for('submitted',exam_id=exam_id))

@app.route('/student/submitted/<int:exam_id>')
@student_required
def submitted(exam_id):
    s=DB();exam=s.get(Exam,exam_id);attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not exam or not attempt:abort(404)
    if attempt.status!='submitted' and now_dt()>=parse_dt(attempt.end_at):finalize_attempt(s,attempt)
    if attempt.status=='submitted':
        try:
            sync_result=sync_practical_viva_from_attempt(s,attempt)
            if sync_result.get('updated'):s.commit()
        except Exception:
            s.rollback()
            try:app.logger.exception('Practical Exam viva repair failed while showing result for attempt %s',attempt.id)
            except Exception:pass
    violations=s.scalar(select(func.count()).select_from(IntegrityEvent).where(IntegrityEvent.attempt_id==attempt.id)) or 0
    percentage,grade,grade_class=result_performance(attempt.score,attempt.total_marks)
    return render_template('submitted.html',exam=exam,display_title=student_exam_display_title(s,exam),attempt=attempt,violations=violations,percentage=percentage,grade=grade,grade_class=grade_class,answer_review_exam_id=exam.id if attempt.status=='submitted' else None)

@app.route('/student/submitted/<int:exam_id>/answers')
@student_required
def submitted_answers(exam_id):
    s=DB();exam=s.get(Exam,exam_id);attempt=get_attempt(s,web_session['user_id'],exam_id)
    if not exam or not attempt or attempt.status!='submitted':abort(404)
    aq_rows=s.scalars(select(AttemptQuestion).where(AttemptQuestion.attempt_id==attempt.id).order_by(AttemptQuestion.position)).all()
    ordered=[(row.position,row.question_id,row.option_order or 'ABCD') for row in aq_rows] if aq_rows else [(pos,qid,'ABCD') for pos,qid in enumerate(attempt_question_ids(s,attempt),1)]
    qids=[qid for _pos,qid,_order in ordered];qrows=s.scalars(select(Question).where(Question.id.in_(qids))).all() if qids else [];qmap={q.id:q for q in qrows}
    saved=s.scalars(select(Answer).where(Answer.attempt_id==attempt.id)).all();aobj={a.question_id:a for a in saved};amap={a.question_id:answer_record_value(a) for a in saved};views=[]
    for position,qid,option_order in ordered:
        q=qmap.get(qid)
        if not q:continue
        qtype=canonical_question_type(q.question_type);student_value=amap.get(q.id,'');correct_value=normalized_key(q);options=[]
        student_display=student_value or 'Not answered';correct_display=correct_value
        if qtype in {'single_choice','multiple_select'}:
            text_map={'A':q.option_a,'B':q.option_b,'C':q.option_c,'D':q.option_d};order=option_order if len(option_order)==4 and set(option_order)==set('ABCD') else 'ABCD';label_map={key:chr(65+i) for i,key in enumerate(order)};student_set=set(student_value.split(',')) if student_value else set();correct_set=set(correct_value.split(',')) if correct_value else set()
            for i,key in enumerate(order):options.append(type('AnswerOptionView',(),{'label':chr(65+i),'key':key,'text':text_map[key],'is_student':key in student_set,'is_correct':key in correct_set})())
            if student_set:student_display='; '.join(f'{label_map.get(key,key)}. {text_map.get(key,"")}' for key in order if key in student_set)
            correct_display='; '.join(f'{label_map.get(key,key)}. {text_map.get(key,"")}' for key in order if key in correct_set)
        elif qtype=='true_false':
            student_display=student_value.title() if student_value else 'Not answered';correct_display=correct_value.title()
        ans_obj=aobj.get(q.id);is_manual=qtype=='essay';manual_score=(ans_obj.manual_score if ans_obj else None);grader_comment=(ans_obj.grader_comment if ans_obj else '')
        if is_manual:correct_display='Manually assessed by faculty' if manual_score is not None else 'Awaiting manual grading'
        views.append(type('AnswerReviewView',(),{'position':position,'question':q.question,'marks':q.marks,'question_type':qtype,'question_type_label':QUESTION_TYPE_LABELS.get(qtype,qtype),'student_answer':student_display,'correct_answer':correct_display,'is_correct':is_answer_correct(q,student_value),'is_manual':is_manual,'manual_score':manual_score,'grader_comment':grader_comment,'options':options})())
    return render_template('submitted_answers.html',exam=exam,display_title=student_exam_display_title(s,exam),attempt=attempt,questions=views,answer_review_exam_id=exam.id)

@app.errorhandler(400)
def bad_request(e):return render_template('error.html',heading='Invalid or expired request',message=getattr(e,'description','Please refresh the page and try again.')),400
@app.errorhandler(404)
def not_found(e):return render_template('error.html',heading='Page not found',message='The requested exam resource could not be found.'),404

if __name__=='__main__':
    port=int(os.getenv('PORT','8080'))
    try:local_ip=socket.gethostbyname(socket.gethostname())
    except Exception:local_ip='SERVER-IP'
    print('='*64);print(f'LEARN WITH HEMANT — EXAM SYSTEM V{APP_VERSION}');print(f"Mode: {'ONLINE' if APP_MODE=='online' else 'OFFLINE / LAN'}");print(f"Database: {'PostgreSQL' if DATABASE_URL.startswith('postgresql') else 'SQLite'}");print('Server: http://127.0.0.1:%s'%port)
    if APP_MODE=='offline':print(f'LAN URL: http://{local_ip}:{port}')
    print('='*64);app.run(host='0.0.0.0',port=port,debug=False,threaded=True)
