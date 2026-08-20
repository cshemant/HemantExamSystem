"""Helpers for the Practical Marks Register.

These functions intentionally have no Flask/database dependency so roster and
experiment imports can be regression-tested independently of the web app.
"""
from __future__ import annotations

import csv
import io
import re
from difflib import SequenceMatcher
from typing import Iterable

from openpyxl import load_workbook


def cell_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _key(value) -> str:
    text=cell_text(value).casefold().strip()
    text=re.sub(r'[^a-z0-9]+','_',text).strip('_')
    aliases={
        'roll':'roll_no','roll_no':'roll_no','rollno':'roll_no','roll_number':'roll_no',
        'reg_no':'roll_no','reg_number':'roll_no','registration_no':'roll_no','registration_number':'roll_no',
        'enrollment_no':'roll_no','enrolment_no':'roll_no','student_id':'roll_no',
        'name':'name','student_name':'name','name_of_student':'name','candidate_name':'name','full_name':'name',
        'experiment':'experiment_no','experiment_no':'experiment_no','experiment_number':'experiment_no','exp':'experiment_no','exp_no':'experiment_no','no':'experiment_no','s_no':'experiment_no','sno':'experiment_no',
        'title':'title','experiment_name':'title','experiment_title':'title','name_of_experiment':'title','description':'title','practical':'title',
        'marks':'max_marks','max_marks':'max_marks','maximum_marks':'max_marks',
        'code':'reference_code','reference_code':'reference_code','reference_program':'reference_code','source_code':'reference_code','solution_code':'reference_code','experiment_code':'reference_code',
    }
    return aliases.get(text,text)


def _table_rows(filename: str, raw: bytes) -> list[list[object]]:
    filename=(filename or '').lower()
    if not raw:
        raise ValueError('The uploaded file is empty.')
    if filename.endswith('.csv'):
        try:
            text=raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise ValueError('CSV must be UTF-8 encoded.') from exc
        return [list(r) for r in csv.reader(io.StringIO(text))]
    if filename.endswith('.xlsx'):
        try:
            wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
            ws=wb.active
            return [list(r) for r in ws.iter_rows(values_only=True)]
        except Exception as exc:
            raise ValueError('The Excel file could not be read. Please upload a valid .xlsx file.') from exc
    raise ValueError('Please upload a CSV or Excel (.xlsx) file.')


def parse_roster_bytes(filename: str, raw: bytes) -> list[dict]:
    """Read a student roster from a clean table or a university evaluation sheet.

    The supplied university practical sheet keeps ``Reg. No`` and ``Name of
    Student`` on different header rows.  Instead of assuming row 1 is the
    header, this scanner locates the roll/name columns independently within the
    first 15 rows and then reads the student records beneath them.
    """
    rows=_table_rows(filename,raw)
    if not rows:
        raise ValueError('The uploaded file has no rows.')

    # Fast path: clean row-based header (roll_no, name).
    for header_idx,row in enumerate(rows[:15]):
        keys=[_key(v) for v in row]
        if 'roll_no' in keys and 'name' in keys:
            roll_col=keys.index('roll_no');name_col=keys.index('name');start=header_idx+1
            return _collect_roster(rows,start,roll_col,name_col)

    # Flexible university-sheet path: roll/name labels can live on separate rows.
    roll_hit=name_hit=None
    for r_idx,row in enumerate(rows[:15]):
        for c_idx,value in enumerate(row):
            key=_key(value)
            if key=='roll_no' and roll_hit is None:
                roll_hit=(r_idx,c_idx)
            if key=='name' and name_hit is None:
                name_hit=(r_idx,c_idx)
    if not roll_hit or not name_hit:
        raise ValueError('Could not find student columns. The sheet must contain a registration/roll number and student name column.')
    return _collect_roster(rows,max(roll_hit[0],name_hit[0])+1,roll_hit[1],name_hit[1])


def _collect_roster(rows: list[list[object]], start: int, roll_col: int, name_col: int) -> list[dict]:
    out=[];seen=set();blank_run=0
    for row in rows[start:]:
        roll=cell_text(row[roll_col] if roll_col<len(row) else '')
        name=cell_text(row[name_col] if name_col<len(row) else '')
        if not roll and not name:
            blank_run+=1
            if blank_run>=8 and out:
                break
            continue
        blank_run=0
        # Skip footer/summary lines and partially populated records.
        if not roll or not name:
            continue
        low=(roll+' '+name).casefold()
        if any(token in low for token in ('total marks','signature','faculty name','average marks')):
            continue
        key=roll.casefold()
        if key in seen:
            continue
        seen.add(key);out.append({'roll_no':roll,'name':name})
    if not out:
        raise ValueError('No student records were found below the detected roll number/name columns.')
    return out


def normalize_experiment_code(value: str, fallback: int) -> str:
    """Return a stable practical number such as ``12-A``.

    Common source formats such as ``12A``, ``12 A``, ``12-A`` and ``12/A``
    are rendered consistently.  Letter-only continuation cells are handled by
    the sequence-aware import logic below rather than being treated as a full
    experiment number.
    """
    raw=cell_text(value).upper().replace('EXPERIMENT','').replace('EXP.','').replace('EXP','').strip(' :-_')
    raw=re.sub(r'\s+','',raw)
    if not raw:
        return str(fallback)
    match=re.fullmatch(r'(\d+)[\-_/]?([A-Z])',raw)
    if match:
        return f"{int(match.group(1))}-{match.group(2)}"
    if re.fullmatch(r'\d+',raw):
        return str(int(raw))
    return raw[:24]


def normalize_experiment_sequence(codes: Iterable[str]) -> list[str]:
    """Repair/display experiment codes using their ordered sequence.

    Older builds could store continuation rows as ``B``, ``B-5`` or ``C-4``.
    Once a numeric major practical is known, those legacy labels can be safely
    interpreted as its B/C sub-parts.
    """
    out=[];current_major=None
    for pos,value in enumerate(codes,start=1):
        raw=cell_text(value).upper().strip()
        canonical=normalize_experiment_code(raw,pos)
        major_match=re.fullmatch(r'(\d+)(?:-([A-Z]))?',canonical)
        if major_match:
            current_major=str(int(major_match.group(1)))
            out.append(canonical)
            continue
        continuation=re.fullmatch(r'([A-Z])(?:-\d+)?',raw)
        if continuation and current_major:
            out.append(f'{current_major}-{continuation.group(1)}')
        else:
            out.append(canonical)
    return out


def parse_experiment_text(text: str, default_marks: int=10) -> list[dict]:
    """Parse one experiment per line.

    Accepted examples: ``2-A | Linear Layout``, ``2-A - Linear Layout`` or simply
    ``Linear Layout``.  Numbered lists pasted from Word are also accepted.
    """
    rows=[]
    for idx,line in enumerate((text or '').splitlines(),start=1):
        line=line.strip()
        if not line:
            continue
        code='';title=line
        if '|' in line:
            code,title=[x.strip() for x in line.split('|',1)]
        else:
            match=re.match(r'^\s*([0-9]+(?:\s*[-_/]?\s*[A-Za-z])?)\s*[.):-]?\s+(.+)$',line)
            if match:
                code=match.group(1);title=match.group(2).strip()
        if not title:
            continue
        rows.append({'experiment_no':normalize_experiment_code(code,idx),'title':title,'max_marks':max(1,int(default_marks)),'reference_code':''})
    return _dedupe_experiments(rows)


def parse_experiment_bytes(filename: str, raw: bytes, default_marks: int=10) -> list[dict]:
    rows=_table_rows(filename,raw)
    if not rows:
        raise ValueError('The uploaded experiment file has no rows.')

    # Structured header path.
    for header_idx,row in enumerate(rows[:10]):
        keys=[_key(v) for v in row]
        if 'title' in keys:
            title_col=keys.index('title')
            code_col=keys.index('experiment_no') if 'experiment_no' in keys else None
            marks_col=keys.index('max_marks') if 'max_marks' in keys else None
            reference_col=keys.index('reference_code') if 'reference_code' in keys else None
            output=[]
            for pos,data in enumerate(rows[header_idx+1:],start=1):
                title=cell_text(data[title_col] if title_col<len(data) else '')
                if not title:
                    continue
                code=cell_text(data[code_col] if code_col is not None and code_col<len(data) else '')
                marks=cell_text(data[marks_col] if marks_col is not None and marks_col<len(data) else '')
                reference_code=cell_text(data[reference_col] if reference_col is not None and reference_col<len(data) else '')
                try:marks_value=max(1,int(float(marks))) if marks else max(1,int(default_marks))
                except ValueError:marks_value=max(1,int(default_marks))
                output.append({'experiment_no':normalize_experiment_code(code,pos),'title':title,'max_marks':marks_value,'reference_code':reference_code})
            output=_dedupe_experiments(output)
            if output:return output

    # Loose university/list path.  Many lab sheets use one column for the
    # major practical number and another for A/B/C sub-parts.  The major number
    # is commonly shown only on the first row of the group (Excel merged-cell
    # style), so carry it forward for following letter-only rows.
    output=[];current_major=None
    for pos,row in enumerate(rows,start=1):
        cells=[cell_text(v) for v in row]
        values=[v for v in cells if v]
        if not values:
            continue
        if len(values)==1:
            if _key(values[0]) in {'experiment','experiment_no','title','experiment_list'}:
                continue
            code='';title=values[0]
        else:
            title=max(values,key=len)
            # Only inspect non-title cells for numbering tokens.
            tokens=[v.strip().upper() for v in values if v != title]
            combined=next((v for v in tokens if re.fullmatch(r'\d+\s*[-_/]?\s*[A-Z]',v)),None)
            major=next((v for v in tokens if re.fullmatch(r'\d+',v)),None)
            letter=next((v for v in tokens if re.fullmatch(r'[A-Z]',v)),None)
            if combined:
                code=combined
                m=re.match(r'(\d+)',combined);current_major=str(int(m.group(1))) if m else current_major
            elif major:
                current_major=str(int(major))
                code=f'{current_major}-{letter}' if letter else current_major
            elif letter and current_major:
                code=f'{current_major}-{letter}'
            else:
                code=''
            if title==code:
                continue
        if len(title)<4:
            continue
        output.append({'experiment_no':normalize_experiment_code(code,pos),'title':title,'max_marks':max(1,int(default_marks)),'reference_code':''})
    output=_dedupe_experiments(output)
    if not output:
        raise ValueError('No experiments could be detected. Use the template or paste one experiment per line.')
    return output


def _dedupe_experiments(rows: Iterable[dict]) -> list[dict]:
    out=[];used_codes=set();used_titles=set()
    for pos,row in enumerate(rows,start=1):
        title=cell_text(row.get('title'))
        if not title:
            continue
        title_key=title.casefold()
        if title_key in used_titles:
            continue
        base=normalize_experiment_code(row.get('experiment_no',''),pos)
        code=base;counter=2
        while code.casefold() in used_codes:
            code=f'{base}-{counter}';counter+=1
        used_codes.add(code.casefold());used_titles.add(title_key)
        try:marks=max(1,int(row.get('max_marks') or 10))
        except (TypeError,ValueError):marks=10
        out.append({'experiment_no':code,'title':title,'max_marks':marks,'reference_code':cell_text(row.get('reference_code'))})
    return out

# ---------------------------------------------------------------------------
# Practical file first-page OCR matching
# ---------------------------------------------------------------------------

def normalize_scan_identifier(value: str) -> str:
    """Normalize a roll/registration number for OCR-safe comparison."""
    return re.sub(r'[^A-Z0-9]+','',cell_text(value).upper())


def _digit_like_identifier(value: str) -> str:
    """Repair common OCR substitutions only when an identifier is digit-heavy."""
    value=normalize_scan_identifier(value)
    if not value:
        return ''
    digit_count=sum(ch.isdigit() for ch in value)
    if digit_count < max(3,len(value)//2):
        return value
    table=str.maketrans({'O':'0','Q':'0','I':'1','L':'1','S':'5','B':'8','Z':'2'})
    return value.translate(table)


def normalize_scan_name(value: str) -> str:
    text=re.sub(r'[^A-Z ]+',' ',cell_text(value).upper())
    return re.sub(r'\s+',' ',text).strip()


def extract_practical_scan_fields(ocr_text: str) -> dict:
    """Extract likely labelled roll/registration number and name from OCR text."""
    text=(ocr_text or '').replace('\r','\n')
    lines=[re.sub(r'\s+',' ',line).strip() for line in text.split('\n') if line.strip()]
    roll='';name=''
    roll_re=re.compile(r'(?i)\b(?:roll|reg(?:istration)?|enrol(?:lment)?|student\s*id)\s*(?:no\.?|number|#)?\s*[:\-]?\s*([A-Z0-9][A-Z0-9 ./_\-]{3,30})')
    name_re=re.compile(r'(?i)\b(?:student\s*)?name\s*[:\-]?\s*([A-Z][A-Z .]{2,60})')
    for line in lines:
        if not roll:
            m=roll_re.search(line)
            if m:
                candidate=m.group(1).strip(' .:-_')
                # Avoid swallowing a following label such as "Name".
                candidate=re.split(r'(?i)\b(?:name|class|section|semester|branch)\b',candidate,maxsplit=1)[0].strip()
                roll=candidate
        if not name:
            m=name_re.search(line)
            if m:
                candidate=m.group(1).strip(' .:-_')
                candidate=re.split(r'(?i)\b(?:roll|reg(?:istration)?|class|section|semester|branch|date)\b',candidate,maxsplit=1)[0].strip()
                name=candidate
    return {'roll_no':roll,'name':name}


def practical_scan_candidates(ocr_text: str, students: Iterable, limit: int=5) -> list[dict]:
    """Rank practical-register students against OCR text.

    Roll/registration identifiers drive automatic matching. Name similarity is
    deliberately only a supporting signal so a common name cannot silently mark
    the wrong student's file as received.
    """
    text=ocr_text or ''
    fields=extract_practical_scan_fields(text)
    compact_text=normalize_scan_identifier(text)
    raw_tokens=re.findall(r'[A-Za-z0-9][A-Za-z0-9./_\-]{3,30}',text)
    token_ids={normalize_scan_identifier(x) for x in raw_tokens}
    digit_tokens={_digit_like_identifier(x) for x in raw_tokens}
    detected_roll=normalize_scan_identifier(fields.get('roll_no',''))
    detected_roll_digit=_digit_like_identifier(fields.get('roll_no',''))
    detected_name=normalize_scan_name(fields.get('name',''))
    lines=[normalize_scan_name(x) for x in text.splitlines() if normalize_scan_name(x)]
    ranked=[]
    for st in students:
        roll=normalize_scan_identifier(getattr(st,'roll_no',''))
        roll_digit=_digit_like_identifier(getattr(st,'roll_no',''))
        name=normalize_scan_name(getattr(st,'name',''))
        roll_strength=0.0;reason=''
        if roll:
            if detected_roll and detected_roll==roll:
                roll_strength=1.0;reason='labelled roll number matched'
            elif detected_roll_digit and detected_roll_digit==roll_digit:
                roll_strength=.98;reason='OCR-corrected roll number matched'
            elif detected_roll and len(roll)>=5 and (detected_roll.endswith(roll) or roll.endswith(detected_roll)):
                roll_strength=.94;reason='roll number suffix matched'
            elif roll in token_ids:
                roll_strength=.96;reason='roll number found on page'
            elif roll_digit in digit_tokens:
                roll_strength=.93;reason='OCR-corrected roll number found on page'
            elif len(roll)>=5 and roll in compact_text:
                roll_strength=.90;reason='roll number found in OCR text'
            elif detected_roll:
                ratio=SequenceMatcher(None,roll_digit,detected_roll_digit).ratio()
                if ratio>=.86:
                    roll_strength=.72*ratio;reason='similar roll number detected'
        name_strength=0.0
        if name:
            if detected_name:
                name_strength=SequenceMatcher(None,name,detected_name).ratio()
            elif lines:
                name_strength=max(SequenceMatcher(None,name,line).ratio() for line in lines)
        # Identifier contributes up to 90 points; name can add at most 10.
        score=min(100.0,roll_strength*90.0+name_strength*10.0)
        ranked.append({
            'student_id':getattr(st,'id',None),'roll_no':getattr(st,'roll_no',''),
            'name':getattr(st,'name',''),'score':round(score,1),
            'roll_strength':round(roll_strength,3),'name_strength':round(name_strength,3),
            'reason':reason or ('name similarity' if name_strength>.6 else 'weak match'),
        })
    ranked.sort(key=lambda x:(x['score'],x['roll_strength'],x['name_strength']),reverse=True)
    return ranked[:max(1,int(limit))]


def practical_scan_auto_match(ocr_text: str, students: Iterable) -> dict:
    """Return fields, candidates and a safe high-confidence automatic match."""
    fields=extract_practical_scan_fields(ocr_text)
    ranked=practical_scan_candidates(ocr_text,students,limit=5)
    top=ranked[0] if ranked else None
    second=ranked[1] if len(ranked)>1 else None
    # Require a strong identifier match and a useful separation from runner-up.
    automatic=bool(top and top['score']>=84 and top['roll_strength']>=.90 and (not second or top['score']-second['score']>=6))
    return {'fields':fields,'candidates':ranked,'match':top if automatic else None}

